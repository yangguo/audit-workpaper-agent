import json

import openpyxl
import pytest
from langchain_core.messages import AIMessage

from review.findings_review import _llm_review_findings
from review.models import Finding


class _FakeRunnable:
    def __init__(self, content_or_exc):
        self.content_or_exc = content_or_exc
        self.calls = 0

    async def ainvoke(self, messages):
        self.calls += 1
        if isinstance(self.content_or_exc, BaseException):
            raise self.content_or_exc
        return AIMessage(content=self.content_or_exc)


class _FakeLLM:
    def __init__(self, content_or_exc):
        self.content_or_exc = content_or_exc

    def bind(self, **kwargs):
        return _FakeRunnable(self.content_or_exc)


def _make_finding(**kw):
    base = dict(
        issue_type="程序执行不到位/仅依赖访谈", severity="P1", sheet="Sheet",
        cell="B5", snippet="我们通过访谈了解。", basis="依据", suggestion="建议",
    )
    base.update(kw)
    return Finding(**base)


@pytest.mark.asyncio
async def test_llm_review_findings_returns_structured_result(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    finding = _make_finding()
    payload = json.dumps({
        "results": [{
            "id": 1,
            "status": "fail",
            "llm_validity": "成立",
            "severity": "P1",
            "conclusion": "结论文字至少四个字",
            "reasons": ["理由一"],
            "evidence_refs": [{"cell_or_range": "B5", "excerpt": "访谈"}],
            "risk_type": "证据不足",
            "fix_suggestion": {"supplement_explanation": "补充截图"},
        }]
    }, ensure_ascii=False)
    llm = _FakeLLM(payload)

    results = await _llm_review_findings(wb, [finding], llm, batch_size=6, sleep_seconds=0)

    assert 1 in results
    assert results[1]["llm_status"] == "fail"
    assert results[1]["llm_severity_p"] == "P1"
    assert results[1]["llm_conclusion"] == "结论文字至少四个字"


@pytest.mark.asyncio
async def test_llm_review_findings_degrades_to_unknown_on_failure(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    finding = _make_finding()
    llm = _FakeLLM(RuntimeError("timed out"))

    results = await _llm_review_findings(wb, [finding], llm, batch_size=6, sleep_seconds=0)

    assert results[1]["llm_status"] == "unknown"
    assert results[1]["llm_unknown_reason"]


@pytest.mark.asyncio
async def test_llm_review_findings_skips_llm_tagged_findings(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    finding = _make_finding(issue_type="LLM判定：检查要点存在问题")
    llm = _FakeLLM('{"results": []}')

    results = await _llm_review_findings(wb, [finding], llm, batch_size=6, sleep_seconds=0)

    assert results == {}
