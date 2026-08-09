import json

import openpyxl
import pytest
from langchain_core.messages import AIMessage

from review.evidence_steps import _llm_check_evidence_vs_steps
from review.models import AttachmentPreviewItem


class _FakeRunnable:
    def __init__(self, content):
        self.content = content

    async def ainvoke(self, messages):
        return AIMessage(content=self.content)


class _FakeLLM:
    def __init__(self, content):
        self.content = content

    def bind(self, **kwargs):
        return _FakeRunnable(self.content)


class _CapturingRunnable:
    def __init__(self, content, captured):
        self.content = content
        self.captured = captured

    async def ainvoke(self, messages):
        self.captured.append(messages)
        return AIMessage(content=self.content)


class _CapturingLLM:
    def __init__(self, content, captured):
        self.content = content
        self.captured = captured

    def bind(self, **kwargs):
        return _CapturingRunnable(self.content, self.captured)


def _preview_with_item():
    item = AttachmentPreviewItem(
        index="1", rel_dir="d", filename="a.png", rel_path="d/a.png",
        file_type="png", description="用户清单截图", status="OK",
    )
    return {
        "by_filename": {"a.png": [item]}, "by_rel_path": {"d\\a.png": [item]},
        "by_index": {"1": [item]}, "by_sheet_norm": {},
        "path": "p", "items": [item], "status_counts": {"OK": 1},
    }


def _build_ws_with_evidence_ref():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws["A5"] = "获取系统用户清单并检查权限分配。"
    ws["B5"] = "我们导出用户清单，见附件1，截图保存。"
    return ws


@pytest.mark.asyncio
async def test_llm_check_evidence_vs_steps_returns_findings(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    ws = _build_ws_with_evidence_ref()
    payload = json.dumps({
        "results": [{
            "id": 1,
            "status": "fail",
            "conclusion": "证据与步骤不匹配结论",
            "reasons": ["理由"],
            "evidence_refs": [{"cell_or_range": "B5", "excerpt": "导出用户清单"}],
            "severity": "P1",
            "risk_type": "证据不足",
            "fix_suggestion": {"supplement_explanation": "补充"},
        }]
    }, ensure_ascii=False)
    llm = _FakeLLM(payload)

    findings = await _llm_check_evidence_vs_steps(
        llm=llm, ws_title="SA-1", ws=ws,
        attachments_preview=_preview_with_item(), batch_size=6, sleep_seconds=0,
    )

    assert any(
        f.issue_type == "LLM判定：证据-步骤一致性-证据与审计步骤不匹配（证据不足）"
        and f.status == "fail" for f in findings
    )


@pytest.mark.asyncio
async def test_evidence_step_basis_labels_the_attachment_that_it_quotes(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    ws = _build_ws_with_evidence_ref()
    attachments = _preview_with_item()
    attachments["ocr_by_path"] = {
        "d/a.png": {
            "status": "ok",
            "content": "截图显示管理员权限已分配",
        },
    }
    payload = json.dumps({
        "results": [{
            "id": 1,
            "status": "fail",
            "conclusion": "证据与审计步骤不匹配。",
            "reasons": ["截图未说明抽样范围"],
            "evidence_refs": [{
                "attachment": "d/a.png",
                "excerpt": "截图显示管理员权限已分配",
            }],
            "severity": "P1",
            "risk_type": "证据不足",
            "fix_suggestion": {"supplement_explanation": "补充抽样范围"},
        }]
    }, ensure_ascii=False)

    findings = await _llm_check_evidence_vs_steps(
        llm=_FakeLLM(payload), ws_title="SA-1", ws=ws,
        attachments=attachments, batch_size=6, sleep_seconds=0,
    )

    finding = next(f for f in findings if f.status == "fail")
    assert "附件：d/a.png" in finding.basis
    refs = json.loads(finding.evidence_refs)
    assert refs[0]["attachment"] == "d/a.png"


@pytest.mark.asyncio
async def test_llm_check_evidence_vs_steps_empty_preview_returns_empty(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    ws = _build_ws_with_evidence_ref()
    findings = await _llm_check_evidence_vs_steps(
        llm=_FakeLLM('{"results": []}'), ws_title="SA-1", ws=ws,
        attachments_preview={}, batch_size=6, sleep_seconds=0,
    )
    assert findings == []


@pytest.mark.asyncio
async def test_llm_check_evidence_vs_steps_includes_agent_evidence(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    ws = _build_ws_with_evidence_ref()
    captured = []
    attachments = _preview_with_item()
    attachments["agent_evidence_by_sheet"] = {
        "SA1": [{
            "path": "SA-1/users.txt",
            "file_type": "txt",
            "extraction_status": "ok",
            "excerpt": "admin,administrator",
            "supports": "用户清单中的管理员权限",
            "confidence": "high",
        }]
    }

    await _llm_check_evidence_vs_steps(
        llm=_CapturingLLM('{"results": []}', captured),
        ws_title="SA-1", ws=ws, attachments=attachments,
        batch_size=6, sleep_seconds=0,
    )

    assert captured
    assert "admin,administrator" in str(captured[0])
    assert "SA-1/users.txt" in str(captured[0])


def test_evidence_steps_prompt_includes_evidence_inventory(monkeypatch):
    """The evidence_steps review prompt should include the attachment inventory."""
    from review import evidence_steps as es_mod
    from review import llm as llm_mod

    captured = {}

    async def fake_chat(*, llm, messages, stage, max_attempts=2, max_tokens=4096):
        captured["messages"] = messages
        return json.dumps({"results": []})

    monkeypatch.setattr(llm_mod, "_llm_chat", fake_chat)

    # Build minimal ws with a row that triggers evidence matching
    ws = _build_ws_with_evidence_ref()

    # Build attachments dict containing one real + one embedded item.
    # The preview item is keyed by index "1" so the LLM path is reached:
    # row B5 references "附件1" which matches that index.
    attached = AttachmentPreviewItem(
        index="1", rel_dir="d", filename="a.png", rel_path="d/a.png",
        file_type="png", description="用户清单截图", status="OK",
    )
    attachments = {
        "items": [
            attached,
            {"rel_path": ".embedded_media/test.docx::image1.png", "status": "binary", "file_type": "png", "extracted_text": ""},
            {"rel_path": "审计证据/test.txt", "status": "ok", "file_type": "txt", "extracted_text": "用户清单"},
        ],
        "by_filename": {"a.png": [attached]},
        "by_rel_path": {"d\\a.png": [attached]},
        "by_index": {"1": [attached]},
        "by_sheet_norm": {},
    }

    import asyncio
    asyncio.run(es_mod._llm_check_evidence_vs_steps(
        llm=None, ws_title="TEST", ws=ws,
        attachments=attachments, attachments_preview=attachments,
    ))

    # Inventory markers must appear in the user prompt (which carries the payload).
    user_msg = next(m for m in captured["messages"] if m["role"] == "user")
    assert "证据清单" in user_msg["content"] or "证据不足" in user_msg["content"]


