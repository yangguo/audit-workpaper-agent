import json
import re

import openpyxl
import pytest
from langchain_core.messages import AIMessage

from review.pipeline import (
    _backfill_embedded_evidence_refs,
    _finding_to_dict,
    _model_re_review_gate,
    _parse_sheet_filter,
    run_review,
)
from review.finding_taxonomy import default_assertion_catalog
from review.quality_gates import build_quality_gate_context
from review.models import AttachmentFile, Finding


def test_backfill_embedded_evidence_ref_uses_validated_agent_excerpt():
    logical_path = ".embedded_media/sap密码策略.docx::image1.png"
    item = AttachmentFile(
        index="",
        rel_dir=".embedded_media",
        filename="sap密码策略.docx::image1.png",
        rel_path=logical_path,
        file_type="png",
        status="binary",
        extraction_status="binary",
        extracted_text="",
    )
    attachments = {
        "items": [item],
        "by_filename": {item.filename.lower(): [item]},
        "by_rel_path": {logical_path.lower(): [item]},
        "by_index": {},
        "by_sheet_norm": {},
        "ocr_by_path": {
            logical_path.lower(): {
                "status": "ok",
                "content": "密码最小长度为 12 个字符",
            },
        },
        "agent_evidence_by_sheet": {
            "SA10": [{
                "path": logical_path,
                "file_type": "png",
                "extraction_status": "ocr",
                "excerpt": "密码最小长度为 12 个字符",
                "supports": "密码策略参数",
            }],
        },
    }
    finding = {
        "sheet": "SA-10",
        "basis": f"截图 {logical_path} 显示密码最小长度为 12 个字符。",
        "evidence_refs": [],
    }
    out = _backfill_embedded_evidence_refs([finding], attachments)
    assert out[0]["evidence_refs"] == [{
        "sheet": "SA-10",
        "cell_or_range": "",
        "attachment": logical_path,
        "excerpt": "密码最小长度为 12 个字符",
        "full_text": "密码最小长度为 12 个字符",
    }]


def test_backfill_requires_the_exact_path_in_basis_even_for_agent_evidence():
    logical_path = ".embedded_media/policy.docx::image1.png"
    item = AttachmentFile(
        index="",
        rel_dir=".embedded_media",
        filename="policy.docx::image1.png",
        rel_path=logical_path,
        file_type="png",
        status="binary",
        extraction_status="binary",
        extracted_text="",
    )
    attachments = {
        "items": [item],
        "by_filename": {item.filename.lower(): [item]},
        "by_rel_path": {logical_path.lower(): [item]},
        "by_index": {},
        "by_sheet_norm": {},
        "ocr_by_path": {
            logical_path.lower(): {"status": "ok", "content": "密码最小长度为 12 个字符"},
        },
        "agent_evidence_by_sheet": {
            "SA10": [{"path": logical_path, "excerpt": "密码最小长度为 12 个字符"}],
        },
    }
    finding = {
        "sheet": "SA-10",
        "basis": "检查《policy》后认为密码策略需要复核。",
        "evidence_refs": [],
    }

    out = _backfill_embedded_evidence_refs([finding], attachments)

    assert out[0]["evidence_refs"] == []


def test_backfill_embedded_evidence_refs_no_duplicate():
    finding = {
        "sheet": "SA-10",
        "basis": ".embedded_media/foo.docx::image1.png 已有。",
        "evidence_refs": [
            {"sheet": "SA-10", "cell_or_range": "", "attachment": ".embedded_media/foo.docx::image1.png", "excerpt": "x"}
        ],
    }
    out = _backfill_embedded_evidence_refs([finding])
    assert len(out[0]["evidence_refs"]) == 1


def test_backfill_does_not_add_an_uncited_real_attachment():
    finding = {
        "sheet": "PE-6",
        "basis": "见 审计证据/PE-6/C10-演练记录.pdf（含截图）。",
        "evidence_refs": [],
    }
    out = _backfill_embedded_evidence_refs([finding])
    assert out[0]["evidence_refs"] == []


def test_backfill_does_not_turn_a_document_name_into_all_embedded_images():
    """A document mention alone is not a verified citation to each image."""
    attachments = {
        "items": [
            {
                "rel_path": ".embedded_media/doca.docx::image1.png",
                "status": "binary",
                "file_type": "png",
            },
            {
                "rel_path": ".embedded_media/doca.docx::image2.png",
                "status": "binary",
                "file_type": "png",
            },
        ]
    }
    finding = {
        "sheet": "SA-10",
        "basis": "底稿仅引用《doca》文档，未提供截图或导出文件作为直接证据。",
        "evidence_refs": [],
    }
    out = _backfill_embedded_evidence_refs([finding], attachments)
    assert out[0]["evidence_refs"] == []


def test_backfill_does_not_infer_an_attachment_from_a_check_verb():
    """A document mention is not a verified citation to an embedded image."""
    attachments = {
        "items": [
            {"rel_path": ".embedded_media/doca.docx::image1.png", "status": "binary", "file_type": "png"},
        ]
    }
    finding = {
        "sheet": "SA-10",
        "basis": "通过检查《doca》，我们获取了系统密码策略。",
        "evidence_refs": [],
    }
    out = _backfill_embedded_evidence_refs([finding], attachments)
    assert out[0]["evidence_refs"] == []


def test_backfill_embedded_evidence_refs_no_match_leaves_attachment_empty():
    """Document name that doesn't match any attachment should not add refs."""
    attachments = {
        "items": [
            {"rel_path": ".embedded_media/foo.docx::image1.png", "status": "binary", "file_type": "png"},
        ]
    }
    finding = {
        "sheet": "SA-10",
        "basis": "引用了《完全不存在的文档》.",
        "evidence_refs": [],
    }
    out = _backfill_embedded_evidence_refs([finding], attachments)
    assert out[0]["evidence_refs"] == []


def test_backfill_does_not_resolve_a_generic_title_via_token_overlap():
    """Fuzzy document-name matching cannot create an auditable citation."""
    attachments = {
        "items": [
            {"rel_path": ".embedded_media/sap应用系统密码策略.docx::image1.png", "status": "binary", "file_type": "png"},
            {"rel_path": ".embedded_media/sap系统数据库密码策略.docx::image1.png", "status": "binary", "file_type": "png"},
            {"rel_path": ".embedded_media/操作系统密码策略.docx::image1.png", "status": "binary", "file_type": "png"},
        ]
    }
    finding = {
        "sheet": "SA-10",
        "basis": "通过检查《SAP系统密码策略》，我们获取了系统密码策略。",
        "evidence_refs": [],
    }
    out = _backfill_embedded_evidence_refs([finding], attachments)
    assert out[0]["evidence_refs"] == []


def test_backfill_does_not_use_a_snippet_to_infer_an_attachment():
    """Only an explicit basis path plus validated Agent excerpt may backfill."""
    attachments = {
        "items": [
            {"rel_path": ".embedded_media/sap应用系统密码策略.docx::image1.png", "status": "binary", "file_type": "png"},
        ]
    }
    finding = {
        "sheet": "SA-10",
        "cell": "C14",
        "basis": "标准审计程序要求获取/检查证据，但执行描述未体现对应证据。",
        "snippet": "1.在系统管理员协助下，通过检查《SAP系统密码策略》<C22.SA-10-1>，我们获取了系统密码策略：...",
        "evidence_refs": [],
    }
    out = _backfill_embedded_evidence_refs([finding], attachments)
    assert out[0]["evidence_refs"] == []


def test_backfill_does_not_use_unverified_llm_evidence_refs():
    """An LLM-supplied document title cannot substitute for verified evidence."""
    attachments = {
        "items": [
            {"rel_path": ".embedded_media/sap应用系统密码策略.docx::image1.png", "status": "binary", "file_type": "png"},
        ]
    }
    finding = {
        "sheet": "SA-10",
        "basis": "标准审计程序要求获取/检查证据，但执行描述未体现对应证据。",
        "snippet": "",
        "llm_evidence_refs": json.dumps([{
            "sheet": "SA-10",
            "cell_or_range": "C14",
            "attachment": "",
            "excerpt": "通过检查《SAP系统密码策略》获取系统密码策略",
        }], ensure_ascii=False),
        "evidence_refs": [],
    }
    out = _backfill_embedded_evidence_refs([finding], attachments)
    assert out[0]["evidence_refs"] == []


def test_backfill_does_not_mint_a_citation_from_ocr_cache_alone():
    """Cached OCR must still be linked by a validated Agent evidence record."""
    attachments = {
        "items": [
            {"rel_path": ".embedded_media/sap应用系统密码策略.docx::image1.png", "status": "binary", "file_type": "png"},
        ],
        "ocr_by_path": {
            ".embedded_media/sap应用系统密码策略.docx::image1.png": {
                "status": "ok",
                "provider": "mineru-precise",
                "content": "<table><tr><td>min_password_lng</td><td>6</td></tr>"
                           "<tr><td>min_password_specials</td><td>0</td></tr></table>",
            },
        },
    }
    finding = {
        "sheet": "SA-10",
        "basis": "未发现《sap应用系统密码策略》截图中的参数设置",
        "snippet": "",
        "evidence_refs": [],
    }
    out = _backfill_embedded_evidence_refs([finding], attachments)
    assert out[0]["evidence_refs"] == []


class _FakeRunnable:
    def __init__(self, content):
        self.content = content

    async def ainvoke(self, messages):
        return AIMessage(content=self.content)


class _FakeLLM:
    """Returns a canned review result for any LLM call (used by _llm_review_findings)."""
    def __init__(self, content):
        self.content = content

    def bind(self, **kwargs):
        return _FakeRunnable(self.content)


def _pass_review_payload():
    return json.dumps({
        "results": [{
            "id": 1,
            "status": "pass",
            "conclusion": "该发现经复核不成立结论",
            "evidence_refs": [],
            "llm_validity": "不成立",
            "llm_severity": "低",
            "severity": "P2",
            "reasons": ["理由"],
            "risk_type": "证据不足",
        }]
    }, ensure_ascii=False)


def test_parse_sheet_filter():
    assert _parse_sheet_filter(None) is None
    assert _parse_sheet_filter("all") is None
    assert _parse_sheet_filter("SA-4c,SA-5") == ["SA-4c", "SA-5"]
    assert _parse_sheet_filter("SA-4c，SA-5") == ["SA-4c", "SA-5"]


def test_finding_to_dict_parses_json_fields():
    f = Finding(
        issue_type="t", severity="P0", sheet="SA-1", cell="A1",
        snippet="s", basis="b", suggestion="sug",
        evidence_refs='[{"cell_or_range":"A1"}]', reasons='["理由"]',
        fix_suggestion_detail='{"supplement_explanation":"补充"}',
    )
    d = _finding_to_dict(f)
    assert d["evidence_refs"] == [{"cell_or_range": "A1"}]
    assert d["reasons"] == ["理由"]
    assert d["fix_suggestion_detail"] == {"supplement_explanation": "补充"}
    assert d["severity_display"] == "高"


@pytest.mark.asyncio
async def test_run_review_scope_finding_only(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-4c"
    ws["A1"] = "管理员账号识别情况"
    # no layout, no checkpoints, no preview -> only _check_sheet_scope fires
    llm = _FakeLLM(_pass_review_payload())

    findings, stats = await run_review(
        wb=wb, checkpoints={}, attachments_preview={}, sheets=None, llm=llm,
    )

    assert len(findings) == 1
    f0 = findings[0]
    assert f0["issue_type"] == "特权账号识别范围可能不完整"
    assert f0["severity"] == "P1"
    assert f0["severity_display"] == "中"
    assert f0["llm_status"] == "pass"
    assert f0["cross_validate_issues"] == []
    assert f0["challenge_verdict"] is None
    assert stats["total_findings"] == 1
    assert stats["by_severity"]["P1"] == 1


@pytest.mark.asyncio
async def test_run_review_assigns_static_taxonomy_to_deterministic_findings(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    wb.active.title = "SA-4c"
    wb.active["A1"] = "管理员账号识别情况"

    findings, _ = await run_review(
        wb=wb,
        checkpoints={},
        attachments_preview={},
        sheets=None,
        llm=_FakeLLM(_pass_review_payload()),
    )

    assert findings[0]["origin"] == "sheet_scope"
    assert findings[0]["rule_hint"] == "privileged_account_scope"


@pytest.mark.asyncio
async def test_run_review_assigns_controlled_assertion_and_claim(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    wb.active.title = "SA-4c"
    wb.active["A1"] = "管理员账号识别情况"

    findings, _ = await run_review(
        wb=wb,
        checkpoints={},
        attachments_preview={},
        sheets=None,
        llm=_FakeLLM(_pass_review_payload()),
    )

    finding = findings[0]
    assert finding["assertion_id"] == "scope.privileged_account.coverage"
    assert finding["claim_type"] == "population_coverage"
    assert finding["claim_subject"] == "SA-4c|scope:privileged_account_scope"
    assert finding["claim_value"] == "coverage_insufficient"


@pytest.mark.asyncio
async def test_run_review_records_gate_status_for_non_p0_findings(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    monkeypatch.setenv("REVIEW_DETERMINISTIC_CROSSCHECK_MODE", "all_findings")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-4c"
    ws["A1"] = "管理员账号识别情况"
    llm = _FakeLLM(_pass_review_payload())
    gate_context = build_quality_gate_context(
        workbook=wb,
        evidence_registry=None,
        assertion_catalog=default_assertion_catalog(),
    )

    findings, _ = await run_review(
        wb=wb,
        checkpoints={},
        attachments_preview={},
        sheets=None,
        llm=llm,
        quality_gate_context=gate_context,
    )

    gates = findings[0]["quality_gates"]
    assert gates["evidence_excerpt_matches_frozen_source"]["status"] == "passed"
    assert gates["evidence_excerpt_matches_frozen_source"]["duration_ms"] >= 0
    assert gates["adversarial_challenge"]["status"] == "not_run"
    assert gates["adversarial_challenge"]["reason"]
    assert gates["model_re_review"]["status"] in {
        "passed", "flagged", "not_run", "error",
    }


@pytest.mark.asyncio
async def test_run_review_marks_cross_check_not_run_when_disabled(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    monkeypatch.setenv("REVIEW_DETERMINISTIC_CROSSCHECK_MODE", "off")
    wb = openpyxl.Workbook()
    wb.active.title = "SA-4c"
    wb.active["A1"] = "管理员账号识别情况"

    findings, _ = await run_review(
        wb=wb, checkpoints={}, attachments_preview={}, sheets=None,
        llm=_FakeLLM(_pass_review_payload()),
    )

    gate = findings[0]["quality_gates"]["evidence_excerpt_matches_frozen_source"]
    assert gate["status"] == "not_run"
    assert "disabled" in gate["reason"]
    assert gate["duration_ms"] == 0


@pytest.mark.asyncio
async def test_direct_pipeline_does_not_pass_registry_gate_without_context(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    monkeypatch.setenv("REVIEW_DETERMINISTIC_CROSSCHECK_MODE", "all_findings")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-1"
    ws["A1"] = "见附件9"
    attachments = {
        "source_type": "preview",
        "items": [],
        "by_filename": {},
        "by_rel_path": {},
        "by_index": {},
        "by_sheet_norm": {},
    }

    findings, _ = await run_review(
        wb=wb,
        checkpoints={},
        attachments=attachments,
        sheets=None,
        llm=_FakeLLM(_pass_review_payload()),
    )

    finding = next(
        item
        for item in findings
        if item["assertion_id"] == "attachment.inventory.presence"
    )
    gates = finding["quality_gates"]
    assert gates["attachment_inventory_consistent"]["status"] == "not_run"
    assert gates["attachment_inventory_consistent"]["reason"] == "quality_context_unavailable"
    assert gates["claim_has_required_source_kind"]["status"] == "not_run"


def test_model_re_review_accepts_llm_origin_with_review_result():
    """LLM-origin findings now flow through `_llm_review_findings` so they
    also get a model re-review verdict — gating them as ``not_run`` based on
    origin alone was the root cause of the unactioned duplicates/severity
    issues in the audit report."""
    deterministic = _model_re_review_gate(
        {"origin": "sheet_scope", "issue_type": "LLM判定：历史文案", "status": "fail"},
        {"llm_status": "fail"},
    )
    llm_origin_agrees = _model_re_review_gate(
        {"origin": "llm", "issue_type": "普通标题", "status": "fail"},
        {"llm_status": "fail"},
    )
    llm_origin_no_review = _model_re_review_gate(
        {"origin": "llm", "issue_type": "普通标题", "status": "fail"},
        None,
    )

    assert deterministic["status"] == "passed"
    assert llm_origin_agrees["status"] == "passed"
    assert llm_origin_no_review["status"] == "not_run"
    assert llm_origin_no_review["reason"]


@pytest.mark.asyncio
async def test_run_review_empty_workbook(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    wb.active.title = "Empty"
    llm = _FakeLLM('{"results": []}')
    findings, stats = await run_review(
        wb=wb, checkpoints={}, attachments_preview={}, sheets=None, llm=llm,
    )
    assert findings == []
    assert stats["total_findings"] == 0


@pytest.mark.asyncio
async def test_run_review_falls_back_when_specified_sheet_not_reviewable(monkeypatch):
    """If `sheets` selects a sheet with no layout/checkpoints, fall back to all sheets."""
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    import openpyxl as _ox
    wb = _ox.Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "封面内容，无审计程序布局"
    ws = wb.create_sheet("SA-4c")
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws["A5"] = "审计期间获取系统用户清单并检查权限分配情况。"
    ws["B5"] = "我们导出了系统用户清单并进行了权限核查与记录。"
    llm = _FakeLLM('{"status":"pass","reason":"符合","evidence_refs":[]}')

    findings, stats = await run_review(
        wb=wb, checkpoints={}, attachments_preview={}, sheets="Cover", llm=llm,
    )
    # fell back to all sheets -> SA-4c was reviewed -> procedure_pair LLM call attempted
    assert stats.get("warning"), "expected a fallback warning"
    assert stats["llm_call_stats"].get("procedure_pair", {}).get("calls", 0) >= 1


@pytest.mark.asyncio
async def test_run_review_honors_reviewable_sheet_scope(monkeypatch):
    """A reviewable named sheet is reviewed ALONE — never expanded to all sheets.

    Regression guard for the scoping contract: when the caller passes
    `sheets="SA-4c"` and SA-4c has a detectable layout, the pipeline must
    review only SA-4c (no fallback). The agent relies on this to honor a
    user request to review a single control point.
    """
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    import openpyxl as _ox
    import review.pipeline as _pipe

    def _add_layout(ws, std, exe):
        ws["A1"] = "标准审计程序"
        ws["B1"] = "执行审计程序"
        ws["A5"] = std
        ws["B5"] = exe

    wb = _ox.Workbook()
    ws1 = wb.active
    ws1.title = "SA-4c"
    _add_layout(ws1, "审计期间获取系统用户清单并检查权限分配情况。",
                "我们导出了系统用户清单并进行了权限核查与记录。")
    ws2 = wb.create_sheet("SA-5")
    _add_layout(ws2, "抽样核查账号创建审批。", "我们抽样核查了账号创建审批记录。")
    llm = _FakeLLM('{"results": []}')

    reviewed: list = []

    async def _spy(llm, wb, target_sheets, attachments=None):
        reviewed.extend(target_sheets)
        return {}, []

    monkeypatch.setattr(_pipe, "_llm_check_procedure_pairs", _spy)

    findings, stats = await run_review(
        wb=wb, checkpoints={}, attachments_preview={}, sheets="SA-4c", llm=llm,
    )

    assert stats.get("warning", "") == "", "should not fall back to all sheets"
    assert reviewed == ["SA-4c"], (
        f"expected only SA-4c reviewed, got {reviewed} (scope was expanded?)"
    )


@pytest.mark.asyncio
async def test_run_review_resolves_loose_sheet_name(monkeypatch):
    """A loose sheet name (case/dash variant an LLM produces) resolves to the
    actual tab and is reviewed alone — no fallback to all sheets.

    Regression guard for the user-reported bug: agent passed `sheets='pe6'`
    but the actual tab is `PE-6`; the exact-match filter rejected it and fell
    back to all 34 sheets. The filter must normalize-match (PE-6, pe6, pe-6
    all resolve to the same tab).
    """
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    import openpyxl as _ox
    import review.pipeline as _pipe

    def _add_layout(ws, std, exe):
        ws["A1"] = "标准审计程序"
        ws["B1"] = "执行审计程序"
        ws["A5"] = std
        ws["B5"] = exe

    wb = _ox.Workbook()
    ws1 = wb.active
    ws1.title = "PE-6"  # actual tab: uppercase, dash
    _add_layout(ws1, "审计期间获取批处理作业清单。", "我们导出了批处理作业清单。")
    ws2 = wb.create_sheet("SA-4c")
    _add_layout(ws2, "获取系统用户清单并检查权限。", "我们导出用户清单，截图保存。")
    llm = _FakeLLM('{"results": []}')

    reviewed: list = []

    async def _spy(llm, wb, target_sheets, attachments=None):
        reviewed.extend(target_sheets)
        return {}, []

    monkeypatch.setattr(_pipe, "_llm_check_procedure_pairs", _spy)

    findings, stats = await run_review(
        wb=wb, checkpoints={}, attachments_preview={}, sheets="pe6", llm=llm,
    )

    assert stats.get("warning", "") == "", (
        f"loose name 'pe6' should resolve to PE-6, not fall back; warning={stats.get('warning')!r}"
    )
    assert reviewed == ["PE-6"], (
        f"expected only PE-6 reviewed (resolved from 'pe6'), got {reviewed}"
    )


@pytest.mark.asyncio
async def test_run_review_uses_constrained_evidence_agent_and_records_stats(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    monkeypatch.setenv("REVIEW_EVIDENCE_AGENT_MODE", "always")
    import review.pipeline as _pipe
    from review.models import AttachmentFile

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-4c"
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws["A5"] = "获取系统用户清单并检查权限。"
    ws["B5"] = "我们核验用户清单，见附件1。"
    item = AttachmentFile(
        index="1", rel_dir="SA-4c", filename="users.txt",
        rel_path="SA-4c/users.txt", file_type="txt", status="ok",
        extraction_status="ok", extracted_text="admin,administrator", size=19,
    )
    attachments = {
        "source_type": "directory",
        "path": "/pinned/attachments",
        "items": [item],
        "by_filename": {"users.txt": [item]},
        "by_rel_path": {"sa-4c/users.txt": [item]},
        "by_index": {"1": [item]},
        "by_sheet_norm": {"SA4C": [item]},
        "status_counts": {"ok": 1},
    }
    captured = {}

    async def _fake_agent(**kwargs):
        captured["sheet"] = kwargs["ws"].title
        return {
            "status": "completed",
            "evidence": [{
                "path": "SA-4c/users.txt",
                "file_type": "txt",
                "extraction_status": "ok",
                "excerpt": "admin,administrator",
                "supports": "用户清单中的管理员权限",
                "confidence": "high",
            }],
            "unresolved": [],
            "tool_trace": [{"tool": "search_attachment_text"}],
            "tool_calls": 1,
        }

    async def _fake_evidence_steps(**kwargs):
        captured["agent_evidence"] = kwargs["attachments"].get("agent_evidence_by_sheet")
        return []

    monkeypatch.setattr(_pipe, "investigate_sheet", _fake_agent)
    monkeypatch.setattr(_pipe, "_llm_check_evidence_vs_steps", _fake_evidence_steps)

    findings, stats = await _pipe.run_review(
        wb=wb, checkpoints={}, attachments=attachments, sheets=None,
        llm=_FakeLLM('{"results": []}'),
    )

    assert findings == []
    assert captured["sheet"] == "SA-4c"
    assert captured["agent_evidence"]["SA4C"][0]["excerpt"] == "admin,administrator"
    assert stats["evidence_agent"]["runs"] == 1
    assert stats["evidence_agent"]["accepted_evidence"] == 1
    assert stats["evidence_agent"]["tool_calls"] == 1
    assert stats["evidence_agent"]["details"] == [{
        "sheet": "SA-4c",
        "status": "completed",
        "tool_calls": 1,
        "evidence": [{
            "path": "SA-4c/users.txt",
            "file_type": "txt",
            "extraction_status": "ok",
            "excerpt": "admin,administrator",
            "supports": "用户清单中的管理员权限",
            "confidence": "high",
        }],
        "unresolved": [],
        "tool_trace": [{"tool": "search_attachment_text"}],
        "ocr": {},
    }]


@pytest.mark.asyncio
async def test_run_review_partial_match_warns_but_does_not_fall_back(monkeypatch):
    """Some requested names resolve (one reviewable, one not), some don't.

    The reviewable one is reviewed alone (no fallback to all sheets); the
    unresolvable name is surfaced in a warning. Regression guard for the
    mixed case, which the all-resolve / none-resolve tests don't cover.
    """
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    import openpyxl as _ox
    import review.pipeline as _pipe

    def _add_layout(ws, std, exe):
        ws["A1"] = "标准审计程序"
        ws["B1"] = "执行审计程序"
        ws["A5"] = std
        ws["B5"] = exe

    wb = _ox.Workbook()
    ws1 = wb.active
    ws1.title = "PE-6"
    _add_layout(ws1, "审计期间获取批处理作业清单。", "我们导出了批处理作业清单。")
    ws2 = wb.create_sheet("Cover")
    ws2["A1"] = "封面，无审计程序布局"  # resolves but not reviewable (no layout/checkpoints)
    llm = _FakeLLM('{"results": []}')

    reviewed: list = []

    async def _spy(llm, wb, target_sheets, attachments=None):
        reviewed.append(list(target_sheets))
        return {}, []

    monkeypatch.setattr(_pipe, "_llm_check_procedure_pairs", _spy)

    findings, stats = await run_review(
        wb=wb, checkpoints={}, attachments_preview={}, sheets="pe6,ZZZ", llm=llm,
    )

    # PE-6 resolved + reviewable -> reviewed; ZZZ unresolvable -> warned, not fatal
    assert reviewed == [["PE-6"]], f"expected only PE-6 reviewed, got {reviewed}"
    warning = stats.get("warning", "")
    assert warning, "expected a partial-match warning"
    assert "ZZZ" in warning, f"warning should mention unmatched ZZZ: {warning!r}"
    assert "全部 Sheet" not in warning, (
        f"should NOT fall back to all sheets: {warning!r}"
    )


@pytest.mark.asyncio
async def test_run_review_emits_progress_at_stages(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    wb.active.title = "Empty"
    llm = _FakeLLM('{"results": []}')
    recorded: list[dict] = []
    findings, stats = await run_review(
        wb=wb, checkpoints={}, attachments_preview={}, sheets=None, llm=llm,
        on_progress=lambda p: recorded.append(p),
    )
    stages = [p["stage"] for p in recorded]
    assert "starting" in stages
    assert "done" in stages
    assert recorded[-1]["stage"] == "done"
    assert recorded[-1]["findings_so_far"]["total"] == len(findings)
    for p in recorded:
        assert {"stage", "current_sheet", "llm_calls", "findings_so_far", "msg"} <= set(p.keys())
        assert isinstance(p["llm_calls"], dict)
        assert isinstance(p["findings_so_far"]["total"], int)


@pytest.mark.asyncio
async def test_run_review_ignores_on_progress_exceptions(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    wb.active.title = "Empty"
    llm = _FakeLLM('{"results": []}')

    def boom(p):
        raise RuntimeError("progress callback broken")

    findings, stats = await run_review(
        wb=wb, checkpoints={}, attachments_preview={}, sheets=None, llm=llm,
        on_progress=boom,
    )
    # review still completed despite the callback throwing
    assert stats["total_findings"] == len(findings)


@pytest.mark.asyncio
async def test_intra_chunk_progress_emitted_during_long_checkpoints_stage(monkeypatch):
    """A sheet with many checkpoints must emit progress between chunks, not only
    at the stage boundary. Otherwise long checkpoint reviews look frozen for
    many minutes.

    Regression guard for: 1 sheet with 11 checkpoints at batch_size=6 should
    emit at least 2 intra-chunk progress events (one per processed chunk).
    """
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    import openpyxl as _ox
    import review.pipeline as _pipe
    import review.checkpoints as _ckp

    wb = _ox.Workbook()
    ws = wb.active
    ws.title = "SA-12"

    # 11 checkpoints to drive >1 chunk with batch_size=6
    many_ckpts = [f"检查要点 {i}：核查对应证据并形成结论。" for i in range(1, 12)]
    llm = _FakeLLM('{"results": []}')

    intra_events: list[str] = []

    # Spy on the intra-chunk callback to capture messages emitted between chunks.
    import review.checkpoints as _ckp_mod

    async def _spy(llm, ws_title, ws, checkpoints, attachments=None,
                   batch_size=6, sleep_seconds=0, attachments_preview=None,
                   on_progress=None):
        # Mirror the real implementation's chunk loop just enough to invoke the
        # intra-chunk callback the way the real code does.
        for start in range(0, len(checkpoints), max(1, int(batch_size))):
            end = min(start + max(1, int(batch_size)), len(checkpoints))
            if on_progress is not None:
                on_progress(f"checkpoints:{ws_title}",
                            f"已处理 {end} / {len(checkpoints)} 个检查要点")
        return []

    monkeypatch.setattr(_pipe, "_llm_check_sheet_by_checkpoints", _spy)

    def collect(p):
        if p["current_sheet"] == "SA-12" and "已处理" in p["msg"]:
            intra_events.append(p["msg"])

    findings, stats = await run_review(
        wb=wb, checkpoints={"SA-12": many_ckpts}, attachments_preview={},
        sheets=None, llm=llm, on_progress=collect,
    )
    assert len(intra_events) >= 2, (
        f"expected >=2 intra-chunk events for 11 checkpoints, got {intra_events}"
    )
    assert intra_events[-1].endswith("/ 11 个检查要点"), (
        f"final intra-chunk event should report total 11, got {intra_events[-1]!r}"
    )
