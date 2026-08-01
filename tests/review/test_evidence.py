import json
import hashlib

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from review.evidence import build_evidence_graph, build_input_files, sha256_file, sha256_path


def test_build_input_files_hashes_workpaper_and_optional_inputs(tmp_path):
    workpaper = tmp_path / "wp.xlsx"
    checkpoints = tmp_path / "checkpoints.xlsx"
    attachments = tmp_path / "attachments.xlsx"
    workpaper.write_bytes(b"workpaper")
    checkpoints.write_bytes(b"checkpoints")
    attachments.write_bytes(b"attachments")

    inputs = build_input_files(
        workpaper_path=str(workpaper),
        checkpoints_path=str(checkpoints),
        attachments_preview_path=str(attachments),
    )

    assert [item.role for item in inputs] == [
        "workpaper",
        "checkpoints",
        "attachments_preview",
    ]
    assert inputs[0].filename == "wp.xlsx"
    assert inputs[0].size == len(b"workpaper")
    assert inputs[0].sha256 == hashlib.sha256(b"workpaper").hexdigest()


def test_build_input_files_records_attachment_directory_digest(tmp_path):
    workpaper = tmp_path / "wp.xlsx"
    workpaper.write_bytes(b"workpaper")
    attachments_dir = tmp_path / "attachments"
    (attachments_dir / "SA-4c").mkdir(parents=True)
    evidence = attachments_dir / "SA-4c" / "evidence.txt"
    evidence.write_text("evidence", encoding="utf-8")

    inputs = build_input_files(
        workpaper_path=str(workpaper),
        attachments_dir=str(attachments_dir),
    )

    assert [item.role for item in inputs] == ["workpaper", "attachments_dir"]
    assert inputs[1].media_type == "inode/directory"
    assert inputs[1].size == len(b"evidence")
    assert inputs[1].sha256 == sha256_path(attachments_dir)


def test_build_evidence_graph_is_deterministic_and_preserves_formula():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PE-6"
    ws["A1"] = "标准审计程序"
    ws["B2"] = "=1+1"

    first = build_evidence_graph(wb, source_sha256="a" * 64)
    second = build_evidence_graph(wb, source_sha256="a" * 64)

    first_cells = first.sheets[0].cells
    second_cells = second.sheets[0].cells
    formula_cell = next(cell for cell in first_cells if cell.coordinate == "B2")

    assert [cell.evidence_id for cell in first_cells] == [
        cell.evidence_id for cell in second_cells
    ]
    assert first.sheets[0].sheet_hash == second.sheets[0].sheet_hash
    assert formula_cell.value == "=1+1"
    assert formula_cell.formula == "=1+1"


def test_modern_formulas_are_stable_across_independent_reloads(tmp_path):
    path = tmp_path / "modern-formulas.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = ArrayFormula(ref="A1:A2", text="=SEQUENCE(2)")
    sheet["B1"] = DataTableFormula(
        ref="B1:C2",
        ca=True,
        dt2D=True,
        dtr=False,
        r1="A1",
        r2="A2",
        del1=False,
        del2=True,
    )
    workbook.save(path)
    source_sha256 = sha256_file(path)

    first = build_evidence_graph(
        openpyxl.load_workbook(path, data_only=False),
        source_sha256=source_sha256,
    )
    second = build_evidence_graph(
        openpyxl.load_workbook(path, data_only=False),
        source_sha256=source_sha256,
    )
    first_cells = {cell.coordinate: cell for cell in first.sheets[0].cells}
    second_cells = {cell.coordinate: cell for cell in second.sheets[0].cells}

    assert first_cells["A1"].content_hash == second_cells["A1"].content_hash
    assert first_cells["A1"].evidence_id == second_cells["A1"].evidence_id
    assert first_cells["B1"].content_hash == second_cells["B1"].content_hash
    assert first_cells["B1"].evidence_id == second_cells["B1"].evidence_id

    array_formula = json.loads(first_cells["A1"].formula)
    assert array_formula == {
        "ref": "A1:A2",
        "t": "array",
        "text": "=SEQUENCE(2)",
    }
    data_table_formula = json.loads(first_cells["B1"].formula)
    assert data_table_formula == {
        "ca": True,
        "del1": False,
        "del2": True,
        "dt2D": True,
        "dtr": False,
        "r1": "A1",
        "r2": "A2",
        "ref": "B1:C2",
        "t": "dataTable",
    }
    assert first_cells["A1"].value == first_cells["A1"].formula
    assert first_cells["B1"].value == first_cells["B1"].formula


def test_build_evidence_graph_marks_explicit_truncation():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PE-6"
    ws["A1"] = "first"
    ws["B1"] = "second"
    ws["C1"] = "third"

    graph = build_evidence_graph(wb, source_sha256="a" * 64, max_cells=2)

    assert graph.capture_status == "truncated"
    assert graph.captured_cell_count == 2
    assert graph.omitted_cell_count == 1
    assert [cell.coordinate for cell in graph.sheets[0].cells] == ["A1", "B1"]


def test_build_evidence_graph_records_detected_layout_and_merged_ranges():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PE-6"
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws["A2"] = "合并的标准程序"
    ws.merge_cells("A2:B2")

    graph = build_evidence_graph(wb, source_sha256="a" * 64)
    sheet = graph.sheets[0]

    assert sheet.layout_header_row == 1
    assert sheet.standard_column == 1
    assert sheet.execution_columns == [2]
    assert sheet.merged_ranges == ["A2:B2"]


def test_sheet_hash_covers_merged_range_metadata():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "PE-6"
    sheet["A1"] = "标准审计程序"
    sheet["B1"] = "执行审计程序"
    sheet["A2"] = "保持不变的单元格值"

    before_merge = build_evidence_graph(workbook, source_sha256="a" * 64)
    sheet.merge_cells("A2:B2")
    after_merge = build_evidence_graph(workbook, source_sha256="a" * 64)

    assert [cell.coordinate for cell in before_merge.sheets[0].cells] == [
        cell.coordinate for cell in after_merge.sheets[0].cells
    ]
    assert before_merge.sheets[0].sheet_hash != after_merge.sheets[0].sheet_hash
