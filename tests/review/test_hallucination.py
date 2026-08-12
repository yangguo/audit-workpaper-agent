import json

import openpyxl
import pytest
from langchain_core.messages import AIMessage

from review.hallucination import (
    _cross_validate_finding,
    _build_minimal_context,
    _challenge_finding_with_llm,
)
from review.models import Finding


class _FakeRunnable:
    def __init__(self, content):
        self.content = content

    async def ainvoke(self, messages):
        return AIMessage(content=self.content)


class _FakeLLM:
    def __init__(self, content):
        self.content = content

    def bind(self, **kwargs):
        return _FakeRunnable(self.content)


def _make_finding(**kw):
    base = dict(
        issue_type="t", severity="P1", sheet="Sheet", cell=None,
        snippet="", basis="b", suggestion="s",
    )
    base.update(kw)
    return Finding(**base)


def test_cross_validate_exception_flag_contradicts_pass():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = "是否发现异常：是"
    f = _make_finding(status="pass", cell="A1")
    assert "exception_flag_contradicts_pass" in _cross_validate_finding(f, wb)


def test_cross_validate_high_severity_no_evidence():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    f = _make_finding(status="fail", severity="P0", evidence_refs="[]")
    assert "high_severity_no_evidence" in _cross_validate_finding(f, wb)


def test_cross_validate_does_not_apply_sample_size_by_risk_type():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    f = _make_finding(status="fail", severity="P1", risk_type="覆盖性")

    assert "coverage_claim_but_no_sample_size" not in _cross_validate_finding(f, wb)


def test_cross_validate_excerpt_mismatch():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = "这是实际的单元格内容"
    f = _make_finding(
        status="fail", severity="P1",
        evidence_refs=json.dumps([{"cell_or_range": "A1", "excerpt": "完全不相关"}], ensure_ascii=False),
    )
    assert "evidence_excerpt_mismatch" in _cross_validate_finding(f, wb)


def test_cross_validate_missing_sheet_returns_empty():
    f = _make_finding(status="fail", severity="P0", sheet="Missing")
    wb = openpyxl.Workbook()
    assert _cross_validate_finding(f, wb) == []


def test_build_minimal_context_returns_cell_text():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = "表头"
    ws["A5"] = "执行描述内容"
    f = _make_finding(cell="A5")
    ctx = _build_minimal_context(f, ws)
    assert "A5" in ctx and "执行描述内容" in ctx


def test_build_minimal_context_empty_ws_returns_empty():
    f = _make_finding(cell=None, snippet="")
    assert _build_minimal_context(f, None) == ""


@pytest.mark.asyncio
async def test_challenge_returns_agree():
    llm = _FakeLLM("agree")
    f = _make_finding()
    out = await _challenge_finding_with_llm(llm=llm, finding=f, minimal_context="ctx")
    assert out == "agree"


@pytest.mark.asyncio
async def test_challenge_returns_disagree():
    llm = _FakeLLM("disagree")
    f = _make_finding()
    out = await _challenge_finding_with_llm(llm=llm, finding=f, minimal_context="ctx")
    assert out == "disagree"


@pytest.mark.asyncio
async def test_challenge_empty_context_returns_none():
    llm = _FakeLLM("agree")
    f = _make_finding()
    assert await _challenge_finding_with_llm(llm=llm, finding=f, minimal_context="") is None
