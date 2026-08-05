import openpyxl
from review.excel_utils import (
    _is_empty,
    _get_cell_value,
    _get_cell_text,
    _truncate,
    _detect_layout,
    _extract_sheet_text_cells,
    _build_sheet_text_for_llm,
    _normalize_sheet_id,
)


def test_is_empty():
    assert _is_empty(None) is True
    assert _is_empty("   ") is True
    assert _is_empty("x") is False
    assert _is_empty(0) is False


def test_truncate():
    assert _truncate("short", 10) == "short"
    assert _truncate("1234567890ab", 5) == "12345..."


def test_get_cell_value_handles_merged_cells(layout_workbook):
    ws = layout_workbook.active
    ws.merge_cells("A3:B3")
    ws["A3"] = "合并值"
    assert _get_cell_value(ws, "B3") == "合并值"


def test_get_cell_value_empty_returns_none(layout_workbook):
    ws = layout_workbook.active
    assert _get_cell_value(ws, "Z9") is None


def test_get_cell_value_invalid_ref_returns_none(layout_workbook):
    """An LLM-returned cell_or_range that isn't a real coordinate must not crash the review."""
    ws = layout_workbook.active
    assert _get_cell_value(ws, "sheet名") is None
    assert _get_cell_value(ws, "not a cell") is None


def test_get_cell_text_strips_constructed_marker(layout_workbook):
    ws = layout_workbook.active
    ws["A2"] = "用户清单"
    assert _get_cell_text(ws, "A2[非逐字原文]") == "用户清单"
    assert _get_cell_text(ws, "") == ""


def test_detect_layout_finds_standard_and_exec(layout_workbook):
    ws = layout_workbook.active
    header_row, std_col, exec_cols = _detect_layout(ws)
    assert header_row == 1
    assert std_col == 1
    assert exec_cols == [2]


def test_detect_layout_returns_empty_when_no_layout(blank_workbook):
    ws = blank_workbook.active
    ws["A1"] = "无关文本"
    assert _detect_layout(ws) == (None, 0, [])


def test_extract_sheet_text_cells_yields_coord_text(layout_workbook):
    ws = layout_workbook.active
    cells = list(_extract_sheet_text_cells(ws))
    coords = {c for c, _ in cells}
    assert "A1" in coords and "A2" in coords
    text_map = dict(cells)
    assert text_map["B2"].startswith("我们导出用户清单")


def test_build_sheet_text_for_llm_respects_limits(layout_workbook):
    ws = layout_workbook.active
    text = _build_sheet_text_for_llm(ws, max_cells=1, max_chars=10_000)
    assert text.startswith("A1:")
    assert text.count("\n") == 0  # only one cell


def test_normalize_sheet_id():
    assert _normalize_sheet_id("sa-4c") == "SA4C"
    assert _normalize_sheet_id("PM_5") == "PM5"
    assert _normalize_sheet_id("  sa 4c ") == "SA4C"
