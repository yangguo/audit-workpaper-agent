import asyncio
import contextlib
import json
import threading

import openpyxl
import pytest
import pytest_asyncio
from langchain_core.messages import AIMessage

import review.runner as runner
from review.runner import start_review, get_status, cancel_all_running, list_running, _REGISTRY
from review.evidence import sha256_file
from storage.findings_store import load_findings
from storage.review_artifact_store import ReviewArtifactStore


class _FakeRunnable:
    def __init__(self, content):
        self.content = content

    async def ainvoke(self, messages):
        await asyncio.sleep(0.05)  # simulate a slow review step
        return AIMessage(content=self.content)


class _FakeLLM:
    def __init__(self, content):
        self.content = content

    def bind(self, **kwargs):
        return _FakeRunnable(self.content)


def _pass_payload():
    import json
    return json.dumps({"results": [{
        "id": 1, "status": "pass", "conclusion": "复核不成立结论",
        "evidence_refs": [], "llm_validity": "不成立", "llm_severity": "低",
        "severity": "P2", "reasons": ["理由"], "risk_type": "证据不足",
    }]}, ensure_ascii=False)


@pytest_asyncio.fixture(autouse=True)
async def _isolate(monkeypatch, tmp_path):
    """Each test gets a clean registry + fake LLM + temp workspace."""
    _REGISTRY.clear()
    monkeypatch.setenv("WORKSPACE_PATH", str(tmp_path))
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    monkeypatch.setattr(runner, "get_review_llm", lambda: _FakeLLM(_pass_payload()))
    yield

    review_tasks = [
        entry["task"]
        for entry in _REGISTRY.values()
        if isinstance(entry.get("task"), asyncio.Task) and not entry["task"].done()
    ]
    for task in review_tasks:
        task.cancel()
    if review_tasks:
        await asyncio.gather(*review_tasks, return_exceptions=True)

    shadow_tasks = [
        entry["shadow_task"]
        for entry in _REGISTRY.values()
        if isinstance(entry.get("shadow_task"), asyncio.Task)
        and not entry["shadow_task"].done()
    ]
    if shadow_tasks:
        await asyncio.gather(*shadow_tasks, return_exceptions=True)


def _make_workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-4c"
    ws["A1"] = "管理员账号识别情况"
    path = tmp_path / "wp.xlsx"
    wb.save(str(path))
    return str(path)


def _make_judgement_workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-4c"
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws["A5"] = "审计期间获取系统用户清单并检查管理员权限。"
    ws["B5"] = "我们获取用户清单截图，并核对管理员权限。"
    path = tmp_path / "judgement-wp.xlsx"
    wb.save(str(path))
    return str(path)


@pytest.mark.asyncio
async def test_start_review_returns_running_status_immediately(tmp_path):
    wp = _make_workbook(tmp_path)
    review_id = await start_review(file_path=wp, source="wp.xlsx")

    assert isinstance(review_id, str) and len(review_id) == 32
    st = get_status(review_id)
    assert st["status"] in ("running", "completed")  # may have finished fast
    assert st["source"] == "wp.xlsx"


@pytest.mark.asyncio
async def test_review_completes_and_writes_findings(tmp_path):
    wp = _make_workbook(tmp_path)
    review_id = await start_review(file_path=wp, source="wp.xlsx")

    # wait for the background task to finish
    entry = _REGISTRY[review_id]
    await entry["task"]

    st = get_status(review_id)
    assert st["status"] == "completed"
    assert st["stats"]["total_findings"] >= 1

    from storage.findings_store import load_findings
    payload = load_findings(review_id)
    assert payload is not None
    assert payload["findings"][0]["issue_type"] == "特权账号识别范围可能不完整"


@pytest.mark.asyncio
async def test_review_quality_shadow_adds_provenance_without_changing_legacy_fields(
    monkeypatch, tmp_path
):
    monkeypatch.setenv("REVIEW_RESULT_QUALITY_MODE", "shadow")
    wp = _make_workbook(tmp_path)
    review_id = await start_review(file_path=wp, source="wp.xlsx")

    await _REGISTRY[review_id]["task"]

    payload = load_findings(review_id)
    assert payload is not None
    finding = payload["findings"][0]
    assert finding["quality"]["schema_version"] == "review-quality/1"
    assert finding["quality"]["provenance"]["input_sha256"]
    assert finding["quality"]["finding_id"].startswith("legacy:")
    assert finding["origin"] == "sheet_scope"
    assert finding["rule_hint"] == "privileged_account_scope"
    assert finding["quality"]["citation_validation"]["status"] in {
        "verified",
        "partial",
        "invalid",
        "not_available",
    }
    assert "grouping" in finding["quality"]
    assert finding["quality"]["grouping"]["root_cause_id"]
    assert finding["quality"]["remediation"]["status"] in {
        "actionable",
        "needs_human_refinement",
        "not_available",
    }
    assert payload["stats"]["quality"]["raw_findings"] == len(payload["findings"])
    assert finding["issue_type"] == "特权账号识别范围可能不完整"


@pytest.mark.asyncio
async def test_execution_manifest_records_assertion_catalog_component(tmp_path):
    review_id = await start_review(file_path=_make_workbook(tmp_path), source="wp.xlsx")

    await _REGISTRY[review_id]["task"]
    await _REGISTRY[review_id]["shadow_task"]

    manifest = ReviewArtifactStore().load_manifest(review_id)
    assert manifest is not None
    assert {
        component["component_id"] for component in manifest["components"]
    } >= {"review-quality-assertions"}


@pytest.mark.asyncio
async def test_manifest_is_started_before_v1_and_quality_reuses_execution_identity(
    monkeypatch, tmp_path
):
    started_manifests = []
    original_begin = ReviewArtifactStore.begin

    def _record_begin(self, manifest):
        started_manifests.append(manifest.model_dump(mode="json"))
        return original_begin(self, manifest)

    async def _assert_manifest_exists_before_review(**kwargs):
        assert len(started_manifests) == 1
        manifest = started_manifests[0]
        assert manifest["input_set_sha256"]
        assert manifest["execution_sha256"]
        return (
            [
                {
                    "issue_type": "受控测试发现",
                    "severity": "P1",
                    "status": "fail",
                    "sheet": "SA-4c",
                    "cell": "A1",
                    "origin": "checkpoint",
                    "rule_hint": "test",
                    "evidence_refs": [
                        {
                            "sheet": "SA-4c",
                            "cell_or_range": "A1",
                            "excerpt": "管理员账号识别情况",
                        }
                    ],
                    "quality_gates": {},
                }
            ],
            {"total_findings": 1},
        )

    monkeypatch.setenv("REVIEW_RESULT_QUALITY_MODE", "shadow")
    monkeypatch.setattr(ReviewArtifactStore, "begin", _record_begin)
    monkeypatch.setattr(runner, "run_review", _assert_manifest_exists_before_review)
    review_id = await start_review(file_path=_make_workbook(tmp_path), source="wp.xlsx")

    await _REGISTRY[review_id]["task"]
    payload = load_findings(review_id)
    await _REGISTRY[review_id]["shadow_task"]

    manifest = ReviewArtifactStore().load_manifest(review_id)
    assert payload is not None
    assert manifest is not None
    assert len(started_manifests) == 1
    assert payload["findings"][0]["quality"]["provenance"]["input_set_sha256"] == manifest[
        "input_set_sha256"
    ]
    assert payload["findings"][0]["quality"]["provenance"]["execution_sha256"] == manifest[
        "execution_sha256"
    ]


@pytest.mark.asyncio
async def test_completed_review_starts_shadow_artifact_without_changing_v1_result(tmp_path):
    review_id = await start_review(file_path=_make_workbook(tmp_path), source="wp.xlsx")

    await _REGISTRY[review_id]["task"]
    await _REGISTRY[review_id]["shadow_task"]

    assert get_status(review_id)["status"] == "completed"
    assert get_status(review_id)["artifact_status"] == "completed"
    assert load_findings(review_id)["review_id"] == review_id
    assert ReviewArtifactStore().load_manifest(review_id)["artifact_status"] == "completed"


@pytest.mark.asyncio
async def test_stage_c_shadow_writes_judgements_and_v2_findings_without_changing_v1(
    monkeypatch, tmp_path
):
    monkeypatch.setenv("REVIEW_JUDGEMENT_MODE", "shadow")
    review_id = await start_review(
        file_path=_make_judgement_workbook(tmp_path), source="wp.xlsx"
    )

    await _REGISTRY[review_id]["task"]
    v1_payload = load_findings(review_id)
    await _REGISTRY[review_id]["shadow_task"]

    artifact_dir = tmp_path / "assets" / "reviews" / review_id
    manifest = json.loads((artifact_dir / "manifest.json").read_text("utf-8"))
    judgements = json.loads((artifact_dir / "judgements.json").read_text("utf-8"))
    v2_findings = json.loads(
        (artifact_dir / "v2-findings.json").read_text("utf-8")
    )
    comparison = json.loads(
        (artifact_dir / "comparison.json").read_text("utf-8")
    )

    assert get_status(review_id)["status"] == "completed"
    assert get_status(review_id)["artifact_status"] == "completed"
    assert manifest["judgement_policy_pack"] == {
        "id": "itgc-judgement",
        "version": "1.0.0",
    }
    assert judgements["schema_version"] == "stage-c-judgements/1"
    assert judgements["requests"]
    assert judgements["results"]
    assert v2_findings["schema_version"] == "stage-c-v2-findings/1"
    assert v2_findings["findings"]
    assert comparison["schema_version"] == "review-finding-comparison/1"
    assert set(comparison["counts"]) == {
        "agreement",
        "legacy_only",
        "shadow_only",
        "status_conflict",
        "evidence_conflict",
        "not_comparable",
    }
    assert load_findings(review_id) == v1_payload


@pytest.mark.asyncio
async def test_stage_c_off_does_not_write_judgement_artifacts(monkeypatch, tmp_path):
    monkeypatch.setenv("REVIEW_JUDGEMENT_MODE", "off")
    review_id = await start_review(
        file_path=_make_judgement_workbook(tmp_path), source="wp.xlsx"
    )

    await _REGISTRY[review_id]["task"]
    await _REGISTRY[review_id]["shadow_task"]

    artifact_dir = tmp_path / "assets" / "reviews" / review_id
    manifest = json.loads((artifact_dir / "manifest.json").read_text("utf-8"))
    assert manifest["judgement_policy_pack"] is None
    assert not (artifact_dir / "judgements.json").exists()
    assert not (artifact_dir / "v2-findings.json").exists()


@pytest.mark.asyncio
async def test_stage_c_failure_does_not_fail_completed_v1_review(monkeypatch, tmp_path):
    monkeypatch.setenv("REVIEW_JUDGEMENT_MODE", "shadow")

    async def _fail_judgement(*args, **kwargs):
        raise RuntimeError("stage c boom")

    monkeypatch.setattr(runner, "execute_judgement_requests", _fail_judgement)
    review_id = await start_review(
        file_path=_make_judgement_workbook(tmp_path), source="wp.xlsx"
    )

    await _REGISTRY[review_id]["task"]
    v1_payload = load_findings(review_id)
    await _REGISTRY[review_id]["shadow_task"]

    status = get_status(review_id)
    assert status["status"] == "completed"
    assert status["artifact_status"] == "error"
    assert "RuntimeError: stage c boom" in status["artifact_error"]
    manifest = ReviewArtifactStore().load_manifest(review_id)
    assert manifest is not None
    assert manifest["artifact_status"] == "error"
    assert "RuntimeError: stage c boom" in manifest["artifact_error"]
    assert load_findings(review_id) == v1_payload


@pytest.mark.asyncio
async def test_completed_review_writes_stage_b_policy_artifacts_without_changing_v1(
    tmp_path,
):
    review_id = await start_review(file_path=_make_workbook(tmp_path), source="wp.xlsx")

    await _REGISTRY[review_id]["task"]
    v1_payload = load_findings(review_id)
    await _REGISTRY[review_id]["shadow_task"]

    artifact_dir = tmp_path / "assets" / "reviews" / review_id
    manifest = json.loads((artifact_dir / "manifest.json").read_text("utf-8"))
    plan = json.loads((artifact_dir / "review-plan.json").read_text("utf-8"))
    policy_findings = json.loads(
        (artifact_dir / "policy-findings.json").read_text("utf-8")
    )

    assert manifest["policy_pack"] == {"id": "itgc-core", "version": "1.0.0"}
    assert plan["schema_version"] == "stage-b-plan/1"
    assert policy_findings["schema_version"] == "stage-b-policy-findings/1"
    assert load_findings(review_id) == v1_payload
    assert get_status(review_id)["status"] == "completed"


@pytest.mark.asyncio
async def test_invalid_stage_b_policy_only_fails_shadow_artifact(
    monkeypatch, tmp_path
):
    monkeypatch.setenv("REVIEW_POLICY_PACK_ROOT", str(tmp_path / "missing-policy"))
    review_id = await start_review(file_path=_make_workbook(tmp_path), source="wp.xlsx")

    await _REGISTRY[review_id]["task"]
    v1_payload = load_findings(review_id)
    await _REGISTRY[review_id]["shadow_task"]

    status = get_status(review_id)
    assert status["status"] == "completed"
    assert status["artifact_status"] == "error"
    assert "PolicyPackError" in status["artifact_error"]
    assert load_findings(review_id) == v1_payload


@pytest.mark.asyncio
async def test_completed_artifact_uses_v1_input_versions_when_sources_change(
    monkeypatch, tmp_path
):
    workpaper_path = _make_workbook(tmp_path)
    workpaper = openpyxl.load_workbook(workpaper_path)
    workpaper.active["A1"] = "ORIGINAL-WORKPAPER"
    workpaper.save(workpaper_path)

    checkpoints_path = tmp_path / "checkpoints.xlsx"
    checkpoints = openpyxl.Workbook()
    checkpoints.active["A1"] = "SA-4c"
    checkpoints.active["C1"] = "ORIGINAL-CHECKPOINT"
    checkpoints.save(checkpoints_path)

    attachments_path = tmp_path / "attachments.xlsx"
    attachments = openpyxl.Workbook()
    attachments.active.title = "图片描述"
    attachments.active["A1"] = "附件文件名"
    attachments.active["B1"] = "详细描述"
    attachments.active["A2"] = "evidence.png"
    attachments.active["B2"] = "ORIGINAL-ATTACHMENT"
    attachments.save(attachments_path)

    source_paths = [workpaper_path, str(checkpoints_path), str(attachments_path)]
    original_hashes = [sha256_file(path) for path in source_paths]

    def _replace_workbook(path, value):
        replacement = openpyxl.Workbook()
        replacement.active["A1"] = value
        replacement.save(path)

    async def _mutating_review(*, wb, checkpoints, attachments_preview, **kwargs):
        assert wb.active["A1"].value == "ORIGINAL-WORKPAPER"
        assert checkpoints == {"SA-4c": ["ORIGINAL-CHECKPOINT"]}
        assert "ORIGINAL-ATTACHMENT" in str(attachments_preview)
        _replace_workbook(workpaper_path, "REPLACED-WORKPAPER")
        _replace_workbook(checkpoints_path, "REPLACED-CHECKPOINT")
        _replace_workbook(attachments_path, "REPLACED-ATTACHMENT")
        return ([{"issue_type": "V1-ORIGINAL"}], {"total_findings": 1})

    monkeypatch.setattr(runner, "run_review", _mutating_review)
    review_id = await start_review(
        file_path=workpaper_path,
        checkpoints_path=str(checkpoints_path),
        attachments_preview_path=str(attachments_path),
        source="wp.xlsx",
    )

    await _REGISTRY[review_id]["task"]
    await _REGISTRY[review_id]["shadow_task"]

    artifact_dir = tmp_path / "assets" / "reviews" / review_id
    manifest = json.loads((artifact_dir / "manifest.json").read_text("utf-8"))
    evidence = json.loads((artifact_dir / "evidence.json").read_text("utf-8"))
    artifact_findings = json.loads(
        (artifact_dir / "findings.json").read_text("utf-8")
    )

    assert get_status(review_id)["artifact_status"] == "completed"
    assert [item["sha256"] for item in manifest["inputs"]] == original_hashes
    assert [sha256_file(path) for path in source_paths] != original_hashes
    assert evidence["sheets"][0]["cells"][0]["value"] == "ORIGINAL-WORKPAPER"
    assert artifact_findings["findings"] == [{"issue_type": "V1-ORIGINAL"}]


@pytest.mark.asyncio
async def test_review_runner_indexes_pinned_attachment_directory(monkeypatch, tmp_path):
    workpaper_path = _make_workbook(tmp_path)
    attachments_dir = tmp_path / "attachments"
    (attachments_dir / "SA-4c").mkdir(parents=True)
    (attachments_dir / "SA-4c" / "附件1-user-list.txt").write_text(
        "admin,管理员", encoding="utf-8"
    )
    captured = {}

    async def _capture_review(*, wb, checkpoints, attachments, **kwargs):
        captured["attachments"] = attachments
        return ([], {"total_findings": 0})

    monkeypatch.setattr(runner, "run_review", _capture_review)
    review_id = await start_review(
        file_path=workpaper_path,
        attachments_dir=str(attachments_dir),
        source="wp.xlsx",
    )

    await _REGISTRY[review_id]["task"]
    await _REGISTRY[review_id]["shadow_task"]

    item = captured["attachments"]["items"][0]
    assert item.rel_path == "SA-4c/附件1-user-list.txt"
    assert item.extracted_text == "admin,管理员"
    assert item.extraction_status == "ok"


@pytest.mark.asyncio
async def test_snapshot_failure_does_not_fail_v1_review(monkeypatch, tmp_path):
    def _fail_snapshot(*args, **kwargs):
        raise OSError("snapshot storage unavailable")

    monkeypatch.setattr(ReviewArtifactStore, "snapshot_inputs", _fail_snapshot)
    review_id = await start_review(file_path=_make_workbook(tmp_path), source="wp.xlsx")

    await _REGISTRY[review_id]["task"]

    status = get_status(review_id)
    assert status["status"] == "completed"
    assert status["artifact_status"] == "error"
    assert status["artifact_error"] == "OSError: snapshot storage unavailable"
    assert load_findings(review_id) is not None
    assert "shadow_task" not in _REGISTRY[review_id]
    assert not (tmp_path / "assets" / "reviews" / review_id / "manifest.json").exists()


@pytest.mark.asyncio
async def test_snapshot_failure_is_retained_when_v1_fallback_is_cancelled(
    monkeypatch, tmp_path
):
    def _fail_snapshot(*args, **kwargs):
        raise OSError("snapshot storage unavailable")

    review_entered = asyncio.Event()

    async def _blocked_review(**kwargs):
        review_entered.set()
        await asyncio.Event().wait()

    monkeypatch.setattr(ReviewArtifactStore, "snapshot_inputs", _fail_snapshot)
    monkeypatch.setattr(runner, "run_review", _blocked_review)
    review_id = await start_review(file_path=_make_workbook(tmp_path), source="wp.xlsx")

    await asyncio.wait_for(review_entered.wait(), timeout=0.5)
    assert cancel_all_running() == 1
    with contextlib.suppress(asyncio.CancelledError):
        await _REGISTRY[review_id]["task"]

    status = get_status(review_id)
    assert status["status"] == "cancelled"
    assert status["artifact_status"] == "error"
    assert status["artifact_error"] == "OSError: snapshot storage unavailable"


@pytest.mark.asyncio
async def test_shadow_failure_does_not_fail_existing_review(monkeypatch, tmp_path):
    monkeypatch.setattr(
        runner,
        "build_evidence_graph",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("boom")),
    )
    review_id = await start_review(file_path=_make_workbook(tmp_path), source="wp.xlsx")

    await _REGISTRY[review_id]["task"]
    await _REGISTRY[review_id]["shadow_task"]

    assert get_status(review_id)["status"] == "completed"
    assert get_status(review_id)["artifact_status"] == "error"
    assert load_findings(review_id) is not None


@pytest.mark.asyncio
async def test_shadow_artifact_capture_keeps_event_loop_responsive(monkeypatch, tmp_path):
    release_capture = threading.Event()
    original_build_input_files = runner.build_input_files

    def _blocked_build_input_files(*args, **kwargs):
        release_capture.wait(timeout=0.2)
        return original_build_input_files(*args, **kwargs)

    monkeypatch.setattr(runner, "build_input_files", _blocked_build_input_files)
    review_id = "shadow-responsive"
    _REGISTRY[review_id] = {"status": "completed"}
    shadow_task = asyncio.create_task(
        runner._capture_shadow_artifact(
            review_id=review_id,
            file_path=_make_workbook(tmp_path),
            checkpoints_path="",
            attachments_preview_path="",
            sheets=None,
            source="wp.xlsx",
            findings=[],
            stats={},
        )
    )

    timer = threading.Timer(0.2, release_capture.set)
    timer.start()
    try:
        loop = asyncio.get_running_loop()
        started_at = loop.time()
        await asyncio.sleep(0.01)
        assert loop.time() - started_at < 0.1
    finally:
        release_capture.set()
        timer.cancel()
        await shadow_task


@pytest.mark.asyncio
async def test_cancelled_shadow_capture_retains_its_original_workspace(monkeypatch, tmp_path):
    capture_started = threading.Event()
    release_capture = threading.Event()
    capture_finished = threading.Event()
    observed = {}
    original_build_input_files = runner.build_input_files

    class _RecordingArtifactStore:
        def __init__(self, *, workspace_path=None):
            self.workspace_path = workspace_path

        def begin(self, manifest):
            observed["workspace_path"] = self.workspace_path or runner.os.getenv(
                "WORKSPACE_PATH"
            )

        def write_evidence(self, review_id, graph):
            return None

        def write_v1_findings(self, review_id, findings, stats):
            return None

        def write_review_plan(self, review_id, plan):
            return None

        def write_policy_findings(self, review_id, payload):
            return None

        def complete(self, review_id):
            capture_finished.set()
            return None

        def fail(self, review_id, error):
            capture_finished.set()
            return None

    def _blocked_build_input_files(*args, **kwargs):
        capture_started.set()
        release_capture.wait(timeout=0.5)
        return original_build_input_files(*args, **kwargs)

    monkeypatch.setattr(runner, "ReviewArtifactStore", _RecordingArtifactStore)
    monkeypatch.setattr(runner, "build_input_files", _blocked_build_input_files)
    review_id = "shadow-cleanup"
    _REGISTRY[review_id] = {"status": "completed"}
    shadow_task = asyncio.create_task(
        runner._capture_shadow_artifact(
            review_id=review_id,
            file_path=_make_workbook(tmp_path),
            checkpoints_path="",
            attachments_preview_path="",
            sheets=None,
            source="wp.xlsx",
            findings=[],
            stats={},
        )
    )

    await asyncio.to_thread(capture_started.wait, 0.5)
    shadow_task.cancel()
    await asyncio.gather(shadow_task, return_exceptions=True)
    monkeypatch.setenv("WORKSPACE_PATH", str(tmp_path / "restored-workspace"))
    release_capture.set()
    try:
        await asyncio.wait_for(asyncio.to_thread(capture_finished.wait, 0.5), 0.6)
        assert observed["workspace_path"] == str(tmp_path)
    finally:
        release_capture.set()
        await asyncio.to_thread(capture_finished.wait, 0.5)


@pytest.mark.asyncio
async def test_start_review_cancels_prior_running(tmp_path):
    wp = _make_workbook(tmp_path)
    first = await start_review(file_path=wp, source="first")
    # start a second while the first is still running
    second = await start_review(file_path=wp, source="second")

    # let tasks settle (the cancelled first task raises CancelledError on await)
    with contextlib.suppress(asyncio.CancelledError):
        await _REGISTRY[first]["task"]
    await _REGISTRY[second]["task"]

    assert _REGISTRY[first]["status"] == "cancelled"
    assert _REGISTRY[second]["status"] == "completed"
    assert list_running() == []


@pytest.mark.asyncio
async def test_cancel_all_running_cancels_in_flight(tmp_path):
    wp = _make_workbook(tmp_path)
    rid = await start_review(file_path=wp, source="wp")
    n = cancel_all_running()
    assert n == 1
    assert _REGISTRY[rid]["status"] == "cancelled"
    # draining the cancelled task does not raise to the caller
    entry = _REGISTRY[rid]
    with contextlib_suppress():
        await entry["task"]


def contextlib_suppress():
    return contextlib.suppress(asyncio.CancelledError)


@pytest.mark.asyncio
async def test_get_status_unknown_returns_none():
    assert get_status("does-not-exist") is None


@pytest.mark.asyncio
async def test_progress_callback_updates_status():
    _REGISTRY.clear()
    review_id = "test-rid"
    _REGISTRY[review_id] = {
        "status": "running",
        "started_at": runner._now(),
        "source": "wp.xlsx",
    }
    # before any progress, status.progress is None
    assert get_status(review_id)["progress"] is None

    cb = runner._make_progress_cb(review_id)
    cb({
        "stage": "checkpoints",
        "current_sheet": "SA-4c",
        "llm_calls": {"checkpoints": 2},
        "findings_so_far": {"P0": 0, "P1": 1, "P2": 0, "total": 1},
        "msg": "完成 SA-4c checkpoint 评审",
    })
    st = get_status(review_id)
    assert st["progress"]["stage"] == "checkpoints"
    assert st["progress"]["current_sheet"] == "SA-4c"
    assert st["progress"]["findings_so_far"]["total"] == 1
    assert st["progress"]["recent_events"][-1]["msg"] == "完成 SA-4c checkpoint 评审"
    assert len(st["progress"]["recent_events"]) == 1

    # rolling window caps recent_events at 15
    for i in range(20):
        cb({
            "stage": "checkpoints", "current_sheet": "", "llm_calls": {},
            "findings_so_far": {"P0": 0, "P1": 0, "P2": 0, "total": 0},
            "msg": f"evt {i}",
        })
    assert len(get_status(review_id)["progress"]["recent_events"]) == 15
