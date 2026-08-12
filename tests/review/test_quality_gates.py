import openpyxl
import pytest

from review.contracts import EvidenceFact
from review.evidence_facts import EvidenceFactRegistry
from review.finding_taxonomy import AssertionCatalog, AssertionSpec, default_assertion_catalog
from review.quality_gates import (
    QualityGateConfigurationError,
    TRUSTED_QUALITY_GATES,
    build_quality_gate_context,
    run_assertion_gates,
)


def _finding(assertion_id: str, *, subject: str = "SA-1|attachment:config.txt"):
    assertion = default_assertion_catalog().assertion(assertion_id)
    return {
        "assertion_id": assertion_id,
        "claim_type": assertion.claim_type,
        "claim_subject": subject,
        "claim_value": "coverage_insufficient",
        "risk_type": "覆盖性",
        "status": "fail",
        "severity": "P1",
        "sheet": "SA-1",
        "evidence_refs": [],
    }


@pytest.fixture
def gate_context():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-1"
    ws["A1"] = "样本量"
    ws["B1"] = 25
    registry = EvidenceFactRegistry(
        [
            EvidenceFact(
                fact_id="attachment:config",
                fact_type="attachment",
                source_ref="config.txt",
                source_sha256="source-sha",
                content_hash="content-sha",
                sheet_scope=["SA-1"],
                extraction_status="ok",
                source_type="directory",
            )
        ]
    )
    return build_quality_gate_context(
        workbook=wb,
        evidence_registry=registry,
        assertion_catalog=default_assertion_catalog(),
    )


def test_asset_coverage_does_not_receive_sample_size_gate(gate_context):
    outcomes = run_assertion_gates(
        _finding("configuration.asset_coverage.support"), gate_context
    )

    assert "sample_size_present" not in outcomes
    assert "configuration_scope_supported" in outcomes
    assert outcomes["configuration_scope_supported"]["duration_ms"] >= 0


def test_every_default_catalog_gate_has_a_trusted_evaluator(gate_context):
    declared = {
        gate_id
        for assertion in gate_context.assertion_catalog.assertions
        for gate_id in assertion.deterministic_gate_ids
    }

    assert declared <= set(TRUSTED_QUALITY_GATES)


def test_population_sample_size_assertion_runs_only_its_declared_gate(gate_context):
    outcomes = run_assertion_gates(
        _finding("population.sample_size.present", subject="SA-1|cell:A1"),
        gate_context,
    )

    assert set(outcomes) == {"sample_size_present"}
    assert outcomes["sample_size_present"]["status"] == "passed"


def test_registry_dependent_gate_is_not_run_without_registry(gate_context):
    context_without_registry = build_quality_gate_context(
        workbook=gate_context.workbook,
        evidence_registry=None,
        assertion_catalog=gate_context.assertion_catalog,
    )

    outcomes = run_assertion_gates(
        _finding("attachment.inventory.presence"), context_without_registry
    )

    assert outcomes["attachment_inventory_consistent"]["status"] == "not_run"
    assert outcomes["attachment_inventory_consistent"]["reason"] == "quality_context_unavailable"


def test_excerpt_gate_does_not_pass_when_finding_sheet_is_unavailable(gate_context):
    finding = _finding(
        "scope.privileged_account.coverage", subject="Missing|scope:privileged_account_scope"
    )
    finding["sheet"] = "Missing"

    outcomes = run_assertion_gates(finding, gate_context)

    assert outcomes["evidence_excerpt_matches_frozen_source"]["status"] == "not_run"
    assert outcomes["evidence_excerpt_matches_frozen_source"]["reason"] == "finding_sheet_unavailable"


def test_context_rejects_catalog_gate_that_has_no_trusted_evaluator(gate_context):
    invalid_catalog = AssertionCatalog(
        id="test-catalog",
        version="1.0.0",
        assertions=[
            AssertionSpec(
                assertion_id="test.invalid.gate",
                version="1.0.0",
                claim_type="workpaper_text",
                allowed_origins=["checkpoint"],
                requires_attachment_support=False,
                exclusive_claim=False,
                deterministic_gate_ids=["not_registered"],
                remediation_template_id="human_refinement",
            )
        ],
    )

    with pytest.raises(QualityGateConfigurationError, match="not_registered"):
        build_quality_gate_context(
            workbook=gate_context.workbook,
            evidence_registry=gate_context.evidence_registry,
            assertion_catalog=invalid_catalog,
        )
