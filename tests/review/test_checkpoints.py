import json

import openpyxl
import pytest
from langchain_core.messages import AIMessage

from review.checkpoints import (
    load_checkpoints_xlsx,
    _split_checkpoints,
    _extract_checkpoint_keywords,
    _llm_check_sheet_by_checkpoints,
)


class _FakeRunnable:
    def __init__(self, content):
        self.content = content

    async def ainvoke(self, messages):
        return AIMessage(content=self.content)


class _FakeLLM:
    def __init__(self, content):
        self.content = content

    def bind(self, **kwargs):
        return _FakeRunnable(self.content)


def test_split_checkpoints_strips_numbering():
    assert _split_checkpoints("1. 检查用户清单\n2. 核对权限") == ["检查用户清单", "核对权限"]


def test_split_checkpoints_handles_empty():
    assert _split_checkpoints("") == []
    assert _split_checkpoints("   ") == []


def test_split_checkpoints_single_chunk():
    assert _split_checkpoints("单条检查要点") == ["单条检查要点"]


def test_extract_checkpoint_keywords_vocab_hits():
    assert _extract_checkpoint_keywords("获取用户清单并核对权限") == ["用户清单"]


def test_extract_checkpoint_keywords_segments_when_no_vocab():
    out = _extract_checkpoint_keywords("某段含中文的关键描述文字")
    assert out == ["某段含中文的关键描述文字"]


def test_extract_checkpoint_keywords_empty():
    assert _extract_checkpoint_keywords("") == []


def test_load_checkpoints_xlsx_groups_by_sheet(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="SA-4c")
    ws.cell(row=1, column=3, value="检查用户清单")
    ws.cell(row=2, column=1, value=None)
    ws.cell(row=2, column=3, value="核对权限矩阵")
    ws.cell(row=3, column=1, value="SA-5")
    ws.cell(row=3, column=3, value="检查变更日志")
    path = tmp_path / "checkpoints.xlsx"
    wb.save(str(path))

    result = load_checkpoints_xlsx(str(path))
    assert result == {
        "SA-4c": ["检查用户清单", "核对权限矩阵"],
        "SA-5": ["检查变更日志"],
    }


def test_load_checkpoints_xlsx_empty_path_returns_empty():
    assert load_checkpoints_xlsx("") == {}


@pytest.mark.asyncio
async def test_llm_check_sheet_by_checkpoints_returns_findings(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "表头"
    ws["A2"] = "用户清单导出记录"
    payload = json.dumps({
        "results": [{
            "id": 1,
            "checkpoint": "检查用户清单",
            "status": "fail",
            "conclusion": "结论文字至少四个字",
            "reasons": ["理由一"],
            "evidence_refs": [{"cell_or_range": "A2", "excerpt": "用户清单"}],
            "severity": "P1",
            "risk_type": "证据不足",
            "fix_suggestion": {"supplement_explanation": "补充截图"},
        }]
    }, ensure_ascii=False)
    llm = _FakeLLM(payload)

    findings = await _llm_check_sheet_by_checkpoints(
        llm=llm, ws_title="SA-1", ws=ws,
        checkpoints=["检查用户清单"], batch_size=6, sleep_seconds=0,
    )

    assert len(findings) == 1
    assert findings[0].issue_type == "LLM判定：检查要点存在问题（证据不足）"
    assert findings[0].status == "fail"
    assert findings[0].severity == "P1"


@pytest.mark.asyncio
async def test_llm_check_sheet_by_checkpoints_degrades_on_failure(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "表头"

    class _ErrRunnable:
        async def ainvoke(self, messages):
            raise RuntimeError("timed out")

    class _ErrLLM:
        def bind(self, **kwargs):
            return _ErrRunnable()

    findings = await _llm_check_sheet_by_checkpoints(
        llm=_ErrLLM(), ws_title="SA-1", ws=ws,
        checkpoints=["检查用户清单"], batch_size=6, sleep_seconds=0,
    )
    assert len(findings) == 1
    assert findings[0].status == "unknown"
    assert findings[0].issue_type == "LLM判定：检查要点复核失败"


def test_checkpoints_prompt_includes_evidence_inventory(monkeypatch):
    """The checkpoints review prompt should include attachment inventory."""
    from review import checkpoints as cp
    from review import llm as llm_mod

    captured = {}

    async def fake_chat(*, llm, messages, stage, max_attempts=2, max_tokens=4096):
        captured["messages"] = messages
        return json.dumps({"results": []})

    monkeypatch.setattr(llm_mod, "_llm_chat", fake_chat)

    # Build minimal ws
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A5"] = "测试检查要点"

    # Build minimal attachments
    attachments = {
        "items": [
            {"rel_path": ".embedded_media/test.docx::image1.png", "status": "binary", "file_type": "png", "extracted_text": ""},
            {"rel_path": "审计证据/test.txt", "status": "ok", "file_type": "txt", "extracted_text": "hello"},
        ],
    }

    import asyncio
    asyncio.run(cp._llm_check_sheet_by_checkpoints(
        llm=None, ws_title="TEST", ws=ws,
        checkpoints=["检查点1"], attachments=attachments, attachments_preview=attachments,
        on_progress=None,
    ))

    # Check the prompt includes inventory
    sys_msg = next(m for m in captured["messages"] if m["role"] == "system")
    assert "证据清单" in sys_msg["content"] or "证据清单" in str(captured["messages"])
