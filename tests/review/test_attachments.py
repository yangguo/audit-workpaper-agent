import openpyxl

from review.attachments import (
    _extract_attachment_refs,
    _compact_keywords,
    _evidence_matches_step,
    _match_preview_items,
    build_evidence_inventory,
    load_attachments_preview_xlsx,
    _check_attachment_references,
)
from review.models import AttachmentPreviewItem


def test_extract_attachment_refs():
    filenames, rel_paths, indices = _extract_attachment_refs(
        "见附件1，截图 screenshot.png，路径 dir/sub/file.xlsx，压缩包 export.rar"
    )
    assert "screenshot.png" in filenames
    assert "dir/sub/file.xlsx" in rel_paths
    assert "export.rar" in filenames
    assert "1" in indices


def test_extract_attachment_refs_empty():
    assert _extract_attachment_refs("") == ([], [], [])


def test_compact_keywords_filters_stopwords():
    out = _compact_keywords("审计 程序 用户清单 权限矩阵")
    assert out == ["用户清单", "权限矩阵"]


def test_evidence_matches_step_overlap():
    assert _evidence_matches_step("获取用户清单", "用户清单导出截图") is True


def test_evidence_matches_step_no_overlap():
    assert _evidence_matches_step("密码策略复杂度", "用户清单导出") is False


def test_evidence_matches_step_empty_is_true():
    assert _evidence_matches_step("", "x") is True
    assert _evidence_matches_step("x", "") is True


def _make_item(**kw):
    base = dict(index="1", rel_dir="d", filename="a.png", rel_path="d/a.png",
                file_type="png", description="desc", status="OK")
    base.update(kw)
    return AttachmentPreviewItem(**base)


def test_match_preview_items_by_filename_and_missing_index():
    item = _make_item()
    preview = {
        "by_filename": {"a.png": [item]},
        "by_rel_path": {},
        "by_index": {"1": [item]},
        "by_sheet_norm": {},
        "path": "p",
        "items": [item],
        "status_counts": {"OK": 1},
    }
    picked, missing = _match_preview_items(
        preview, filenames=["a.png"], rel_paths=[], indices=["2"],
    )
    assert item in picked
    assert "索引2" in missing


def test_load_attachments_preview_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "图片描述"
    headers = ["目录索引", "相对目录", "附件文件名", "相对路径", "文件类型", "详细描述", "状态"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="SA-4c")
    ws.cell(row=2, column=3, value="a.png")
    ws.cell(row=2, column=4, value="SA-4c/a.png")
    ws.cell(row=2, column=5, value="png")
    ws.cell(row=2, column=6, value="用户清单截图")
    ws.cell(row=2, column=7, value="OK")
    path = tmp_path / "preview.xlsx"
    wb.save(str(path))

    preview = load_attachments_preview_xlsx(str(path))
    assert len(preview["items"]) == 1
    assert preview["items"][0].filename == "a.png"
    assert "a.png" in preview["by_filename"]
    assert "1" in preview["by_index"]
    assert "SA4C" in preview["by_sheet_norm"]


def test_check_attachment_references_flags_missing_index():
    item = _make_item(status="OK")
    preview = {
        "by_filename": {}, "by_rel_path": {},
        "by_index": {"1": [item]}, "by_sheet_norm": {},
        "path": "p", "items": [item], "status_counts": {"OK": 1},
    }
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "见附件9"
    findings = _check_attachment_references("SA-1", ws, preview)
    assert len(findings) == 1
    assert findings[0].issue_type == "附件证据引用未匹配到附件目录"
    assert findings[0].severity == "P2"


def test_check_attachment_references_empty_preview_returns_empty(layout_workbook):
    assert _check_attachment_references("SA-1", layout_workbook.active, {}) == []


# Add to existing test file
from review.attachments import build_evidence_inventory


def test_build_evidence_inventory_empty():
    assert build_evidence_inventory({}) == ""
    assert build_evidence_inventory(None or {}) == ""  # type: ignore


def test_build_evidence_inventory_lists_real_attachments(tmp_path):
    """Two real text-extractable attachments should appear with status [ok]."""
    from review.attachments import build_attachment_index
    from pathlib import Path
    d = tmp_path / "atts"
    d.mkdir()
    (d / "a.txt").write_text("hello world " * 20, encoding="utf-8")
    # Use a real zip-based docx so build_attachment_index can extract its text
    # and mark it with status [ok] (the brief's snippet had a typo here).
    import io
    import zipfile
    docx = d / "b.docx"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("word/document.xml", b"<?xml version='1.0'?><doc><p>fake docx content</p></doc>")
    docx.write_bytes(buf.getvalue())
    idx = build_attachment_index(str(d))
    inv = build_evidence_inventory(idx)
    assert "证据清单" in inv
    assert "[ok] a.txt" in inv
    assert "[ok] b.docx" in inv


def test_build_evidence_inventory_groups_embedded_media(tmp_path):
    """Embedded media items should be grouped by source_document with [EMBED] header."""
    # Use the actual extract pipeline on a real DOCX
    import io
    import zipfile
    from review.attachments import build_attachment_index
    d = tmp_path / "atts"
    d.mkdir()
    docx = d / "test.docx"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("word/document.xml", b"<?xml version='1.0'?><doc/>")
        zf.writestr("word/media/image1.png", b"\x89PNG\r\n\x1a\n" + b"\x00" * 4)
    docx.write_bytes(buf.getvalue())
    idx = build_attachment_index(str(d))
    inv = build_evidence_inventory(idx)
    # .embedded_media section present
    assert ".embedded_media" in inv
    assert "test.docx" in inv


def test_build_evidence_inventory_truncates(tmp_path):
    """More than max_entries + max_embedded should be truncated, with hint text."""
    from review.attachments import build_attachment_index
    d = tmp_path / "atts"
    d.mkdir()
    for i in range(50):
        (d / f"f{i}.txt").write_text(f"file {i} content", encoding="utf-8")
    idx = build_attachment_index(str(d))
    inv = build_evidence_inventory(idx, max_entries=10, max_embedded=5)
    # Should mention truncation
    assert "实际有" in inv or "前 10" in inv or "前 5" in inv
