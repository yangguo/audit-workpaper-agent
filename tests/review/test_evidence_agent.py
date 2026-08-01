import json

import openpyxl
import pytest
from langchain_core.messages import AIMessage

from review.evidence_agent import (
    build_evidence_tools,
    investigate_sheet,
    should_run_evidence_agent,
    validate_agent_result,
)
from review.mineru_client import MinerUResult
from review.models import AttachmentFile


def _item(**overrides):
    values = {
        "index": "1",
        "rel_dir": "SA-4c",
        "filename": "user-list.txt",
        "rel_path": "SA-4c/user-list.txt",
        "file_type": "txt",
        "description": "admin,administrator",
        "status": "ok",
        "size": 18,
        "extracted_text": "admin,administrator\nroot,administrator",
        "extraction_status": "ok",
    }
    values.update(overrides)
    return AttachmentFile(**values)


def _index(*items, source_type="directory"):
    items = list(items)
    by_filename = {}
    by_rel_path = {}
    by_index = {}
    by_sheet_norm = {}
    for item in items:
        by_filename.setdefault(item.filename.lower(), []).append(item)
        by_rel_path.setdefault(item.rel_path.lower(), []).append(item)
        if item.index:
            by_index.setdefault(item.index, []).append(item)
        if item.rel_dir:
            by_sheet_norm.setdefault(item.rel_dir.replace("-", "").upper(), []).append(item)
    return {
        "path": "/pinned/review/attachments",
        "source_type": source_type,
        "items": items,
        "by_filename": by_filename,
        "by_rel_path": by_rel_path,
        "by_index": by_index,
        "by_sheet_norm": by_sheet_norm,
        "status_counts": {"ok": sum(item.extraction_status == "ok" for item in items)},
    }


def test_evidence_tools_list_search_and_read_only_return_indexed_content():
    index = _index(_item())
    trace = []
    tools = {tool.name: tool for tool in build_evidence_tools(index, trace=trace)}

    listed = json.loads(tools["list_attachment_files"].invoke({"query": "user"}))
    assert listed["files"] == [{
        "path": "SA-4c/user-list.txt",
        "index": "1",
        "rel_dir": "SA-4c",
        "file_type": "txt",
        "size": 18,
        "extraction_status": "ok",
    }]

    searched = json.loads(tools["search_attachment_text"].invoke({"query": "administrator"}))
    assert searched["matches"][0]["path"] == "SA-4c/user-list.txt"
    assert "admin,administrator" in searched["matches"][0]["excerpt"]

    read = json.loads(tools["read_attachment"].invoke({"path": "SA-4c/user-list.txt"}))
    assert read["path"] == "SA-4c/user-list.txt"
    assert read["content"] == "admin,administrator\nroot,administrator"
    assert [event["tool"] for event in trace] == [
        "list_attachment_files",
        "search_attachment_text",
        "read_attachment",
    ]


def test_read_attachment_rejects_escape_and_unindexed_paths():
    index = _index(_item())
    tools = {tool.name: tool for tool in build_evidence_tools(index)}

    for path in ("/etc/passwd", "../secret.txt", "SA-4c/missing.txt"):
        result = json.loads(tools["read_attachment"].invoke({"path": path}))
        assert result["status"] == "rejected"
        assert "content" not in result


def test_validate_agent_result_accepts_only_exact_source_excerpts():
    index = _index(_item())
    payload = {
        "evidence": [{
            "path": "SA-4c/user-list.txt",
            "excerpt": "root,administrator",
            "supports": "权限清单包含管理员账号",
            "confidence": "high",
        }, {
            "path": "SA-4c/not-real.txt",
            "excerpt": "编造内容",
        }, {
            "path": "SA-4c/user-list.txt",
            "excerpt": "编造内容",
        }],
        "unresolved": [{"request": "截图内容", "reason": "图片未解析"}],
    }

    result = validate_agent_result(payload, index)

    assert result["evidence"] == [{
        "path": "SA-4c/user-list.txt",
        "file_type": "txt",
        "extraction_status": "ok",
        "excerpt": "root,administrator",
        "supports": "权限清单包含管理员账号",
        "confidence": "high",
    }]
    assert len(result["unresolved"]) == 3
    assert result["unresolved"][0]["reason"] == "source_not_indexed"
    assert result["unresolved"][1]["reason"] == "excerpt_not_in_source"


def test_ocr_tool_uploads_only_indexed_binary_file_and_caches_verified_text(tmp_path):
    item = _item(
        filename="screenshot.png",
        rel_path="SA-4c/screenshot.png",
        file_type="png",
        status="binary",
        extraction_status="binary",
        extracted_text="",
        description="",
        size=12,
    )
    index = _index(item)
    index["path"] = str(tmp_path)
    (tmp_path / "SA-4c").mkdir()
    (tmp_path / "SA-4c" / "screenshot.png").write_bytes(b"image")
    client_calls = []

    class _FakeMinerU:
        def parse_file(self, path, **kwargs):
            client_calls.append((path, kwargs))
            return MinerUResult(
                status="ok",
                text="截图中显示：用户 admin 具有管理员权限",
                provider="mineru-lightweight",
                task_id="task-ocr",
            )

    tools = {tool.name: tool for tool in build_evidence_tools(index, mineru_client=_FakeMinerU())}
    result = json.loads(tools["ocr_attachment"].invoke({"path": "SA-4c/screenshot.png"}))

    assert result["status"] == "ok"
    assert result["path"] == "SA-4c/screenshot.png"
    assert "用户 admin" in result["content"]
    assert client_calls[0][0].name == "screenshot.png"
    assert client_calls[0][1]["is_ocr"] is True
    assert index["ocr_by_path"]["sa-4c/screenshot.png"]["content"] == result["content"]

    validated = validate_agent_result({
        "evidence": [{
            "path": "SA-4c/screenshot.png",
            "excerpt": "用户 admin 具有管理员权限",
            "supports": "截图中的管理员权限",
        }],
    }, index)
    assert validated["evidence"][0]["path"] == "SA-4c/screenshot.png"


def test_ocr_tool_rejects_unindexed_path_and_preserves_remote_failure(tmp_path):
    index = _index(_item(
        filename="screenshot.png",
        rel_path="SA-4c/screenshot.png",
        file_type="png",
        status="binary",
        extraction_status="binary",
        extracted_text="",
    ))
    index["path"] = str(tmp_path)
    (tmp_path / "SA-4c").mkdir()
    (tmp_path / "SA-4c" / "screenshot.png").write_bytes(b"image")

    class _FakeMinerU:
        def parse_file(self, path, **kwargs):
            return MinerUResult(status="error", error="HTTP 429", provider="mineru-lightweight")

    tools = {tool.name: tool for tool in build_evidence_tools(index, mineru_client=_FakeMinerU())}
    assert json.loads(tools["ocr_attachment"].invoke({"path": "../secret.png"}))["reason"] == "path_not_indexed"
    failed = json.loads(tools["ocr_attachment"].invoke({"path": "SA-4c/screenshot.png"}))
    assert failed["status"] == "error"
    assert failed["reason"] == "mineru_failed"


def test_fallback_trigger_requires_directory_and_evidence_gap(monkeypatch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-4c"
    ws["A1"] = "执行程序"
    ws["A2"] = "我们核验用户清单，见附件1"

    ok_index = _index(_item())
    assert should_run_evidence_agent(ws, ok_index, mode="fallback") is False

    unresolved_index = _index(_item(filename="screenshot.png", rel_path="SA-4c/screenshot.png",
                                    file_type="png", status="binary", extraction_status="binary",
                                    extracted_text="", description=""))
    assert should_run_evidence_agent(ws, unresolved_index, mode="fallback") is True

    preview_index = _index(unresolved_index["items"][0], source_type="preview")
    assert should_run_evidence_agent(ws, preview_index, mode="fallback") is False
    unrelated_bad = _index(
        _item(),
        _item(index="9", rel_dir="SA-9", rel_path="SA-9/broken.rar", filename="broken.rar",
              file_type="rar", status="unsupported", extraction_status="unsupported",
              extracted_text=""),
    )
    assert should_run_evidence_agent(ws, unrelated_bad, mode="fallback") is False
    monkeypatch.setenv("REVIEW_EVIDENCE_AGENT_MODE", "off")
    assert should_run_evidence_agent(ws, unresolved_index) is False


@pytest.mark.asyncio
async def test_investigate_sheet_uses_constrained_agent_and_returns_validated_evidence():
    index = _index(_item())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-4c"
    ws["A1"] = "执行程序"
    ws["A2"] = "我们核验用户清单，见附件1"

    class _FakeAgent:
        async def ainvoke(self, payload, config=None):
            assert payload["messages"]
            return {
                "messages": [AIMessage(content=json.dumps({
                    "evidence": [{
                        "path": "SA-4c/user-list.txt",
                        "excerpt": "admin,administrator",
                        "supports": "用户清单中的管理员权限",
                        "confidence": "high",
                    }],
                    "unresolved": [],
                }, ensure_ascii=False))]
            }

    captured = {}

    def _factory(**kwargs):
        captured.update(kwargs)
        return _FakeAgent()

    result = await investigate_sheet(
        ws=ws,
        attachments=index,
        llm=object(),
        agent_factory=_factory,
        mode="always",
    )

    assert result["status"] == "completed"
    assert result["evidence"][0]["path"] == "SA-4c/user-list.txt"
    assert result["evidence"][0]["excerpt"] == "admin,administrator"
    assert captured["tools"]
    assert captured["system_prompt"]


@pytest.mark.asyncio
async def test_investigate_sheet_marks_non_json_agent_response_as_error():
    index = _index(_item())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-4c"
    ws["A1"] = "执行程序，见附件1"

    class _BadAgent:
        async def ainvoke(self, payload, config=None):
            return {"messages": [AIMessage(content="无法按格式输出")]}

    result = await investigate_sheet(
        ws=ws,
        attachments=index,
        llm=object(),
        agent_factory=lambda **kwargs: _BadAgent(),
        mode="always",
    )

    assert result["status"] == "error"
    assert result["unresolved"][0]["reason"] == "invalid_agent_json"
