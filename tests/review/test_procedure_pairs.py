import openpyxl
import pytest
from langchain_core.messages import AIMessage

from review.procedure_pairs import (
    _requires_evidence_by_standard,
    _likely_interview_only,
    _check_procedure_pairs,
    _check_sheet_scope,
    _llm_judge_procedure_pair,
    _llm_check_procedure_pairs,
    _classify_mismatch,
)


class _FakeRunnable:
    def __init__(self, content):
        self.content = content

    async def ainvoke(self, messages):
        return AIMessage(content=self.content)


class _FakeLLM:
    def __init__(self, content):
        self.content = content

    def bind(self, **kwargs):
        return _FakeRunnable(self.content)


def test_requires_evidence_by_standard():
    assert "截图/参数界面" in _requires_evidence_by_standard("获取系统截图界面")
    assert "导出清单" in _requires_evidence_by_standard("导出用户清单")
    assert "审批/授权证据" in _requires_evidence_by_standard("检查审批记录")
    assert list(_requires_evidence_by_standard("无关键词的标准程序")) == []


def test_likely_interview_only():
    assert _likely_interview_only("通过访谈了解权限设置") is True
    assert _likely_interview_only("导出用户清单截图") is False
    assert _likely_interview_only("我们检查了配置") is False


def _build_procedure_ws(standard: str, execution: str, row: int = 5):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws.cell(row=row, column=1, value=standard)
    ws.cell(row=row, column=2, value=execution)
    return ws


def test_check_procedure_pairs_flags_unreplaced_template():
    ws = _build_procedure_ws(
        standard="审计期间获取系统用户清单并检查权限分配情况。",
        execution="如果系统存在以下情况，被认为需要复核方可执行。",
    )
    findings = _check_procedure_pairs("SA-1", ws)
    assert len(findings) == 1
    assert findings[0].issue_type == "执行列疑似未替换模板/未按要求填列"
    assert findings[0].severity == "P0"


def test_check_procedure_pairs_flags_interview_only():
    ws = _build_procedure_ws(
        standard="审计期间访谈了解权限管理流程并记录结果。",
        execution="我们通过访谈了解权限管理流程，未获取其他证据。",
    )
    findings = _check_procedure_pairs("SA-1", ws)
    assert any(f.issue_type == "程序执行不到位/仅依赖访谈" for f in findings)


def test_check_procedure_pairs_no_layout_returns_empty(blank_workbook):
    ws = blank_workbook.active
    ws["A1"] = "无关文本"
    assert _check_procedure_pairs("SA-1", ws) == []


def test_check_sheet_scope_flags_missing_os_db_admins():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "管理员账号识别情况"
    findings = _check_sheet_scope("SA-4c", ws)
    assert len(findings) == 1
    assert findings[0].issue_type == "特权账号识别范围可能不完整"


def test_check_sheet_scope_no_finding_when_os_db_covered():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "管理员账号与操作系统数据库账号核对"
    assert _check_sheet_scope("SA-4c", ws) == []


def test_check_sheet_scope_supplier_without_agreement():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "供应商维护系统管理员账号"
    findings = _check_sheet_scope("PM-5", ws)
    assert len(findings) == 1
    assert findings[0].issue_type == "供应商托管场景证据可能不足"


@pytest.mark.asyncio
async def test_llm_judge_procedure_pair_parses_json(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    llm = _FakeLLM('{"status":"fail","reason":"理由文字","evidence_refs":[]}')
    success, is_match, reason, raw = await _llm_judge_procedure_pair(
        llm=llm, standard_text="获取系统用户清单并检查权限。",
        execution_text="我们导出用户清单并核查权限。",
    )
    assert success is True
    assert is_match is False
    assert "理由文字" in reason


@pytest.mark.asyncio
async def test_llm_check_procedure_pairs_flags_mismatch(monkeypatch):
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws["A5"] = "审计期间获取系统用户清单并检查权限分配情况。"
    ws["B5"] = "我们导出了系统用户清单并进行了权限核查与记录。"
    llm = _FakeLLM('{"status":"fail","reason":"执行未覆盖关键条件","evidence_refs":[]}')

    report, findings = await _llm_check_procedure_pairs(
        llm=llm, wb=wb, target_sheets=[ws.title], start_row=5, sleep_seconds=0,
    )
    assert report["total"] >= 1
    assert any(f.issue_type.startswith("LLM判定：") for f in findings)


@pytest.mark.asyncio
async def test_procedure_pairs_prompt_includes_evidence_inventory(monkeypatch):
    """The A-C judgement prompt should carry the attachment inventory.

    Captures the messages at the fake LLM boundary rather than monkeypatching
    ``_llm_chat``: ``procedure_pairs`` binds ``_llm_chat`` at import time, so
    patching ``review.llm._llm_chat`` would not intercept the call here.
    """
    monkeypatch.setenv("REVIEW_LLM_BACKOFF_SCALE", "0")
    captured = []

    class _CapturingRunnable:
        async def ainvoke(self, messages):
            captured.append(messages)
            return AIMessage(content='{"status":"pass","reason":"符合","evidence_refs":[]}')

    class _CapturingLLM:
        def bind(self, **kwargs):
            return _CapturingRunnable()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws["A5"] = "审计期间获取系统用户清单并检查权限分配情况。"
    ws["B5"] = "我们导出了系统用户清单并进行了权限核查与记录。"

    # One real attachment + one embedded media item.
    attachments = {
        "items": [
            {
                "rel_path": "审计证据/用户清单.txt",
                "status": "ok",
                "file_type": "txt",
                "extracted_text": "admin,administrator",
            },
            {
                "rel_path": ".embedded_media/密码策略.docx::image1.png",
                "status": "binary",
                "file_type": "png",
                "extracted_text": "",
            },
        ],
    }

    await _llm_check_procedure_pairs(
        llm=_CapturingLLM(), wb=wb, target_sheets=[ws.title],
        start_row=5, sleep_seconds=0, attachments=attachments,
    )

    assert captured, "LLM was never called"
    user_text = "\n".join(
        str(getattr(m, "content", m)) for m in captured[0]
    )
    assert "证据清单" in user_text or "证据不足" in user_text
    assert "审计证据/用户清单.txt" in user_text


def test_classify_mismatch_distinguishes_sa12_reasons():
    """Regression guard: SA-12 produced 5 P0 findings that all looked identical
    because they shared the generic '执行程序不符合标准审计程序' label.
    The classifier should now map them to distinct issue_type buckets.
    """
    cases = [
        # C15 - missing personnel list
        ("获取准确及完整的执行特权级别用户日志复核的人员清单",
         "实际执行中获取并查看的是审阅记录，未明确提及获取复核人员清单，未覆盖标准要求的关键审计动作和证据类型。",
         "LLM判定：执行程序未覆盖复核人员/权限范围"),
        # C18 - missing log/system scope
        ("获取准确及完整的信息系统特权级别用户日志复核清单（检查范围包括应用程序、操作系统及数据库/堡垒机等日志）",
         "实际执行仅确认对应用层管理员操作日志进行了全量复核，未覆盖操作系统、数据库/堡垒机等日志范围，未满足标准的关键检查范围要求。",
         "LLM判定：执行程序未覆盖日志/系统范围"),
        # C19 - review precision / exception follow-up
        ("确认日志复核工作是否足够精确、复核内容是否适当，以保证相关异常操作能够被发现并追查原因",
         "实际执行仅说明查看了日志复核审阅记录并确认未见明显异常，但未提供足够证据证明复核工作足够精确、复核内容适当。",
         "LLM判定：执行程序未充分评估复核精确性与异常跟进"),
        # C5 - design effectiveness conclusion lacks evidence
        ("询问IT管理层并了解日志管理制度、记录方式、级别、范围；检查证据以证实控制设计。",
         "标准要求检查证据以证实控制设计，而实际执行未描述对证据的检查，仅基于访谈得出结论，证据不足。",
         "LLM判定：设计有效性结论缺乏证据支持"),
    ]
    for standard, reason, expected_issue in cases:
        issue_type, sev, _ = _classify_mismatch(standard, "", reason)
        assert issue_type == expected_issue, (
            f"expected {expected_issue!r}, got {issue_type!r} for reason: {reason[:80]}"
        )
        assert sev == "高"


def test_classify_mismatch_falls_back_to_generic():
    issue_type, sev, _ = _classify_mismatch("标准程序", "执行程序", "不符合要求")
    assert issue_type == "LLM判定：执行程序不符合标准审计程序"
    assert sev == "高"
