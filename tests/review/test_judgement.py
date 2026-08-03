import json

import openpyxl
import pytest
from langchain_core.messages import AIMessage

from review.evidence import build_evidence_graph
from review.judgement import build_judgement_requests, execute_judgement_requests
from review.models import AttachmentFile
from review.policy import load_policy_pack


def _workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA-4c"
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws["A5"] = "审计期间获取系统用户清单并检查管理员权限。"
    ws["B5"] = "我们访谈管理员，获取用户清单截图，并核对管理员权限。"
    return wb


def _attachments():
    item = AttachmentFile(
        index="1",
        rel_dir="SA-4c",
        filename="user-list.txt",
        rel_path="SA-4c/user-list.txt",
        file_type="text",
        extracted_text="admin,管理员,权限清单",
        extraction_status="ok",
    )
    return {"items": [item], "by_sheet_norm": {"SA4C": [item]}}


def test_build_judgement_requests_are_bounded_and_stable():
    wb = _workbook()
    graph = build_evidence_graph(wb, source_sha256="a" * 64)
    pack = load_policy_pack(pack_id="itgc-judgement", version="1.0.0")

    first = build_judgement_requests(
        workbook=wb,
        evidence_graph=graph,
        policy_pack=pack,
        sheets="SA-4c",
        attachments=_attachments(),
    )
    second = build_judgement_requests(
        workbook=wb,
        evidence_graph=graph,
        policy_pack=pack,
        sheets="SA-4c",
        attachments=_attachments(),
    )

    assert {request.rule_id for request in first} == {
        "itgc.judgement.evidence_step_alignment",
        "itgc.judgement.procedure_correspondence",
    }
    assert [request.request_id for request in first] == [
        request.request_id for request in second
    ]
    assert all(request.evidence for request in first)
    assert all(
        evidence.evidence_id.startswith(("ev:", "att:"))
        for request in first
        for evidence in request.evidence
    )
    assert all("entire workbook" not in request.question for request in first)


class _RequestAwareLLM:
    def __init__(self, *, invalid=False):
        self.invalid = invalid
        self.calls = 0

    def bind(self, **kwargs):
        return self

    async def ainvoke(self, messages):
        self.calls += 1
        request = json.loads(messages[-1].content)
        if self.invalid:
            payload = {
                "decision": "contradicted",
                "conclusion": "证据引用无法验证",
                "evidence_refs": [
                    {
                        "evidence_id": "ev:not-in-request",
                        "quote": "编造内容",
                        "start_offset": 0,
                        "end_offset": 4,
                        "role": "supporting",
                    }
                ],
                "reasoning_summary": ["引用不在白名单中"],
            }
        else:
            evidence = request["evidence"][0]
            quote = evidence["quote"]
            payload = {
                "decision": "contradicted",
                "conclusion": "执行描述未能充分支持该审计步骤。",
                "evidence_refs": [
                    {
                        "evidence_id": evidence["evidence_id"],
                        "quote": quote,
                        "start_offset": 0,
                        "end_offset": len(quote),
                        "content_hash": evidence["content_hash"],
                        "role": "supporting",
                    }
                ],
                "reasoning_summary": ["引用了请求中的原始证据"],
            }
        return AIMessage(content=json.dumps(payload, ensure_ascii=False))


@pytest.mark.asyncio
async def test_execute_judgement_requests_verifies_exact_references():
    wb = _workbook()
    graph = build_evidence_graph(wb, source_sha256="b" * 64)
    requests = build_judgement_requests(
        workbook=wb,
        evidence_graph=graph,
        policy_pack=load_policy_pack(pack_id="itgc-judgement", version="1.0.0"),
        sheets="SA-4c",
        attachments=_attachments(),
    )
    llm = _RequestAwareLLM()

    results = await execute_judgement_requests(requests, llm=llm)

    assert len(results) == len(requests)
    assert all(result.decision == "contradicted" for result in results)
    assert all(result.verification_status == "contradicted" for result in results)
    assert all(result.evidence_refs_v2[0]["end_offset"] > 0 for result in results)


@pytest.mark.asyncio
async def test_invalid_references_retry_once_then_downgrade_to_unknown():
    wb = _workbook()
    graph = build_evidence_graph(wb, source_sha256="c" * 64)
    requests = build_judgement_requests(
        workbook=wb,
        evidence_graph=graph,
        policy_pack=load_policy_pack(pack_id="itgc-judgement", version="1.0.0"),
        sheets="SA-4c",
        attachments=_attachments(),
    )[:1]
    llm = _RequestAwareLLM(invalid=True)

    results = await execute_judgement_requests(requests, llm=llm)

    assert llm.calls == 2
    assert results[0].decision == "insufficient"
    assert results[0].verification_status == "invalid"
    assert results[0].unknown_reason
    assert "evidence" in " ".join(results[0].errors).lower()
