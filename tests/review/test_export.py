import io
import json
import pytest
import openpyxl
from fastapi.testclient import TestClient

from main import app
from review.export import generate_findings_xlsx
from review.contracts import ReviewManifest
from storage.review_artifact_store import ReviewArtifactStore


@pytest.fixture
def client():
    return TestClient(app)


def test_generate_findings_xlsx_includes_all_columns():
    findings = [{
        "issue_type": "问题A",
        "severity": "P0",
        "severity_display": "高",
        "sheet": "SA-1",
        "cell": "C5",
        "risk_type": "一致性",
        "status": "fail",
        "conclusion": "结论",
        "basis": "依据",
        "suggestion": "建议",
        "evidence_refs": [{"sheet": "SA-1", "cell_or_range": "C5", "excerpt": "原文"}],
        "cross_validate_issues": ["矛盾1"],
        "llm_status": "pass",
        "llm_comment": "复核说明",
        "unknown_reason": "",
    }]
    data = generate_findings_xlsx(findings)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.title == "审阅发现汇总"
    headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
    assert headers[0] == "序号"
    assert headers[3] == "问题类型"
    assert ws.cell(row=2, column=4).value == "问题A"
    assert ws.cell(row=2, column=5).value == "P0 / 高"


def test_export_findings_returns_xlsx(client, monkeypatch, tmp_path):
    monkeypatch.setenv("WORKSPACE_PATH", str(tmp_path))
    results_dir = tmp_path / "assets" / "results"
    results_dir.mkdir(parents=True)
    payload = {
        "review_id": "r123",
        "created_at": "2026-08-06T00:00:00",
        "source": "test.xlsx",
        "stats": {"total_findings": 1, "by_severity": {"P0": 1}},
        "findings": [{
            "issue_type": "问题A", "severity": "P0", "severity_display": "高",
            "sheet": "SA-1", "cell": "C5", "risk_type": "一致性", "status": "fail",
            "conclusion": "结论", "basis": "依据", "suggestion": "建议",
            "evidence_refs": [], "cross_validate_issues": [],
        }],
    }
    (results_dir / "r123_findings.json").write_text(json.dumps(payload), encoding="utf-8")

    res = client.get("/findings/r123/export?format=xlsx")
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "findings_r123.xlsx" in res.headers["content-disposition"]
    # Should be a valid xlsx
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert wb.active.title == "审阅发现汇总"
    summary_values = [
        wb["审阅运行摘要"].cell(row=row, column=2).value
        for row in range(2, wb["审阅运行摘要"].max_row + 1)
    ]
    assert "r123" in summary_values


def test_export_route_includes_execution_manifest_metadata(client, monkeypatch, tmp_path):
    monkeypatch.setenv("WORKSPACE_PATH", str(tmp_path))
    results_dir = tmp_path / "assets" / "results"
    results_dir.mkdir(parents=True)
    (results_dir / "r-manifest_findings.json").write_text(
        json.dumps(
            {
                "review_id": "r-manifest",
                "stats": {"total_findings": 1, "by_severity": {"P1": 1}},
                "findings": [{"issue_type": "问题A", "evidence_refs": []}],
            }
        ),
        encoding="utf-8",
    )
    ReviewArtifactStore().begin(
        ReviewManifest(
            review_id="r-manifest",
            source="test.xlsx",
            input_set_sha256="input-set-route",
            execution_sha256="execution-route",
            engine_version="engine-route",
        )
    )

    response = client.get("/findings/r-manifest/export?format=xlsx")

    assert response.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    values = [
        workbook["输入与运行清单"].cell(row=row, column=3).value
        for row in range(2, workbook["输入与运行清单"].max_row + 1)
    ]
    assert "input-set-route" in values
    assert "execution-route" in values


def test_export_findings_missing_returns_404(client, monkeypatch, tmp_path):
    monkeypatch.setenv("WORKSPACE_PATH", str(tmp_path))
    res = client.get("/findings/notexist/export?format=xlsx")
    assert res.status_code == 404


def test_export_findings_explicit_stage_c_shadow_uses_candidate_artifact(
    client, monkeypatch, tmp_path
):
    monkeypatch.setenv("WORKSPACE_PATH", str(tmp_path))
    results_dir = tmp_path / "assets" / "results"
    results_dir.mkdir(parents=True)
    payload = {
        "review_id": "r-shadow",
        "created_at": "2026-08-11T00:00:00",
        "source": "test.xlsx",
        "stats": {"total_findings": 0, "by_severity": {}},
        "findings": [],
    }
    (results_dir / "r-shadow_findings.json").write_text(
        json.dumps(payload), encoding="utf-8"
    )
    store = ReviewArtifactStore()
    store.begin(ReviewManifest(review_id="r-shadow", source="test.xlsx"))
    store.write_v2_findings(
        "r-shadow",
        {
            "schema_version": "stage-c-v2-findings/1",
            "source_sha256": "sha-shadow",
            "findings": [
                {
                    "finding_id": "finding:v2-1",
                    "identity_key": "judgement:1",
                    "issue_type": "候选问题",
                    "severity": "P1",
                    "risk_type": "证据不足",
                    "sheet": "SA-1",
                    "cell": "C5",
                    "status": "unknown",
                    "decision": "insufficient",
                    "verification_status": "insufficient",
                    "basis": "shadow basis",
                    "suggestion": "shadow suggestion",
                    "evidence_refs_v2": [],
                }
            ],
            "stats": {"total_findings": 1},
        },
    )
    store.complete("r-shadow")

    res = client.get("/findings/r-shadow/export?format=xlsx&source=stage_c_shadow")

    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert wb["审阅发现汇总"].cell(row=2, column=4).value == "候选问题"
    summary_values = [
        wb["审阅运行摘要"].cell(row=row, column=2).value
        for row in range(2, wb["审阅运行摘要"].max_row + 1)
    ]
    assert "stage_c_shadow" in summary_values


def test_export_findings_missing_stage_c_shadow_returns_conflict(
    client, monkeypatch, tmp_path
):
    monkeypatch.setenv("WORKSPACE_PATH", str(tmp_path))
    results_dir = tmp_path / "assets" / "results"
    results_dir.mkdir(parents=True)
    (results_dir / "r-no-shadow_findings.json").write_text(
        json.dumps(
            {
                "review_id": "r-no-shadow",
                "findings": [{"issue_type": "V1"}],
                "stats": {},
            }
        ),
        encoding="utf-8",
    )

    res = client.get(
        "/findings/r-no-shadow/export?format=xlsx&source=stage_c_shadow"
    )

    assert res.status_code == 409


EXPECTED_HEADERS = [
    "序号", "Sheet", "单元格", "问题类型", "严重级别", "风险类型",
    "状态", "结论", "判定依据", "整改建议", "证据引用",
    "交叉校验问题", "LLM 复核状态", "LLM 复核说明", "不确定原因",
]


def test_generate_findings_xlsx_uses_exact_15_headers():
    findings = [{
        "issue_type": "占位",
        "evidence_refs": [],
    }]
    data = generate_findings_xlsx(findings)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
    assert headers == EXPECTED_HEADERS


def test_generate_findings_xlsx_preserves_attachment_in_evidence_refs():
    findings = [{
        "issue_type": "附件支撑缺失",
        "evidence_refs": [{
            "sheet": "",
            "cell_or_range": "",
            "attachment": "attachments/contracts/contract-001.pdf",
            "excerpt": "合同条款摘录：付款周期…",
        }, {
            "sheet": "SA-2",
            "cell_or_range": "B7",
            "excerpt": "底稿单元格摘录",
        }],
    }]
    data = generate_findings_xlsx(findings)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    evidence_cell = ws.cell(row=2, column=11).value
    # Must preserve both refs as JSON text, including the attachment path.
    parsed = [json.loads(line) for line in evidence_cell.splitlines() if line]
    assert any(
        r.get("attachment") == "attachments/contracts/contract-001.pdf" for r in parsed
    )
    assert any(r.get("sheet") == "SA-2" and r.get("cell_or_range") == "B7" for r in parsed)
    # Chinese characters must NOT be escaped to \uXXXX.
    assert "\\u" not in evidence_cell
    assert "付款周期" in evidence_cell


def test_export_findings_empty_list_returns_404(client, monkeypatch, tmp_path):
    monkeypatch.setenv("WORKSPACE_PATH", str(tmp_path))
    results_dir = tmp_path / "assets" / "results"
    results_dir.mkdir(parents=True)
    payload = {
        "review_id": "r-empty",
        "created_at": "2026-08-06T00:00:00",
        "source": "test.xlsx",
        "stats": {"total_findings": 0, "by_severity": {}},
        "findings": [],
    }
    (results_dir / "r-empty_findings.json").write_text(json.dumps(payload), encoding="utf-8")

    res = client.get("/findings/r-empty/export?format=xlsx")
    assert res.status_code == 404


def test_generate_findings_xlsx_adds_quality_and_provenance_sheets_compatibly():
    findings = [{
        "issue_type": "覆盖性",
        "severity": "P1",
        "severity_display": "中",
        "sheet": "SA-1",
        "cell": None,
        "risk_type": "覆盖性",
        "status": "fail",
        "conclusion": "结论",
        "basis": "依据",
        "suggestion": "补充完整范围清单",
        "evidence_refs": [
            {"sheet": "SA-1", "cell_or_range": "C5", "excerpt": "不应出现在已验证表"},
            {"attachment": "foreign.txt", "excerpt": "拒绝引用"},
        ],
        "quality": {
            "schema_version": "review-quality/1",
            "finding_id": "legacy:f1",
            "primary_location": {
                "source_kind": "cell",
                "sheet": "SA-1",
                "cell_or_range": "C5",
                "evidence_id": "cell:1",
            },
            "citation_validation": {
                "status": "partial",
                "verified_count": 1,
                "rejected_count": 1,
                "rejection_codes": ["out_of_scope_source"],
                "evidence_ids": ["cell:1"],
                "verified_refs": [{
                    "evidence_id": "cell:1",
                    "source_kind": "cell",
                    "sheet": "SA-1",
                    "cell_or_range": "C5",
                    "excerpt": "已验证摘录",
                    "source_ref": "workpaper:SA-1!C5",
                    "source_sha256": "sha-1",
                    "content_hash": "hash-1",
                    "start_offset": 0,
                    "end_offset": 5,
                }],
            },
            "gates": {
                "deterministic_cross_check": {"status": "passed", "issues": []},
                "model_re_review": {"status": "not_run", "reason": "same-model"},
                "adversarial_challenge": {"status": "not_run", "reason": "P0 only"},
            },
            "provenance": {
                "input_sha256": "sha-1",
                "engine_version": "engine-1",
                "policy_pack": None,
            },
            "grouping": {
                "root_cause_id": "root:r1",
                "duplicate_of": None,
                "related_finding_ids": [],
            },
            "remediation": {
                "status": "actionable",
                "action": "补充完整范围清单",
                "required_evidence": ["范围清单"],
                "acceptance_criteria": ["范围可复核"],
                "missing_fields": [],
            },
        },
    }]

    data = generate_findings_xlsx(
        findings,
        report_metadata={
            "review_id": "r-quality",
            "created_at": "2026-08-11T00:00:00",
            "source": "sample.xlsx",
            "stats": {
                "quality": {
                    "mode": "shadow",
                    "raw_findings": 1,
                    "canonical_findings": 1,
                    "duplicate_findings": 0,
                    "root_cause_count": 1,
                }
            },
        },
    )
    wb = openpyxl.load_workbook(io.BytesIO(data))

    assert wb.sheetnames == [
        "审阅发现汇总",
        "审阅运行摘要",
        "审阅质量摘要",
        "证据溯源明细",
        "输入与运行清单",
        "审阅一致性与冲突",
    ]
    legacy_ws = wb["审阅发现汇总"]
    assert legacy_ws.cell(row=2, column=3).value == "C5"

    quality_ws = wb["审阅质量摘要"]
    quality_headers = [quality_ws.cell(row=1, column=c).value for c in range(1, 18)]
    assert "引用校验状态" in quality_headers
    assert quality_ws.cell(row=2, column=2).value == "legacy:f1"
    assert "partial" in [quality_ws.cell(row=2, column=c).value for c in range(1, 18)]

    provenance_ws = wb["证据溯源明细"]
    assert provenance_ws.max_row == 3
    provenance_values = [
        provenance_ws.cell(row=row, column=column).value
        for row in range(2, provenance_ws.max_row + 1)
        for column in range(1, provenance_ws.max_column + 1)
    ]
    assert "cell:1" in provenance_values
    assert "foreign.txt" not in str(provenance_values)


def test_generate_findings_xlsx_handles_legacy_payload_without_quality_envelope():
    data = generate_findings_xlsx(
        [{"issue_type": "历史发现", "sheet": "SA-1", "cell": "C5"}],
        report_metadata={"review_id": "legacy"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(data))

    assert wb["审阅质量摘要"].cell(row=2, column=7).value == "not_available"
    summary_values = [
        wb["审阅运行摘要"].cell(row=row, column=2).value
        for row in range(2, wb["审阅运行摘要"].max_row + 1)
    ]
    assert any("历史结果" in str(value) for value in summary_values)


def test_generate_findings_xlsx_exports_quality_v2_run_identity_and_conflicts():
    findings = [
        {
            "finding_id": "finding:quality-1",
            "issue_type": "附件内容无法支持结论",
            "severity": "P1",
            "sheet": "SA-1",
            "cell": "C5",
            "status": "fail",
            "assertion_id": "attachment.content.support",
            "claim_type": "attachment_content",
            "claim_subject": "SA-1|attachment:contract.pdf",
            "claim_value": "unsupported",
            "evidence_refs": [
                {"attachment": "untrusted.pdf", "excerpt": "不得导出的拒绝摘录"}
            ],
            "quality": {
                "schema_version": "review-quality/2",
                "finding_id": "finding:quality-1",
                "citation_validation": {
                    "status": "partial",
                    "verified_count": 1,
                    "rejected_count": 1,
                    "rejection_codes": ["out_of_scope_source"],
                    "verified_refs": [
                        {
                            "evidence_id": "attachment:1",
                            "source_kind": "attachment",
                            "attachment": "contract.pdf",
                            "excerpt": "已接受的附件摘录",
                            "source_sha256": "source-sha",
                            "content_hash": "content-sha",
                        }
                    ],
                },
                "claim_support": {
                    "status": "partial",
                    "supporting_evidence_ids": ["attachment:1"],
                    "missing_requirements": ["attachment_content"],
                    "reason_codes": ["attachment_required"],
                },
                "consistency": {
                    "status": "conflicted",
                    "conflict_ids": ["conflict:1"],
                    "related_finding_ids": ["finding:quality-2"],
                    "reason_codes": ["exclusive_claim_values"],
                },
                "provenance": {
                    "input_sha256": "input-sha",
                    "input_set_sha256": "input-set-sha",
                    "execution_sha256": "execution-sha",
                    "engine_version": "review-engine/2",
                    "assertion_catalog": {
                        "id": "review-quality",
                        "version": "1.0.0",
                    },
                },
                "remediation": {
                    "status": "actionable",
                    "action": "补充原始合同附件",
                    "required_evidence": ["冻结附件原件"],
                    "acceptance_criteria": ["结论可定位到附件摘录"],
                    "missing_fields": [],
                },
            },
        }
    ]
    data = generate_findings_xlsx(
        findings,
        report_metadata={
            "review_id": "r-quality-v2",
            "quality_stats": {"mode": "shadow", "total_findings": 1},
            "conflicts": [
                {
                    "conflict_id": "conflict:1",
                    "assertion_id": "attachment.content.support",
                    "claim_subject": "SA-1|attachment:contract.pdf",
                    "finding_ids": ["finding:quality-1", "finding:quality-2"],
                    "values": ["present", "absent"],
                    "status": "unresolved",
                }
            ],
            "manifest": {
                "inputs": [
                    {
                        "role": "workpaper",
                        "path": "/private/server/workpaper.xlsx",
                        "filename": "workpaper.xlsx",
                        "sha256": "workpaper-sha",
                        "size": 42,
                        "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    }
                ],
                "requested_sheets": ["SA-1"],
                "engine_version": "review-engine/2",
                "input_set_sha256": "input-set-sha",
                "execution_sha256": "execution-sha",
                "runtime_config": {
                    "review_model": "review-model-a",
                    "review_temperature": 0.2,
                    "quality_mode": "shadow",
                    "judgement_mode": "off",
                    "secret": "must-not-export",
                },
                "components": [
                    {
                        "component_id": "review-quality-remediation",
                        "version": "1.0.0",
                        "sha256": "component-sha",
                    }
                ],
            },
        },
    )

    wb = openpyxl.load_workbook(io.BytesIO(data))
    legacy_headers = [
        wb["审阅发现汇总"].cell(row=1, column=column).value
        for column in range(1, 16)
    ]
    assert legacy_headers == EXPECTED_HEADERS
    assert wb.sheetnames[-2:] == ["输入与运行清单", "审阅一致性与冲突"]

    quality_ws = wb["审阅质量摘要"]
    quality_columns = {
        quality_ws.cell(row=1, column=column).value: column
        for column in range(1, quality_ws.max_column + 1)
    }
    assert quality_ws.cell(row=2, column=quality_columns["assertion_id"]).value == (
        "attachment.content.support"
    )
    assert quality_ws.cell(row=2, column=quality_columns["Claim 类型"]).value == (
        "attachment_content"
    )
    assert quality_ws.cell(row=2, column=quality_columns["结论支持状态"]).value == "partial"
    assert quality_ws.cell(row=2, column=quality_columns["一致性状态"]).value == "conflicted"
    assert quality_ws.cell(row=2, column=quality_columns["冲突编号"]).value == "conflict:1"
    assert quality_ws.cell(row=2, column=quality_columns["输入集SHA256"]).value == "input-set-sha"
    assert quality_ws.cell(row=2, column=quality_columns["执行SHA256"]).value == "execution-sha"

    provenance_ws = wb["证据溯源明细"]
    provenance_headers = {
        provenance_ws.cell(row=1, column=column).value: column
        for column in range(1, provenance_ws.max_column + 1)
    }
    provenance_values = [
        [provenance_ws.cell(row=row, column=column).value for column in range(1, provenance_ws.max_column + 1)]
        for row in range(2, provenance_ws.max_row + 1)
    ]
    assert any(
        row[provenance_headers["验证状态"] - 1] == "accepted"
        and row[provenance_headers["支持该声明"] - 1] == "是"
        and "已接受的附件摘录" in row
        for row in provenance_values
    )
    assert any(
        row[provenance_headers["验证状态"] - 1] == "rejected"
        and row[provenance_headers["拒绝码"] - 1] == "out_of_scope_source"
        for row in provenance_values
    )
    assert "不得导出的拒绝摘录" not in str(provenance_values)

    execution_ws = wb["输入与运行清单"]
    execution_values = [
        execution_ws.cell(row=row, column=3).value
        for row in range(2, execution_ws.max_row + 1)
    ]
    assert "input-set-sha" in execution_values
    assert "execution-sha" in execution_values
    assert "review-model-a" in execution_values
    assert "/private/server/workpaper.xlsx" not in str(execution_values)
    assert "must-not-export" not in str(execution_values)

    conflicts_ws = wb["审阅一致性与冲突"]
    assert conflicts_ws.cell(row=2, column=1).value == "conflict:1"
