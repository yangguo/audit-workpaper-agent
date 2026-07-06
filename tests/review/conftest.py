import openpyxl
import pytest


@pytest.fixture
def blank_workbook():
    return openpyxl.Workbook()


@pytest.fixture
def layout_workbook():
    """A sheet with a standard layout: header row 1, standard col A, exec col B."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "标准审计程序"
    ws["B1"] = "执行审计程序"
    ws["A2"] = "获取系统用户清单并检查权限。"
    ws["B2"] = "我们导出用户清单，截图保存。"
    return wb
