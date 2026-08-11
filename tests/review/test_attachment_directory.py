import openpyxl
from review.models import AttachmentFile

from review.attachments import (
    _attachments_context_for_sheet,
    _check_attachment_references,
    _extract_attachment_refs,
    _match_attachment_items,
    _verify_attachment_evidence_refs,
    build_attachment_index,
)


def test_build_attachment_index_recurses_and_extracts_text(tmp_path):
    root = tmp_path / "attachments"
    evidence_dir = root / "SA-4c"
    evidence_dir.mkdir(parents=True)
    evidence = evidence_dir / "附件1-用户清单.txt"
    evidence.write_text("用户名：admin\n权限：管理员", encoding="utf-8")

    index = build_attachment_index(str(root))

    assert index["source_type"] == "directory"
    assert len(index["items"]) == 1
    item = index["items"][0]
    assert item.rel_path == "SA-4c/附件1-用户清单.txt"
    assert item.extracted_text == "用户名：admin\n权限：管理员"
    assert item.extraction_status == "ok"
    assert item in index["by_filename"]["附件1-用户清单.txt"]
    assert item in index["by_index"]["1"]
    assert item in index["by_sheet_norm"]["SA4C"]


def test_match_attachment_items_accepts_index_and_relative_path_suffix(tmp_path):
    root = tmp_path / "attachments"
    evidence_dir = root / "bundle" / "SA-4c"
    evidence_dir.mkdir(parents=True)
    evidence = evidence_dir / "附件1-user-list.csv"
    evidence.write_text("user,role\nadm,admin", encoding="utf-8")
    index = build_attachment_index(str(root))

    filenames, rel_paths, indices = _extract_attachment_refs(
        "见附件1，文件为 SA-4c/附件1-user-list.csv"
    )
    matched, missing = _match_attachment_items(
        index,
        filenames=filenames,
        rel_paths=rel_paths,
        indices=indices,
    )

    assert len(matched) == 1
    assert matched[0].filename == "附件1-user-list.csv"
    assert missing == []


def test_match_attachment_items_rejects_parent_path_reference(tmp_path):
    root = tmp_path / "attachments"
    root.mkdir()
    evidence = root / "evidence.txt"
    evidence.write_text("admin,administrator", encoding="utf-8")
    index = build_attachment_index(str(root))

    matched, missing = _match_attachment_items(
        index,
        filenames=[],
        rel_paths=["../evidence.txt"],
        indices=[],
    )

    assert matched == []
    assert missing == ["../evidence.txt"]


def test_build_attachment_index_reads_xlsx_evidence(tmp_path):
    root = tmp_path / "attachments"
    root.mkdir()
    path = root / "permission-matrix.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "账号"
    workbook.active["B1"] = "角色"
    workbook.active["A2"] = "admin"
    workbook.active["B2"] = "管理员"
    workbook.save(path)

    index = build_attachment_index(str(root))

    item = index["items"][0]
    assert "admin" in item.extracted_text
    assert "管理员" in item.extracted_text


def test_build_attachment_index_keeps_unsupported_files_as_metadata(tmp_path):
    root = tmp_path / "attachments"
    root.mkdir()
    path = root / "export.rar"
    path.write_bytes(b"not parsed")

    index = build_attachment_index(str(root))

    assert index["items"][0].filename == "export.rar"
    assert index["items"][0].extraction_status == "unsupported"
    assert index["items"][0].extracted_text == ""


def test_attachment_context_contains_matched_file_content(tmp_path):
    root = tmp_path / "attachments"
    root.mkdir()
    (root / "user-list.txt").write_text("admin,administrator", encoding="utf-8")
    index = build_attachment_index(str(root))

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "SA-4c"
    sheet["A1"] = "执行程序"
    sheet["A2"] = "导出用户清单，见 user-list.txt"

    context = _attachments_context_for_sheet(sheet, index)

    assert "user-list.txt" in context
    assert "admin,administrator" in context


def test_attachment_context_contains_validated_agent_evidence():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "SA-4c"
    sheet["A1"] = "执行程序"
    attachments = {
        "items": [],
        "by_filename": {},
        "by_rel_path": {},
        "by_index": {},
        "by_sheet_norm": {},
        "agent_evidence_by_sheet": {
            "SA4C": [{
                "path": "SA-4c/user-list.txt",
                "file_type": "txt",
                "extraction_status": "ok",
                "excerpt": "admin,administrator",
                "supports": "用户清单中的管理员权限",
                "confidence": "high",
            }],
        },
    }

    context = _attachments_context_for_sheet(sheet, attachments)

    assert "SA-4c/user-list.txt" in context
    assert "admin,administrator" in context


def test_verify_attachment_evidence_refs_drops_unverified_attachment_hint():
    item = AttachmentFile(
        index="1", rel_dir="SA-4c", filename="users.txt",
        rel_path="SA-4c/users.txt", file_type="txt", status="ok",
        extraction_status="ok", extracted_text="admin,administrator", size=19,
    )
    attachments = {
        "path": "/pinned/attachments",
        "items": [item],
        "by_filename": {"users.txt": [item]},
        "by_rel_path": {"sa-4c/users.txt": [item]},
        "by_index": {"1": [item]},
        "by_sheet_norm": {},
    }
    refs = _verify_attachment_evidence_refs([
        {"sheet": "SA-4c", "cell_or_range": "B5", "attachment": "SA-4c/users.txt", "excerpt": "admin,administrator"},
        {"sheet": "SA-4c", "cell_or_range": "B5", "attachment": "missing.txt", "excerpt": "编造"},
        {"sheet": "SA-4c", "cell_or_range": "B5", "attachment": "users.txt", "excerpt": "编造"},
    ], attachments)

    assert refs[0]["attachment"] == "SA-4c/users.txt"
    assert "attachment" not in refs[1]
    assert "attachment" not in refs[2]


def test_verify_attachment_evidence_refs_keeps_attachment_excerpt_when_cell_is_valid():
    item = AttachmentFile(
        index="1", rel_dir="SA-4c", filename="users.txt",
        rel_path="SA-4c/users.txt", file_type="txt", status="ok",
        extraction_status="ok", extracted_text="admin,administrator", size=19,
    )
    attachments = {
        "items": [item],
        "by_filename": {"users.txt": [item]},
        "by_rel_path": {"sa-4c/users.txt": [item]},
        "by_index": {"1": [item]},
        "by_sheet_norm": {},
    }
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["B5"] = "见附件1并核验用户权限"

    refs = _verify_attachment_evidence_refs([
        {
            "sheet": "SA-4c",
            "cell_or_range": "B5",
            "attachment": "SA-4c/users.txt",
            "excerpt": "admin,administrator",
        }
    ], attachments, ws=sheet)

    assert refs[0]["attachment"] == "SA-4c/users.txt"
    assert refs[0]["excerpt"] == "admin,administrator"


def test_verify_attachment_evidence_refs_accepts_verified_ocr_excerpt():
    item = AttachmentFile(
        index="1", rel_dir="SA-4c", filename="screenshot.png",
        rel_path="SA-4c/screenshot.png", file_type="png", status="binary",
        extraction_status="binary", extracted_text="", size=19,
    )
    attachments = {
        "items": [item],
        "by_filename": {"screenshot.png": [item]},
        "by_rel_path": {"sa-4c/screenshot.png": [item]},
        "by_index": {"1": [item]},
        "by_sheet_norm": {},
        "ocr_by_path": {
            "sa-4c/screenshot.png": {
                "status": "ok",
                "provider": "mineru-lightweight",
                "content": "用户 admin 具有管理员权限",
            },
        },
    }

    refs = _verify_attachment_evidence_refs([
        {
            "sheet": "SA-4c",
            "cell_or_range": "B5",
            "attachment": "SA-4c/screenshot.png",
            "excerpt": "用户 admin 具有管理员权限",
        }
    ], attachments)

    assert refs[0]["attachment"] == "SA-4c/screenshot.png"


def test_check_attachment_references_does_not_flag_successful_ocr_as_unparsed():
    item = AttachmentFile(
        index="1", rel_dir="SA-4c", filename="screenshot.webp",
        rel_path="SA-4c/screenshot.webp", file_type="webp", status="binary",
        extraction_status="binary", extracted_text="", size=19,
    )
    attachments = {
        "items": [item],
        "by_filename": {"screenshot.webp": [item]},
        "by_rel_path": {"sa-4c/screenshot.webp": [item]},
        "by_index": {"1": [item]},
        "by_sheet_norm": {},
        "ocr_by_path": {
            "sa-4c/screenshot.webp": {
                "status": "ok",
                "content": "用户 admin 具有管理员权限",
            },
        },
    }
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "SA-4c"
    sheet["B5"] = "见附件1"

    assert _check_attachment_references("SA-4c", sheet, attachments) == []


def test_build_attachment_index_rejects_missing_directory(tmp_path):
    assert build_attachment_index(str(tmp_path / "missing")) == {}


def test_build_attachment_index_includes_docx_embedded_images(tmp_path):
    from review.attachments import build_attachment_index
    # Set up attachments dir with a docx containing an image
    att_dir = tmp_path / "atts"
    att_dir.mkdir()
    docx = att_dir / "report.docx"
    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("word/document.xml", b"<?xml version='1.0'?><doc/>")
        zf.writestr("word/media/picture.png", b"\x89PNG\r\n\x1a\n" + b"\x00" * 4)
    docx.write_bytes(buf.getvalue())

    idx = build_attachment_index(str(att_dir))
    # Virtual attachment should be present
    virtual = [it for it in idx["items"] if "embedded_media" in it.rel_path]
    assert virtual, "expected virtual attachment from embedded image"
    assert any(v.file_type == "png" for v in virtual)


def test_build_attachment_index_continues_after_bad_docx(tmp_path):
    """A single corrupted office document must not abort extraction for others."""
    from review.attachments import build_attachment_index
    att_dir = tmp_path / "atts"
    att_dir.mkdir()

    # Valid docx with an embedded image.
    good_docx = att_dir / "good.docx"
    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("word/document.xml", b"<?xml version='1.0'?><doc/>")
        zf.writestr("word/media/picture.png", b"\x89PNG\r\n\x1a\n" + b"\x00" * 4)
    good_docx.write_bytes(buf.getvalue())

    # Invalid docx that is not a zip archive.
    bad_docx = att_dir / "bad.docx"
    bad_docx.write_bytes(b"this is not a zip file")

    idx = build_attachment_index(str(att_dir))

    # Both source files are indexed as real attachments.
    real_items = [it for it in idx["items"] if not it.rel_path.startswith(".embedded_media/")]
    assert {it.filename for it in real_items} == {"good.docx", "bad.docx"}

    # The valid docx still yields a virtual embedded-media item.
    virtual = [it for it in idx["items"] if it.rel_path.startswith(".embedded_media/")]
    assert any(it.rel_path == ".embedded_media/good.docx::picture.png" for it in virtual)

    # The on-disk embedded file exists.
    assert (att_dir / ".embedded_media" / "good.docx__picture.png").is_file()


def test_build_attachment_index_cleans_stale_embedded_media(tmp_path):
    """Pre-existing .embedded_media files must not be indexed as real attachments."""
    from review.attachments import build_attachment_index
    att_dir = tmp_path / "atts"
    att_dir.mkdir()

    good_docx = att_dir / "good.docx"
    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("word/document.xml", b"<?xml version='1.0'?><doc/>")
        zf.writestr("word/media/picture.png", b"\x89PNG\r\n\x1a\n" + b"\x00" * 4)
    good_docx.write_bytes(buf.getvalue())

    # Seed a stale extracted file from a previous run.
    embedded_root = att_dir / ".embedded_media"
    embedded_root.mkdir()
    (embedded_root / "old.docx__stale.png").write_bytes(b"stale")

    idx = build_attachment_index(str(att_dir))

    # The stale on-disk file is removed.
    assert not (embedded_root / "old.docx__stale.png").exists()

    # Only the current source doc's virtual item remains.
    virtual = [it for it in idx["items"] if it.rel_path.startswith(".embedded_media/")]
    assert [it.rel_path for it in virtual] == [".embedded_media/good.docx::picture.png"]

    # No .embedded_media/ file is indexed as a real attachment.
    real_rels = [it.rel_path for it in idx["items"] if not it.rel_path.startswith(".embedded_media/")]
    assert all(".embedded_media" not in rel for rel in real_rels)


def test_extract_attachment_text_handles_legacy_via_converter(tmp_path, monkeypatch):
    """Legacy .xls/.doc should be routed through the converter when available."""
    from review import attachments, legacy_convert

    class _FakeLegacy:
        def convert_legacy_to_modern(self, src_path, dest_dir=None):
            out = tmp_path / (src_path.stem + ".xlsx")
            out.write_bytes(b"fake xlsx")
            return out

    def _fake_read_xlsx(_path):
        return "converted content", "ok"

    monkeypatch.setattr(
        legacy_convert, "convert_legacy_to_modern", _FakeLegacy().convert_legacy_to_modern
    )
    monkeypatch.setattr(attachments, "_read_xlsx_file", _fake_read_xlsx)

    fake = tmp_path / "old.xls"
    fake.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")  # OLE/CFB magic
    text, status = attachments._extract_attachment_text(fake)
    # Either ok or unavailable depending on the patched _read_xlsx_file; main goal:
    # the legacy branch was entered (no "unsupported" without calling converter)
    assert status != "unsupported" or text != ""
    assert text == "converted content"
    assert status == "ok"
