"""
底稿解析工具 - 用于读取和分析 Excel 审计底稿
"""
import os
import json
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass

import openpyxl
from openpyxl.utils import get_column_letter

from langchain.tools import tool


@dataclass(frozen=True)
class WorksheetInfo:
    """工作表信息"""
    sheet_name: str
    total_rows: int
    total_cols: int
    layout_info: Dict[str, Any]
    summary: str


def _is_empty(value) -> bool:
    """检查值是否为空"""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _get_cell_value(ws, cell_ref: str) -> Optional[str]:
    """获取单元格值，处理合并单元格"""
    cell = ws[cell_ref]
    value = cell.value
    if value is None and ws.merged_cells.ranges:
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                break
    if _is_empty(value):
        return None
    return str(value).strip()


def _detect_layout(ws) -> Tuple[Optional[int], int, List[int]]:
    """检测表格布局：标准审计程序列和执行程序列"""
    max_scan_row = min(ws.max_row or 0, 40)
    max_scan_col = min(ws.max_column or 0, 30)

    for r in range(1, max_scan_row + 1):
        standard_col = None
        exec_cols: List[int] = []
        for c in range(1, max_scan_col + 1):
            v = _get_cell_value(ws, f"{get_column_letter(c)}{r}")
            if not v:
                continue
            # 检测标准审计程序列
            if ("标准" in v and "审计程序" in v) or ("标准审计程序" in v):
                standard_col = c
            # 检测执行程序列
            if "执行" in v and "审计程序" in v:
                exec_cols.append(c)

        if standard_col and exec_cols:
            exec_cols = sorted({c for c in exec_cols if c != standard_col})
            if exec_cols:
                return r, standard_col, exec_cols

    return None, 0, []


def _extract_sheet_text_cells(ws, limit: int = 1000) -> List[Tuple[str, str]]:
    """提取工作表中的所有文本单元格"""
    result = []
    count = 0
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if _is_empty(cell.value):
                continue
            value = cell.value
            if isinstance(value, str):
                text = value.strip()
            else:
                text = str(value).strip()
            if not text:
                continue
            result.append((cell.coordinate, text))
            count += 1
            if count >= limit:
                return result
    return result


@tool
def analyze_worksheet(file_path: str) -> str:
    """
    分析 Excel 审计底稿工作表，提取结构和内容信息

    Args:
        file_path: Excel 文件路径（相对于 assets 目录的路径，或绝对路径）

    Returns:
        JSON 字符串，包含工作表的布局、内容和摘要信息
    """
    workspace_path = os.getenv("COZE_WORKSPACE_PATH", os.getcwd())
    if not os.path.isabs(file_path):
        full_path = os.path.join(workspace_path, file_path)
    else:
        full_path = file_path

    if not os.path.exists(full_path):
        return json.dumps({
            "error": f"文件不存在: {full_path}",
            "success": False
        }, ensure_ascii=False)

    try:
        # 读取 Excel 文件
        wb = openpyxl.load_workbook(full_path, data_only=True)

        all_sheets_info = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 检测布局
            header_row, standard_col, exec_cols = _detect_layout(ws)

            # 提取文本内容
            text_cells = _extract_sheet_text_cells(ws, limit=500)

            # 构建布局信息
            layout_info = {
                "header_row": header_row,
                "standard_column": standard_col,
                "standard_column_letter": get_column_letter(standard_col) if standard_col else None,
                "exec_columns": exec_cols,
                "exec_columns_letters": [get_column_letter(c) for c in exec_cols] if exec_cols else [],
            }

            # 生成摘要
            if standard_col and exec_cols:
                summary = f"检测到标准审计程序列（{get_column_letter(standard_col)}）和{len(exec_cols)}个执行程序列（{', '.join(get_column_letter(c) for c in exec_cols)}），表头在第{header_row}行"
            else:
                summary = f"未检测到标准的审计程序布局，共{ws.max_row}行{ws.max_column}列数据"

            # 提取审计程序内容（如果有标准列）
            audit_programs = []
            if standard_col and exec_cols:
                for r in range(header_row + 1, min(header_row + 51, ws.max_row + 1)):
                    std_text = _get_cell_value(ws, f"{get_column_letter(standard_col)}{r}")
                    if std_text:
                        exec_texts = []
                        for exec_c in exec_cols:
                            exec_text = _get_cell_value(ws, f"{get_column_letter(exec_c)}{r}")
                            if exec_text:
                                exec_texts.append({
                                    "column": get_column_letter(exec_c),
                                    "content": exec_text[:200]  # 限制长度
                                })
                        if exec_texts:
                            audit_programs.append({
                                "row": r,
                                "standard": std_text[:200],
                                "executions": exec_texts
                            })

            sheet_info = WorksheetInfo(
                sheet_name=sheet_name,
                total_rows=ws.max_row or 0,
                total_cols=ws.max_column or 0,
                layout_info=layout_info,
                summary=summary
            )

            all_sheets_info.append({
                "sheet_name": sheet_info.sheet_name,
                "total_rows": sheet_info.total_rows,
                "total_cols": sheet_info.total_cols,
                "layout_info": sheet_info.layout_info,
                "summary": sheet_info.summary,
                "sample_text_cells": text_cells[:20],  # 返回前20个单元格示例
                "audit_programs_sample": audit_programs[:10],  # 返回前10条审计程序示例
                "total_text_cells": len(text_cells)
            })

        result = {
            "success": True,
            "file_path": full_path,
            "total_sheets": len(all_sheets_info),
            "sheets": all_sheets_info
        }

        return json.dumps(result, ensure_ascii=False, indent=2)

    except Exception as e:
        error_msg = f"分析 Excel 文件失败: {str(e)}"
        return json.dumps({
            "error": error_msg,
            "success": False,
            "exception_type": type(e).__name__
        }, ensure_ascii=False)
