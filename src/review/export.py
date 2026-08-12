"""Export review findings to structured report formats."""

from __future__ import annotations

import io
import json
from typing import Any, Dict, Iterable, List, Mapping

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LEGACY_HEADERS = [
    "序号", "Sheet", "单元格", "问题类型", "严重级别", "风险类型",
    "状态", "结论", "判定依据", "整改建议", "证据引用",
    "交叉校验问题", "LLM 复核状态", "LLM 复核说明", "不确定原因",
]

QUALITY_HEADERS = [
    "序号", "finding_id", "问题类型", "Sheet", "主证据类型", "主证据定位",
    "引用校验状态", "已验证引用数", "拒绝引用数", "交叉校验状态",
    "模型复核状态", "对抗挑战状态", "根因编号", "重复于", "整改状态",
    "整改缺口", "输入SHA256",
    "assertion_id", "Claim 类型", "Claim 对象", "Claim 值", "结论支持状态",
    "支持证据ID", "支持缺口", "支持原因", "一致性状态", "冲突编号",
    "关联发现", "输入集SHA256", "执行SHA256", "断言目录版本",
]

PROVENANCE_HEADERS = [
    "序号", "发现序号", "finding_id", "证据ID", "来源类型", "Sheet",
    "单元格/范围", "相对来源", "逐字摘录", "来源SHA256", "内容Hash",
    "起始偏移", "结束偏移", "验证状态", "拒绝码", "支持该声明",
    "拒绝引用数",
]

EXECUTION_HEADERS = ["类别", "字段", "值", "说明"]

CONFLICT_HEADERS = ["冲突编号", "assertion_id", "Claim 对象", "关联发现", "互斥值", "处理状态"]

_RUNTIME_CONFIG_FIELDS = (
    "review_model",
    "review_endpoint_sha256",
    "review_temperature",
    "review_json_mode",
    "verify_ssl",
    "quality_mode",
    "deterministic_crosscheck_mode",
    "evidence_agent_mode",
    "evidence_snapshot_max_cells",
    "challenger_full_text",
    "mineru_ocr_mode",
    "mineru_ocr_language",
    "mineru_model_version",
    "policy_mode",
    "judgement_mode",
    "judgement_max_requests",
    "prompt_bundle_version",
)

_HEADER_FILL = PatternFill(
    start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"
)
_HEADER_FONT = Font(bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _quality(finding: Mapping[str, Any]) -> dict[str, Any]:
    value = finding.get("quality")
    return dict(value) if isinstance(value, Mapping) else {}


def _list_text(value: Any, separator: str = "；") -> str:
    if isinstance(value, list):
        return separator.join(_text(item) for item in value if _text(item))
    return _text(value)


def _write_table(
    ws,
    headers: list[str],
    rows: Iterable[Iterable[Any]],
    *,
    widths: list[int] | None = None,
) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row_number, values in enumerate(rows, start=2):
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_number, column=col, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if widths:
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    if ws.max_column and ws.max_row:
        ws.auto_filter.ref = ws.dimensions


def _primary_cell(finding: Mapping[str, Any]) -> str:
    cell = _text(finding.get("cell"))
    if cell:
        return cell
    quality = _quality(finding)
    location = quality.get("primary_location")
    if isinstance(location, Mapping) and _text(location.get("source_kind")) == "cell":
        return _text(location.get("cell_or_range"))
    return ""


def _legacy_rows(findings: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for idx, finding in enumerate(findings, start=1):
        severity = _text(finding.get("severity"))
        severity_display = _text(finding.get("severity_display"))
        severity_str = f"{severity} / {severity_display}" if severity_display else severity
        evidence_refs = finding.get("evidence_refs", [])
        if isinstance(evidence_refs, list):
            evidence_str = "\n".join(
                json.dumps(ref, ensure_ascii=False) for ref in evidence_refs
            )
        else:
            evidence_str = str(evidence_refs)
        rows.append(
            [
                idx,
                finding.get("sheet", ""),
                _primary_cell(finding),
                finding.get("issue_type", ""),
                severity_str,
                finding.get("risk_type", ""),
                finding.get("status", ""),
                finding.get("llm_conclusion") or finding.get("conclusion") or "",
                finding.get("basis", ""),
                finding.get("suggestion", ""),
                evidence_str,
                _list_text(finding.get("cross_validate_issues", [])),
                finding.get("llm_status", ""),
                finding.get("llm_comment", ""),
                finding.get("unknown_reason", ""),
            ]
        )
    return rows


def _gate_status(quality: Mapping[str, Any], name: str) -> str:
    gates = quality.get("gates")
    gate = gates.get(name) if isinstance(gates, Mapping) else None
    return _text(gate.get("status")) if isinstance(gate, Mapping) else "not_available"


def _mapping(value: Any) -> Mapping[str, Any]:
    return value if isinstance(value, Mapping) else {}


def _quality_stats(report_metadata: Mapping[str, Any]) -> Mapping[str, Any]:
    direct = report_metadata.get("quality_stats")
    if isinstance(direct, Mapping):
        return direct
    stats = report_metadata.get("stats")
    quality = stats.get("quality") if isinstance(stats, Mapping) else None
    return quality if isinstance(quality, Mapping) else {}


def _report_manifest(report_metadata: Mapping[str, Any]) -> Mapping[str, Any]:
    manifest = report_metadata.get("manifest")
    return manifest if isinstance(manifest, Mapping) else {}


def _catalog_version(provenance: Mapping[str, Any]) -> str:
    catalog = provenance.get("assertion_catalog")
    if not isinstance(catalog, Mapping):
        return ""
    pack_id = _text(catalog.get("id"))
    version = _text(catalog.get("version"))
    if pack_id and version:
        return f"{pack_id}@{version}"
    return version or pack_id


def _quality_rows(findings: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for idx, finding in enumerate(findings, start=1):
        quality = _quality(finding)
        location = quality.get("primary_location")
        citation = quality.get("citation_validation")
        grouping = quality.get("grouping")
        remediation = quality.get("remediation")
        claim_support = quality.get("claim_support")
        consistency = quality.get("consistency")
        provenance = _mapping(quality.get("provenance"))
        rows.append(
            [
                idx,
                _text(quality.get("finding_id") or finding.get("finding_id")),
                finding.get("issue_type", ""),
                finding.get("sheet", ""),
                _text(location.get("source_kind")) if isinstance(location, Mapping) else "not_available",
                (
                    f"{_text(location.get('sheet'))}!{_text(location.get('cell_or_range'))}"
                    if isinstance(location, Mapping) and _text(location.get("cell_or_range"))
                    else _text(location.get("source_ref")) if isinstance(location, Mapping) else ""
                ),
                _text(citation.get("status")) if isinstance(citation, Mapping) else "not_available",
                citation.get("verified_count", 0) if isinstance(citation, Mapping) else 0,
                citation.get("rejected_count", 0) if isinstance(citation, Mapping) else 0,
                _gate_status(quality, "deterministic_cross_check"),
                _gate_status(quality, "model_re_review"),
                _gate_status(quality, "adversarial_challenge"),
                _text(grouping.get("root_cause_id")) if isinstance(grouping, Mapping) else "",
                _text(grouping.get("duplicate_of")) if isinstance(grouping, Mapping) else "",
                _text(remediation.get("status")) if isinstance(remediation, Mapping) else "not_available",
                _list_text(remediation.get("missing_fields")) if isinstance(remediation, Mapping) else "",
                _text(provenance.get("input_sha256")),
                _text(finding.get("assertion_id")),
                _text(finding.get("claim_type")),
                _text(finding.get("claim_subject")),
                _text(finding.get("claim_value")),
                _text(claim_support.get("status"))
                if isinstance(claim_support, Mapping) else "not_available",
                _list_text(claim_support.get("supporting_evidence_ids"))
                if isinstance(claim_support, Mapping) else "",
                _list_text(claim_support.get("missing_requirements"))
                if isinstance(claim_support, Mapping) else "",
                _list_text(claim_support.get("reason_codes"))
                if isinstance(claim_support, Mapping) else "",
                _text(consistency.get("status"))
                if isinstance(consistency, Mapping) else "not_available",
                _list_text(consistency.get("conflict_ids"))
                if isinstance(consistency, Mapping) else "",
                _list_text(consistency.get("related_finding_ids"))
                if isinstance(consistency, Mapping) else "",
                _text(provenance.get("input_set_sha256")),
                _text(provenance.get("execution_sha256")),
                _catalog_version(provenance),
            ]
        )
    return rows


def _provenance_rows(findings: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    row_number = 1
    for finding_number, finding in enumerate(findings, start=1):
        quality = _quality(finding)
        citation = quality.get("citation_validation")
        refs = citation.get("verified_refs") if isinstance(citation, Mapping) else []
        if not isinstance(refs, list):
            refs = []
        claim_support = quality.get("claim_support")
        supporting_ids = {
            _text(value)
            for value in claim_support.get("supporting_evidence_ids", [])
            if _text(value)
        } if isinstance(claim_support, Mapping) else set()
        has_claim_support = isinstance(claim_support, Mapping)
        finding_id = _text(quality.get("finding_id") or finding.get("finding_id"))
        for ref in refs:
            if not isinstance(ref, Mapping):
                continue
            rows.append(
                [
                    row_number,
                    finding_number,
                    finding_id,
                    _text(ref.get("evidence_id")),
                    _text(ref.get("source_kind")),
                    _text(ref.get("sheet")),
                    _text(ref.get("cell_or_range")),
                    _text(ref.get("source_ref") or ref.get("attachment")),
                    _text(ref.get("excerpt") or ref.get("quote")),
                    _text(ref.get("source_sha256")),
                    _text(ref.get("content_hash")),
                    ref.get("start_offset", ""),
                    ref.get("end_offset", ""),
                    "accepted",
                    "",
                    (
                        "是" if _text(ref.get("evidence_id")) in supporting_ids
                        else "否" if has_claim_support else "不可用"
                    ),
                    "",
                ]
            )
            row_number += 1
        rejected_count = citation.get("rejected_count", 0) if isinstance(citation, Mapping) else 0
        try:
            rejected_count = max(0, int(rejected_count or 0))
        except (TypeError, ValueError):
            rejected_count = 0
        rejection_codes = (
            [_text(code) for code in citation.get("rejection_codes", []) if _text(code)]
            if isinstance(citation, Mapping) and isinstance(citation.get("rejection_codes"), list)
            else []
        )
        for index in range(max(rejected_count, len(rejection_codes))):
            # Rejected refs deliberately omit source, locator and excerpt.
            rows.append(
                [
                    row_number,
                    finding_number,
                    finding_id,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "rejected",
                    rejection_codes[index] if index < len(rejection_codes) else "未提供拒绝码",
                    "否",
                    rejected_count,
                ]
            )
            row_number += 1
    return rows


def _summary_rows(
    findings: list[dict[str, Any]], report_metadata: Mapping[str, Any]
) -> list[list[Any]]:
    quality_stats = _quality_stats(report_metadata)
    has_quality = any(isinstance(finding.get("quality"), Mapping) for finding in findings)
    values = [
        ("报告schema版本", "review-report/2"),
        ("review_id", _text(report_metadata.get("review_id"))),
        ("来源", _text(report_metadata.get("source"))),
        ("生成时间", _text(report_metadata.get("created_at"))),
        ("质量模式", _text(quality_stats.get("mode")) or ("shadow" if has_quality else "历史结果")),
        ("输入SHA256", _text(quality_stats.get("input_sha256"))),
        ("引擎版本", _text(quality_stats.get("engine_version"))),
        ("原始发现数", quality_stats.get("raw_findings", len(findings))),
        ("可合并后发现数", quality_stats.get("canonical_findings", len(findings))),
        ("重复发现数", quality_stats.get("duplicate_findings", 0)),
        ("根因数", quality_stats.get("root_cause_count", 0)),
        ("质量信封", "可用" if has_quality else "历史结果，质量信封不可用"),
    ]
    return [[key, value] for key, value in values]


def _first_quality_provenance(findings: list[dict[str, Any]]) -> Mapping[str, Any]:
    for finding in findings:
        provenance = _quality(finding).get("provenance")
        if isinstance(provenance, Mapping):
            return provenance
    return {}


def _execution_rows(
    findings: list[dict[str, Any]], report_metadata: Mapping[str, Any]
) -> list[list[Any]]:
    manifest = _report_manifest(report_metadata)
    quality_stats = _quality_stats(report_metadata)
    provenance = _first_quality_provenance(findings)
    rows: list[list[Any]] = []
    input_set_sha256 = (
        _text(manifest.get("input_set_sha256"))
        or _text(quality_stats.get("input_set_sha256"))
        or _text(provenance.get("input_set_sha256"))
    )
    execution_sha256 = (
        _text(manifest.get("execution_sha256"))
        or _text(quality_stats.get("execution_sha256"))
        or _text(provenance.get("execution_sha256"))
    )
    engine_version = (
        _text(manifest.get("engine_version"))
        or _text(quality_stats.get("engine_version"))
        or _text(provenance.get("engine_version"))
    )
    rows.extend(
        [
            ["运行身份", "input_set_sha256", input_set_sha256, "冻结输入集标识"],
            ["运行身份", "execution_sha256", execution_sha256, "输入与运行选择的联合标识"],
            ["运行身份", "engine_version", engine_version, "审阅引擎版本"],
            [
                "运行范围",
                "requested_sheets",
                _list_text(manifest.get("requested_sheets")),
                "请求审阅的工作表",
            ],
        ]
    )
    inputs = manifest.get("inputs")
    if isinstance(inputs, list):
        for item in inputs:
            if not isinstance(item, Mapping):
                continue
            role = _text(item.get("role"))
            filename = _text(item.get("filename"))
            detail = "；".join(
                value for value in (
                    f"size={item.get('size')}" if item.get("size") is not None else "",
                    _text(item.get("media_type")),
                ) if value
            )
            rows.append(
                ["输入文件", role or "unknown", _text(item.get("sha256")), f"{filename}；{detail}".strip("；")]
            )
    runtime_config = manifest.get("runtime_config")
    if isinstance(runtime_config, Mapping):
        for key in _RUNTIME_CONFIG_FIELDS:
            value = runtime_config.get(key)
            if isinstance(value, (str, int, float, bool)):
                rows.append(["运行配置", key, value, "受控运行参数"])
    components = manifest.get("components")
    if isinstance(components, list):
        for item in components:
            if not isinstance(item, Mapping):
                continue
            component_id = _text(item.get("component_id"))
            version = _text(item.get("version"))
            sha256 = _text(item.get("sha256"))
            if component_id or version or sha256:
                rows.append(["组件", component_id, sha256, version])
    return rows


def _report_conflicts(report_metadata: Mapping[str, Any]) -> list[Mapping[str, Any]]:
    direct = report_metadata.get("conflicts")
    if isinstance(direct, list):
        return [item for item in direct if isinstance(item, Mapping)]
    consistency = _quality_stats(report_metadata).get("consistency")
    conflicts = consistency.get("conflicts") if isinstance(consistency, Mapping) else []
    return [item for item in conflicts if isinstance(item, Mapping)] if isinstance(conflicts, list) else []


def _conflict_rows(report_metadata: Mapping[str, Any]) -> list[list[Any]]:
    return [
        [
            _text(conflict.get("conflict_id")),
            _text(conflict.get("assertion_id")),
            _text(conflict.get("claim_subject")),
            _list_text(conflict.get("finding_ids")),
            _list_text(conflict.get("values")),
            _text(conflict.get("status")) or "unresolved",
        ]
        for conflict in _report_conflicts(report_metadata)
    ]


def generate_findings_xlsx(
    findings: List[Dict[str, Any]],
    report_metadata: Mapping[str, Any] | None = None,
) -> bytes:
    """Render a compatible multi-sheet review package as an .xlsx workbook."""

    metadata = dict(report_metadata or {})
    normalized_findings = [dict(item) for item in findings if isinstance(item, Mapping)]
    wb = openpyxl.Workbook()
    wb.properties.title = "审阅结果报告"
    wb.properties.subject = "V1 findings with review-quality provenance"

    legacy_ws = wb.active
    legacy_ws.title = "审阅发现汇总"
    _write_table(
        legacy_ws,
        LEGACY_HEADERS,
        _legacy_rows(normalized_findings),
        widths=[6, 12, 12, 35, 12, 12, 10, 35, 45, 45, 45, 30, 15, 30, 30],
    )

    summary_ws = wb.create_sheet("审阅运行摘要")
    _write_table(summary_ws, ["指标", "值"], _summary_rows(normalized_findings, metadata), widths=[28, 80])

    quality_ws = wb.create_sheet("审阅质量摘要")
    _write_table(
        quality_ws,
        QUALITY_HEADERS,
        _quality_rows(normalized_findings),
        widths=[6, 38, 30, 12, 14, 28, 14, 14, 14, 16, 16, 16, 34, 34, 22, 30, 68, 38, 20, 44, 28, 18, 34, 30, 30, 18, 34, 34, 68, 68, 30],
    )

    provenance_ws = wb.create_sheet("证据溯源明细")
    _write_table(
        provenance_ws,
        PROVENANCE_HEADERS,
        _provenance_rows(normalized_findings),
        widths=[6, 10, 38, 42, 14, 14, 16, 48, 60, 68, 68, 14, 14, 14, 28, 16, 14],
    )

    execution_ws = wb.create_sheet("输入与运行清单")
    _write_table(
        execution_ws,
        EXECUTION_HEADERS,
        _execution_rows(normalized_findings, metadata),
        widths=[16, 34, 72, 48],
    )

    conflicts_ws = wb.create_sheet("审阅一致性与冲突")
    _write_table(
        conflicts_ws,
        CONFLICT_HEADERS,
        _conflict_rows(metadata),
        widths=[34, 38, 48, 68, 40, 18],
    )

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
