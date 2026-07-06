"""Rule-based procedure-pair and sheet-scope checks + A-C LLM correspondence
(ported from analyze_excel.py)."""
import asyncio
import json
import re
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl.utils import get_column_letter

from review.constants import EVIDENCE_KEYWORDS, INTERVIEW_ONLY_KEYWORDS, OS_DB_KEYWORDS
from review.excel_utils import _detect_layout, _extract_sheet_text_cells, _get_cell_value, _truncate
from review.llm import _llm_chat, _llm_stat, _try_parse_json
from review.models import Finding, _SEVERITY_FROM_CHINESE, _EXCERPT_MAX_LEN
from review.validation import _verify_evidence_refs


def _likely_interview_only(execution_text: str) -> bool:
    t = execution_text or ""
    if not t:
        return False
    has_interview = any(k in t for k in INTERVIEW_ONLY_KEYWORDS)
    has_evidence = any(k in t for k in EVIDENCE_KEYWORDS)
    return has_interview and not has_evidence


def _requires_evidence_by_standard(standard_text: str) -> Sequence[str]:
    required: List[str] = []
    t = standard_text or ""
    if any(k in t for k in ("截图", "参数", "配置")):
        required.append("截图/参数界面")
    if any(k in t for k in ("导出", "清单", "用户清单", "权限清单")):
        required.append("导出清单")
    if any(k in t for k in ("日志", "台账", "变更日志", "变更台账", "任务清单")):
        required.append("日志/台账/清单")
    if any(k in t for k in ("审批", "授权", "批准")):
        required.append("审批/授权证据")
    if any(k in t for k in ("协议", "合同", "供应商")):
        required.append("协议/合同条款")
    return required


def _check_procedure_pairs(ws_title: str, ws) -> List[Finding]:
    findings: List[Finding] = []
    header_row, standard_col, execution_cols = _detect_layout(ws)
    if header_row is None or standard_col <= 0 or not execution_cols:
        return findings
    start_row = max(5, (header_row or 1) + 2)

    skip_a_exact = {"n/a", "na", "不适用", "有效", "无"}
    skip_a_symbol = {"√", "×", "✓", "✗"}
    skip_c_exact = {"n/a", "na", "不适用"}
    keywords_like_procedure = (
        "询问", "访谈", "检查", "获取", "抽取", "观察", "复核", "比对", "重新执行", "分析", "确认", "审查",
    )
    is_design_section = False
    if header_row and standard_col > 0:
        header_text = _get_cell_value(ws, f"{get_column_letter(standard_col)}{header_row}") or ""
        is_design_section = "设计有效性" in header_text

    empty_streak = 0
    for row in range(start_row, (ws.max_row or 0) + 1):
        row_marker = _get_cell_value(ws, f"A{row}")
        if row_marker:
            m = row_marker.strip()
            if (m.startswith("*") and m[1:].isdigit()) or (m.startswith("#") and m[1:].isdigit()):
                continue
        a_text = _get_cell_value(ws, f"{get_column_letter(standard_col)}{row}")
        has_exec = any(_get_cell_value(ws, f"{get_column_letter(c)}{row}") for c in execution_cols)
        if a_text is None and not has_exec:
            empty_streak += 1
            if empty_streak >= 30:
                break
            continue
        empty_streak = 0

        if a_text is None:
            continue

        a_compact = a_text.replace(" ", "").replace("\n", "").strip()
        if a_compact in skip_a_symbol or a_compact.lower() in skip_a_exact:
            continue
        if len(a_compact) <= 10 and all(ch.isalnum() or ch in "-_./" for ch in a_compact):
            continue
        if not any(k in a_text for k in keywords_like_procedure) and "。" not in a_text and "•" not in a_text:
            continue

        required_evidence = _requires_evidence_by_standard(a_text)
        for exec_col in execution_cols:
            c_cell = f"{get_column_letter(exec_col)}{row}"
            c_text = _get_cell_value(ws, c_cell)
            if c_text is None:
                continue

            row_marker = _get_cell_value(ws, f"A{row}")
            if row_marker:
                m = row_marker.strip()
                if (m.startswith("*") and m[1:].isdigit()) or (m.startswith("#") and m[1:].isdigit()):
                    continue

            c_compact = c_text.replace(" ", "").replace("\n", "").strip()
            if c_compact.lower() in skip_c_exact:
                continue
            if len(c_compact) <= 12 and all(ch.isalnum() or ch in "-_./" for ch in c_compact):
                continue
            if len(a_text) < 20 or len(c_text) < 20:
                continue

            execution_like = (
                ("我们" in c_text)
                or any(k in c_text for k in INTERVIEW_ONLY_KEYWORDS)
                or any(k in c_text for k in EVIDENCE_KEYWORDS)
            )
            if not execution_like:
                template_like = (
                    "•" in c_text
                    or "被认为" in c_text
                    or c_text.strip().startswith(("如果", "以下", "在以下", "根据", "用户参与", "文档包括"))
                )
                if template_like:
                    findings.append(
                        Finding(
                            issue_type="执行列疑似未替换模板/未按要求填列",
                            severity="P0",
                            sheet=ws_title,
                            cell=c_cell,
                            snippet=_truncate(c_text, 220),
                            basis="执行列内容更像标准模板/判定口径（如「如果/以下/被认为」及大量条款），缺少「我们获取/检查/抽样/复核」等实际执行描述。",
                            suggestion="将该单元格补充为实际执行步骤与获取证据描述（含样本框定方法、样本来源/编号、证据链接/截图/导出）。",
                        )
                    )
                continue

            if _likely_interview_only(c_text):
                findings.append(
                    Finding(
                        issue_type="程序执行不到位/仅依赖访谈",
                        severity="P1",
                        sheet=ws_title,
                        cell=c_cell,
                        snippet=_truncate(c_text, 220),
                        basis="执行描述出现“访谈/询问/了解”等，但未体现截图、导出清单、日志、台账、审批等实质性证据。",
                        suggestion="补充系统截图/导出清单/日志台账/审批记录等，并在执行程序中明确证据来源与核查步骤。",
                    )
                )

            if required_evidence and not any(k in c_text for k in EVIDENCE_KEYWORDS):
                findings.append(
                    Finding(
                        issue_type="证据类型缺失",
                        severity="P1",
                        sheet=ws_title,
                        cell=c_cell,
                        snippet=_truncate(c_text, 220),
                        basis=f"标准审计程序要求获取/检查证据（{', '.join(required_evidence)}），但执行描述未体现对应证据。",
                        suggestion="对照标准程序逐条补齐证据（截图/清单/日志/审批/协议等），并在底稿中保留可复核的原始文件或路径。",
                    )
                )

            if any(k in a_text for k in ("账号新增", "新增账号", "开通")):
                if "入职" in c_text and not any(k in c_text for k in ("账号创建", "创建时间", "创建日期", "用户清单", "跨期比对")):
                    findings.append(
                        Finding(
                            issue_type="账号新增样本总量基准可能有误",
                            severity="P1",
                            sheet=ws_title,
                            cell=c_cell,
                            snippet=_truncate(c_text, 220),
                            basis="常见问题：样本总量应优先以用户清单“账号创建时间”为基准；仅以入职名单抽样可能遗漏外包/延期开户等。",
                            suggestion="优先获取系统用户清单含账号创建时间字段进行抽样；无该字段时，采用用户清单跨期比对+入职名单交叉验证组合确定样本总量。",
                        )
                    )

            if any(k in a_text for k in ("离职", "禁用", "删除")):
                if "已禁用" in c_text and ("抽样" in c_text or "样本" in c_text) and not any(k in c_text for k in ("离职名单", "全量", "关联", "匹配", "用户清单")):
                    findings.append(
                        Finding(
                            issue_type="离职账号禁用检查方法可能有误",
                            severity="P1",
                            sheet=ws_title,
                            cell=c_cell,
                            snippet=_truncate(c_text, 220),
                            basis="常见问题：不应从“已禁用账户”反向抽样，应将离职名单与用户清单全量关联核查账号状态与禁用时间。",
                            suggestion="获取审计期间离职名单，与系统用户清单关联，全量核查账号状态/禁用时间，必要时补充禁用工单或审批证据。",
                        )
                    )

            if "调岗" in a_text or "岗位变动" in a_text:
                if not any(k in c_text for k in ("权限变更", "权限调整", "角色调整", "禁用", "撤销", "变更")):
                    findings.append(
                        Finding(
                            issue_type="未覆盖调岗权限变更/禁用测试",
                            severity="P1",
                            sheet=ws_title,
                            cell=c_cell,
                            snippet=_truncate(c_text, 220),
                            basis="常见问题：调岗应纳入权限变更/禁用控制测试，仅看账号状态不足以覆盖权限调整实质性测试。",
                            suggestion="获取调岗人员名单与用户清单关联，框定调岗且持有账号人员范围，按调岗前后岗位权限差异抽样核查权限变更/禁用证据。",
                        )
                    )

            if any(k in a_text for k in ("密码策略", "密码", "复杂度", "锁定", "时效")):
                if is_design_section:
                    if not any(k in c_text for k in ("制度", "规程", "政策", "流程", "规定", "办法", "指引", "《", "<")):
                        findings.append(
                            Finding(
                                issue_type="设计有效性证据不足（密码策略）",
                                severity="P2",
                                sheet=ws_title,
                                cell=c_cell,
                                snippet=_truncate(c_text, 220),
                                basis="设计有效性测试通常以制度/流程/政策文件作为证据；当前执行描述未体现已获取/引用相关文件。",
                                suggestion="补充引用密码策略相关制度/规程/政策文件（文件名、条款、编号/链接），并在底稿中说明其适用系统与覆盖范围。",
                            )
                        )
                elif not any(k in c_text for k in ("截图", "参数", "配置", "界面")):
                    findings.append(
                        Finding(
                            issue_type="密码策略证据有效性不足",
                            severity="P1",
                            sheet=ws_title,
                            cell=c_cell,
                            snippet=_truncate(c_text, 220),
                            basis="常见问题：仅文字说明不足以支撑密码策略；应留存参数界面/配置截图等实质性证据。",
                            suggestion="补充密码策略参数界面截图（复杂度/锁定/有效期/历史密码等），并核对底稿描述与系统配置一致。",
                        )
                    )

            if any(k in a_text for k in ("批处理", "定时任务", "任务计划", "job", "Job")):
                if _likely_interview_only(c_text) or not any(k in c_text for k in ("任务", "日志", "清单", "导出")):
                    findings.append(
                        Finding(
                            issue_type="批处理作业证据不足/范围可能未覆盖",
                            severity="P1",
                            sheet=ws_title,
                            cell=c_cell,
                            snippet=_truncate(c_text, 220),
                            basis="常见问题：仅访谈了解批处理设置不充分；应获取任务清单/日志，并覆盖应用层、操作系统、数据库层面。",
                            suggestion="补充应用/OS/DB层面任务清单与执行日志导出，并明确是否覆盖全部相关批处理作业。",
                        )
                    )

            if any(k in a_text for k in ("变更", "发布", "上线", "迁移")):
                if _likely_interview_only(c_text) or not any(k in c_text for k in ("变更台账", "变更日志", "台账", "日志", "工单")):
                    findings.append(
                        Finding(
                            issue_type="系统变更证据不足/样本框定可能有误",
                            severity="P1",
                            sheet=ws_title,
                            cell=c_cell,
                            snippet=_truncate(c_text, 220),
                            basis="常见问题：应以变更日志/变更台账为样本总量基准抽样审批；仅依赖访谈或从审批流程反向框定无法验证“变更均经审批”。",
                            suggestion="获取变更日志/台账作为总体，按期间抽样追溯审批与测试/上线证据；补充工单、发布记录、回滚记录等。",
                        )
                    )

    return findings


def _check_sheet_scope(ws_title: str, ws) -> List[Finding]:
    findings: List[Finding] = []
    sheet_text = " ".join(text for _, text in _extract_sheet_text_cells(ws))

    if ws_title in {"SA-4c"}:
        if ("管理员" in sheet_text or "特权" in sheet_text) and not any(k in sheet_text for k in OS_DB_KEYWORDS):
            findings.append(
                Finding(
                    issue_type="特权账号识别范围可能不完整",
                    severity="P1",
                    sheet=ws_title,
                    cell=None,
                    snippet=_truncate(sheet_text, 220),
                    basis="常见问题：项目检查范围未覆盖操作系统与数据库层面的管理员账号设置情况，可能导致特权账号识别不完整。",
                    suggestion="补充获取并核对OS/DB层面管理员账号清单（或截图/导出），并评估与应用层管理员职责冲突与共享风险。",
                )
            )

    if ws_title in {"PM-5", "PM-6", "PM-4b", "PM-4c", "PM-4e", "PM-3"}:
        if "供应商" in sheet_text and not any(k in sheet_text for k in ("协议", "合同", "SaaS", "托管", "权利义务")):
            findings.append(
                Finding(
                    issue_type="供应商托管场景证据可能不足",
                    severity="P1",
                    sheet=ws_title,
                    cell=None,
                    snippet=_truncate(sheet_text, 220),
                    basis="常见问题：供应商维护/托管时需获取相关协议文件检查双方权利义务与实际履行情况。",
                    suggestion="补充协议/合同/运维报告/工单等证据，明确管理员账号归属（租户级权限）与监控复核责任。",
                )
            )

    return findings


# ---------------------------------------------------------------------------
# A-C correspondence: LLM judgement of standard vs execution programs
# ---------------------------------------------------------------------------

async def _llm_judge_procedure_pair(
    *,
    llm,
    standard_text: str,
    execution_text: str,
) -> Tuple[bool, Optional[bool], str, str]:
    """Judge whether execution satisfies the standard program.

    Returns (success, is_match, reason, raw). is_match True=符合, False=不符合,
    None=不确定. success=False means the LLM call itself failed.
    """
    def _clip(text, limit):
        s = (text or "").strip()
        if len(s) <= limit:
            return s
        return s[:limit] + "..."

    system_prompt_json = (
        "你是一名严格的审计质量复核专家。\n"
        "你的任务：对比A列的【标准审计程序描述】与C列的【实际执行程序】，判断执行是否满足标准的控制意图与关键要求。\n\n"
        "重要说明（用于降低误报）：\n"
        "1) 标准描述通常是“规范流程/参考口径”，不要求与执行描述在措辞、人名、部门称谓、文件标题完全一致。\n"
        "2) 访谈对象：若执行访谈对象属于同职能部门/同职责岗位，且能覆盖标准要验证的控制点，可视为符合；仅当职责明显不匹配或关键岗位未覆盖时才判不符合。\n"
        "3) 检查文件：若执行检查的制度/规范/流程文件与标准要求主题一致、适用范围一致或更高层级覆盖，可视为符合；仅当文件主题不相关或未覆盖关键控制要素时才判不符合。\n"
        "4) 结论表述：允许文字概括，但必须覆盖标准中的关键条件/核查点；若仅泛泛表述而未触及关键条件，则应判不符合或不确定。\n"
        "5) 无发生/不适用：若执行描述明确说明审计期间内该事项/活动未发生（如无迁移/无开发项目/无重大变更等），且给出总体为0或无发生的依据（例如项目清单/变更台账/发布记录/日志导出等），则可判pass并在理由中说明“不适用/总体为0”；若仅口头说明且缺少依据，则判unknown。\n\n"
        "判断规则：\n"
        "1. 标准描述中要求的审计动作，实际执行中是否包含\n"
        "2. 标准描述中指定的审计对象，实际执行中是否覆盖\n"
        "3. 标准描述中提出的具体条件，实际执行中是否满足\n"
        "4. 标准描述中要求获取的审计证据类型，实际执行中是否获取\n\n"
        "输出要求：必须输出严格JSON对象，格式为：\n"
        "{\"status\": \"pass\"|\"fail\"|\"unknown\", \"reason\": \"...\", "
        "\"evidence_refs\": [{\"cell\": \"...\", \"excerpt\": \"...\"}], "
        "\"unknown_reason\": \"...\"}\n"
        "evidence_refs: status=fail时必填，excerpt必须来自执行描述或标准描述的原文片段。"
        "status=unknown时unknown_reason必填（≥10字符）。"
    )
    system_prompt_old = (
        "你是一名严格的审计质量复核专家。\n"
        "你的任务：对比A列的【标准审计程序描述】与C列的【实际执行程序】，判断执行是否满足标准的控制意图与关键要求。\n\n"
        "重要说明（用于降低误报）：\n"
        "1) 标准描述通常是“规范流程/参考口径”，不要求与执行描述在措辞、人名、部门称谓、文件标题完全一致。\n"
        "2) 访谈对象：若执行访谈对象属于同职能部门/同职责岗位，且能覆盖标准要验证的控制点，可视为符合；仅当职责明显不匹配或关键岗位未覆盖时才判不符合。\n"
        "3) 检查文件：若执行检查的制度/规范/流程文件与标准要求主题一致、适用范围一致或更高层级覆盖，可视为符合；仅当文件主题不相关或未覆盖关键控制要素时才判不符合。\n"
        "4) 结论表述：允许文字概括，但必须覆盖标准中的关键条件/核查点；若仅泛泛表述而未触及关键条件，则应判不符合或不确定。\n"
        "5) 无发生/不适用：若执行描述明确说明审计期间内该事项/活动未发生（如无迁移/无开发项目/无重大变更等），且给出总体为0或无发生的依据（例如项目清单/变更台账/发布记录/日志导出等），则可判【符合】并在理由中说明“不适用/总体为0”；若仅口头说明且缺少依据，则判【不确定】。\n\n"
        "判断规则：\n"
        "1. 标准描述中要求的审计动作，实际执行中是否包含\n"
        "2. 标准描述中指定的审计对象，实际执行中是否覆盖\n"
        "3. 标准描述中提出的具体条件，实际执行中是否满足\n"
        "4. 标准描述中要求获取的审计证据类型，实际执行中是否获取\n\n"
        "请只回答：【符合】或【不符合】或【不确定】\n"
        "然后换行写【理由：】后面跟简要说明。"
    )
    user_prompt = (
        "【标准审计程序描述 - A列】：\n"
        f"{_clip(standard_text, 900)}\n\n"
        "【实际执行程序 - C列】：\n"
        f"{_clip(execution_text, 900)}\n\n"
        "请判断C列执行程序是否符合A列标准要求。"
    )

    stage = "procedure_pair"
    try:
        answer = await _llm_chat(
            llm=llm,
            messages=[
                {"role": "system", "content": system_prompt_json},
                {"role": "user", "content": user_prompt},
            ],
            stage=stage,
            max_attempts=3,
            max_tokens=512,
        )
        answer = (answer or "").strip()
        parsed = _try_parse_json(answer)
        if isinstance(parsed, dict) and "status" in parsed:
            status_val = str(parsed.get("status", "")).strip()
            if status_val == "pass":
                is_match = True
            elif status_val == "fail":
                is_match = False
            else:
                is_match = None
            reason = str(parsed.get("reason", "")).strip() or str(parsed.get("unknown_reason", "")).strip()
            return True, is_match, reason, answer
    except Exception:
        pass

    try:
        answer = await _llm_chat(
            llm=llm,
            messages=[
                {"role": "system", "content": system_prompt_old},
                {"role": "user", "content": user_prompt},
            ],
            stage=stage,
            max_attempts=2,
            max_tokens=512,
        )
        answer = (answer or "").strip()
        if "不符合" in answer:
            is_match = False
        elif "符合" in answer:
            is_match = True
        else:
            is_match = None
        if "理由：" in answer:
            reason = answer.split("理由：", 1)[1].strip()
        elif "理由:" in answer:
            reason = answer.split("理由:", 1)[1].strip()
        else:
            reason = answer
        reason = str(reason or "").strip("】 \t\n\r").strip()
        return True, is_match, reason, answer
    except Exception as e:
        return False, None, str(e), ""


async def _llm_check_procedure_pairs(
    *,
    wb,
    target_sheets: Sequence[str],
    llm,
    start_row: int = 5,
    sleep_seconds: float = 0.2,
) -> Tuple[Dict[str, object], List[Finding]]:
    skip_a_keywords = {
        "序号", "审计证据", "设计有效性测试结论", "执行有效性测试结论", "测试步骤", "抽样数量",
        "测试期间&样本总量", "样本记录（如果抽样数量>1，则在下表记录其余样本）",
        "缺陷评估", "是否发现异常", "缺陷描述", "标记注释",
    }
    skip_a_exact = {"n/a", "na", "不适用", "有效", "无"}
    skip_a_symbol = {"√", "×", "✓", "✗"}
    skip_c_exact = {"n/a", "na", "不适用"}

    def _execution_label(ws, header_row, execution_col):
        label = None
        if header_row:
            label = _get_cell_value(ws, f"{get_column_letter(execution_col)}{header_row + 1}")
            if not label:
                label = _get_cell_value(ws, f"{get_column_letter(execution_col)}{header_row}")
        return label or get_column_letter(execution_col)

    total = 0
    matched = 0
    failed = 0
    api_errors = 0
    skipped_a_empty = 0
    skipped_c_empty = 0
    skipped_ref = 0
    skipped_header = 0

    sheet_stats: Dict[str, Dict[str, int]] = {}
    result_details: List[Dict[str, Optional[str]]] = []
    issue_details: List[Dict[str, Optional[str]]] = []
    findings: List[Finding] = []

    keywords_like_procedure = ("询问", "访谈", "检查", "获取", "抽取", "观察", "复核", "比对", "重新执行", "分析", "确认", "审查")

    def _is_design_section(ws, header_row, standard_col):
        if not header_row or standard_col <= 0:
            return False
        header_text = _get_cell_value(ws, f"{get_column_letter(standard_col)}{header_row}") or ""
        return "设计有效性" in header_text

    def _looks_password_design_standard(text):
        t = (text or "").replace(" ", "").replace("\n", "").replace("\r", "")
        if not t:
            return False
        if "密码" not in t and "身份验证" not in t:
            return False
        if "设计" not in t and "设计有效性" not in t:
            return False
        return any(k in t for k in ("最短长度", "复杂", "到期", "锁定", "账户", "账号"))

    def _has_policy_evidence(text):
        t = text or ""
        return any(k in t for k in ("制度", "规程", "政策", "流程", "规定", "办法", "指引", "《", "<"))

    def _looks_no_occurrence_exec(execution_text):
        t = (execution_text or "").replace(" ", "").replace("\n", "").replace("\r", "")
        if not t:
            return False
        time_markers = ("审计期间", "本期", "期间内", "测试期间", "本年度", "年度内")
        if not any(x in t for x in time_markers):
            return False
        if "未对此控制点进行测试" in t or "故未对" in t or "不适用" in t or t.lower().find("n/a") >= 0:
            return True
        action_markers = ("迁移", "开发", "重开发", "上线", "项目", "变更", "发布", "投产", "升级", "实施", "切换", "改造", "功能重开发")
        return re.search(r"(未|无)(发生|进行|开展|实施|发生过)?(.{0,18})(" + "|".join(action_markers) + ")", t) is not None

    def _standard_looks_conditional(standard_text):
        t = (standard_text or "").replace(" ", "").replace("\n", "").replace("\r", "")
        if not t:
            return False
        markers = ("抽取", "样本", "期间", "迁移", "上线", "开发", "项目", "变更", "发布", "投产", "升级", "实施", "切换", "改造")
        return any(m in t for m in markers)

    def _classify_mismatch(standard_text, execution_text, reason):
        r = (reason or "").strip()
        r2 = r.replace(" ", "")
        if _looks_no_occurrence_exec(execution_text) and _standard_looks_conditional(standard_text):
            has_basis_evidence = any(k in (execution_text or "") for k in EVIDENCE_KEYWORDS) or any(
                k in (execution_text or "") for k in ("清单", "导出", "台账", "日志", "工单", "记录", "报告")
            )
            sev = "低" if has_basis_evidence else "中"
            return (
                "LLM判定：期间内无发生/不适用（不应按未执行判缺陷）",
                sev,
                "将该控制点按“不适用/总体为0”处理：在底稿中明确审计期间总体=0，并补充无发生依据（项目清单/变更台账/发布记录/工单/日志导出等）；如仅访谈说明，需补充可复核的清单或系统导出佐证。",
            )
        if any(k in r2 for k in ("访谈对象", "访谈人", "受访", "访谈人员")) and any(k in r2 for k in ("不一致", "不同", "不匹配")):
            return (
                "LLM判定：访谈对象/角色表述差异（可能等效）",
                "低",
                "在底稿中补充访谈对象岗位/职责与标准角色的对应关系（同部门/同职能可视为等效），并说明选择该人员的原因；如涉及关键岗位职责差异，补充关键角色访谈或邮件确认。",
            )
        if any(k in r2 for k in ("检查对象", "制度", "规范", "办法", "要求", "流程", "文件")) and any(k in r2 for k in ("不一致", "不同", "不匹配")):
            return (
                "LLM判定：制度/文件名称差异（需确认覆盖范围）",
                "低",
                "在底稿中说明所检查文件与标准要求的主题一致性（范围/适用系统/章节映射），必要时补充目录截图或关键条款摘录，证明已覆盖标准控制要点。",
            )
        if any(k in r2 for k in ("未满足", "未明确", "未确认", "未覆盖", "缺少", "不足", "没有")) and any(
            k in r2 for k in ("条件", "完善", "方案", "计划", "所有", "全部", "必须")
        ):
            return (
                "LLM判定：执行结论未覆盖标准关键条件",
                "中",
                "补充对标准关键条件的逐条确认记录（例如：抽取若干项目方案/计划验证、引用制度条款、说明样本范围/期间），并在结论中明确对应条件是否满足。",
            )
        return (
            "LLM判定：执行程序不符合标准审计程序",
            "高",
            "补充/修改执行程序以覆盖标准要求的审计动作、对象、条件与证据类型，并在底稿中保留可复核来源（截图/导出清单/日志台账/审批或协议等）。",
        )

    for sheet in target_sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header_row, standard_col, execution_cols = _detect_layout(ws)
        if not standard_col or not execution_cols:
            continue

        execution_labels = {c: _execution_label(ws, header_row, c) for c in execution_cols}
        sheet_is_design = _is_design_section(ws, header_row, standard_col)
        sheet_stats[ws.title] = {
            "total": 0, "matched": 0, "failed": 0, "api_errors": 0,
            "skipped_a_empty": 0, "skipped_c_empty": 0, "skipped_ref": 0, "skipped_header": 0,
        }

        empty_streak = 0
        for row in range(max(1, int(start_row)), (ws.max_row or 0) + 1):
            a_text = _get_cell_value(ws, f"{get_column_letter(standard_col)}{row}")
            has_any_exec = any(_get_cell_value(ws, f"{get_column_letter(c)}{row}") is not None for c in execution_cols)

            if a_text is None and not has_any_exec:
                empty_streak += 1
                if empty_streak >= 30:
                    break
                continue
            empty_streak = 0

            if a_text is None:
                skipped_a_empty += 1
                sheet_stats[ws.title]["skipped_a_empty"] += 1
                continue

            a_compact = a_text.replace(" ", "").replace("\n", "").strip()
            if a_compact in skip_a_symbol or a_compact.lower() in skip_a_exact or a_compact in skip_a_keywords:
                skipped_header += 1
                sheet_stats[ws.title]["skipped_header"] += 1
                continue
            if (a_compact.startswith("*") and a_compact[1:].isdigit()) or (a_compact.startswith("#") and a_compact[1:].isdigit()):
                skipped_header += 1
                sheet_stats[ws.title]["skipped_header"] += 1
                continue
            if len(a_compact) <= 10 and all(ch.isalnum() or ch in "-_./" for ch in a_compact):
                skipped_header += 1
                sheet_stats[ws.title]["skipped_header"] += 1
                continue

            a_for_judge = a_text.strip()
            if not any(k in a_for_judge for k in keywords_like_procedure) and "•" not in a_for_judge and "。" not in a_for_judge:
                skipped_header += 1
                sheet_stats[ws.title]["skipped_header"] += 1
                continue

            for exec_col in execution_cols:
                c_text = _get_cell_value(ws, f"{get_column_letter(exec_col)}{row}")
                exec_label = execution_labels.get(exec_col) or get_column_letter(exec_col)
                if c_text is None:
                    skipped_c_empty += 1
                    sheet_stats[ws.title]["skipped_c_empty"] += 1
                    continue

                c_compact = c_text.replace(" ", "").replace("\n", "").strip()
                if c_compact.lower() in skip_c_exact:
                    skipped_ref += 1
                    sheet_stats[ws.title]["skipped_ref"] += 1
                    continue
                if len(c_compact) <= 12 and all(ch.isalnum() or ch in "-_./" for ch in c_compact):
                    skipped_ref += 1
                    sheet_stats[ws.title]["skipped_ref"] += 1
                    continue

                c_for_judge = c_text.strip()
                if len(a_for_judge) < 20 or len(c_for_judge) < 20:
                    skipped_header += 1
                    sheet_stats[ws.title]["skipped_header"] += 1
                    continue

                total += 1
                sheet_stats[ws.title]["total"] += 1

                standard_cell = f"{get_column_letter(standard_col)}{row}"
                execution_cell = f"{get_column_letter(exec_col)}{row}"

                success, is_match, reason, raw = await _llm_judge_procedure_pair(
                    llm=llm, standard_text=a_for_judge, execution_text=c_for_judge,
                )
                if success and (is_match is False or is_match is None):
                    if sheet_is_design and _looks_password_design_standard(a_for_judge) and _has_policy_evidence(c_for_judge):
                        is_match = True
                        reason = "设计有效性：已引用/获取制度/规程等文件作为证据；参数细节（最短长度/锁定等）与覆盖范围（账户类型）可作为完善建议，不作为实施有效性不足判定。"
                    if _looks_no_occurrence_exec(c_for_judge) and _standard_looks_conditional(a_for_judge):
                        is_match = True
                        reason = "执行说明审计期间内该事项未发生/总体为0，故该控制点不适用；建议在底稿中补充总体为0的可复核依据（清单/台账/日志导出等）。"
                result_label = "不确定"
                if success:
                    if is_match is True:
                        matched += 1
                        sheet_stats[ws.title]["matched"] += 1
                        result_label = "✓ 符合"
                    elif is_match is False:
                        failed += 1
                        sheet_stats[ws.title]["failed"] += 1
                        result_label = "✗ 不符合"
                    else:
                        failed += 1
                        sheet_stats[ws.title]["failed"] += 1
                        result_label = "不确定"
                else:
                    api_errors += 1
                    sheet_stats[ws.title]["api_errors"] += 1
                    result_label = "API错误"

                record = {
                    "sheet": ws.title, "row": str(row),
                    "standard_cell": standard_cell, "execution_cell": execution_cell,
                    "execution_label": exec_label, "result": result_label,
                    "reason": _truncate(reason, 500) if reason else "",
                    "a_text": _truncate(a_for_judge, 800), "c_text": _truncate(c_for_judge, 800),
                    "raw": _truncate(raw, 800) if raw else "",
                }
                result_details.append(record)
                if result_label in {"✗ 不符合", "API错误", "不确定"}:
                    issue_details.append(record)

                    ev_refs = [{"sheet": ws.title, "cell_or_range": execution_cell, "excerpt": c_for_judge[:_EXCERPT_MAX_LEN]}]
                    ev_refs = _verify_evidence_refs(ev_refs, ws)

                    if result_label == "✗ 不符合":
                        issue_type, sev, sug = _classify_mismatch(a_for_judge, c_for_judge, reason or raw or "")
                        sev_p = _SEVERITY_FROM_CHINESE.get(sev, sev)
                        proc_status = "fail"
                        proc_unknown = ""
                        if not ev_refs:
                            proc_status = "unknown"
                            proc_unknown = "无法引用原始证据佐证该判定，降级为不确定"
                            sev_p = "P2"
                        basis_parts = []
                        if a_for_judge:
                            basis_parts.append("标准程序: " + _truncate(a_for_judge, 800))
                        if reason or raw:
                            basis_parts.append("LLM依据: " + _truncate(reason or raw, 2000))
                        if ev_refs:
                            basis_parts.append("引用: " + "; ".join(
                                f"{r.get('cell_or_range', '')}: {r.get('excerpt', '')[:200]}"
                                for r in ev_refs[:2] if r.get('excerpt')
                            ))
                        if proc_unknown:
                            basis_parts.append("不确定原因: " + proc_unknown)
                        findings.append(
                            Finding(
                                issue_type=issue_type,
                                severity=sev_p,
                                sheet=ws.title,
                                cell=execution_cell,
                                snippet=_truncate(c_for_judge, 220),
                                basis=_truncate("\n".join(p for p in basis_parts if p), 3000),
                                suggestion=sug,
                                status=proc_status,
                                risk_type="一致性",
                                evidence_refs=json.dumps(ev_refs, ensure_ascii=False),
                                conclusion=f"执行程序与标准审计程序在 {execution_cell} 不一致",
                                reasons=json.dumps([_truncate(reason or raw, 300)] if (reason or raw) else [], ensure_ascii=False),
                                fix_suggestion_detail=json.dumps({"supplement_explanation": sug[:300]}, ensure_ascii=False),
                                unknown_reason=proc_unknown,
                            )
                        )
                    elif result_label == "API错误":
                        err_detail = reason or raw or "API请求超时或返回错误"
                        findings.append(
                            Finding(
                                issue_type="A-C对应性：LLM调用失败（需人工复核）",
                                severity="P1",
                                sheet=ws.title,
                                cell=execution_cell,
                                snippet=_truncate(c_for_judge, 220),
                                basis=_truncate(
                                    f"标准程序: {_truncate(a_for_judge, 400)}\n"
                                    f"LLM调用失败详情: {err_detail[:800]}\n"
                                    "无法自动判定「标准审计程序」与「实际执行程序」的对应性，"
                                    "请人工对比以下两项：\n"
                                    f"  A列（标准）: {_truncate(a_for_judge, 200)}\n"
                                    f"  C列（执行）: {_truncate(c_for_judge, 200)}",
                                    3000,
                                ),
                                suggestion=(
                                    "人工复核步骤：\n"
                                    "1) 对照A列标准审计程序，逐项检查C列执行是否覆盖要求的审计动作、对象、条件与证据类型；\n"
                                    "2) 如实际已覆盖，在底稿中补充「证据→核查点→结论」的对应说明；\n"
                                    "3) 如确实未覆盖，作为缺陷记录并提出整改。"
                                ),
                                status="unknown",
                                unknown_reason=f"LLM调用失败（{err_detail[:200]}），无法自动判定，需人工复核",
                                risk_type="证据不足",
                                evidence_refs=json.dumps(ev_refs, ensure_ascii=False),
                            )
                        )
                    else:
                        llm_partial = reason or raw or ""
                        findings.append(
                            Finding(
                                issue_type="A-C对应性：LLM无法判定（需人工确认）",
                                severity="P1",
                                sheet=ws.title,
                                cell=execution_cell,
                                snippet=_truncate(c_for_judge, 220),
                                basis=_truncate(
                                    f"标准程序: {_truncate(a_for_judge, 400)}\n"
                                    f"LLM分析: {_truncate(llm_partial, 800)}\n"
                                    f"LLM无法确认标准与执行是否一致，建议人工判断覆盖性。",
                                    3000,
                                ),
                                suggestion=(
                                    "人工确认步骤：\n"
                                    "1) 逐条核对：A列每个审计动作/对象/条件是否在C列中有对应描述；\n"
                                    "2) 关注差异点：访谈对象是否覆盖关键岗位、检查文件是否主题一致、"
                                    "是否有替代性程序覆盖同一控制点；\n"
                                    "3) 如确认不符合，补充完整执行程序；如确认符合，在底稿中写明对应关系。"
                                ),
                                status="unknown",
                                unknown_reason=f"LLM无法明确判定标准审计程序与实际执行程序的一致性"
                                + (f"（LLM分析: {_truncate(llm_partial, 150)}）" if llm_partial else ""),
                                risk_type="证据不足",
                                evidence_refs=json.dumps(ev_refs, ensure_ascii=False),
                            )
                        )

                if sleep_seconds and sleep_seconds > 0:
                    await asyncio.sleep(float(sleep_seconds))

    report: Dict[str, object] = {
        "total": total, "matched": matched, "failed": failed, "api_errors": api_errors,
        "skipped_a_empty": skipped_a_empty, "skipped_c_empty": skipped_c_empty,
        "skipped_ref": skipped_ref, "skipped_header": skipped_header,
        "sheet_stats": sheet_stats, "results": result_details, "issues": issue_details,
    }
    return report, findings
