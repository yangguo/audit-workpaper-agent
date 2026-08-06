"""Checkpoint loading and checkpoint-based LLM review (ported from analyze_excel.py)."""
import asyncio
import json
import os
import re
from collections import defaultdict
from typing import Callable, Dict, List, Optional, Sequence

import openpyxl
from openpyxl.utils import get_column_letter

from review.constants import CHECKPOINT_VOCAB
from review.excel_utils import _build_sheet_text_for_llm, _get_cell_value, _truncate
from review.attachments import _attachments_context_for_sheet, _verify_attachment_evidence_refs
from review.llm import _llm_request_json_list, _llm_stat
from review.models import Finding, _SEVERITY_FROM_CHINESE
from review.validation import _verify_evidence_refs


def load_checkpoints_xlsx(checkpoints_path: str) -> Dict[str, List[str]]:
    """Load a checkpoints workbook into a {sheet_id: [check_text, ...]} map.

    Column 1 holds the sheet id (sticky: following rows inherit it until a new
    non-empty value appears); column 3 holds the check text.
    """
    if not checkpoints_path:
        return {}
    if not os.path.exists(checkpoints_path):
        raise FileNotFoundError(checkpoints_path)

    wb = openpyxl.load_workbook(checkpoints_path, data_only=True)
    ws = wb.active

    checkpoints_by_sheet: Dict[str, List[str]] = defaultdict(list)
    last_sheet = None
    for row in range(1, (ws.max_row or 0) + 1):
        sheet_id = ws.cell(row=row, column=1).value
        check_text = ws.cell(row=row, column=3).value
        if isinstance(sheet_id, str):
            sheet_id = sheet_id.strip()
        if isinstance(check_text, str):
            check_text = check_text.strip()

        if sheet_id:
            last_sheet = str(sheet_id).strip()
        if not last_sheet:
            continue
        if not check_text:
            continue

        checkpoints_by_sheet[last_sheet].append(str(check_text).strip())

    return dict(checkpoints_by_sheet)


def _split_checkpoints(text: str) -> List[str]:
    """Split a checkpoint cell into individual checkpoint strings."""
    if not text:
        return []
    normalized = str(text).replace("\r\n", "\n").replace("\r", "\n")
    parts: List[str] = []
    for chunk in normalized.split("\n"):
        s = chunk.strip()
        if not s:
            continue
        s = re.sub(r"^\s*\d+\s*[.、]\s*", "", s)
        if s:
            parts.append(s)
    if parts:
        return parts
    return [normalized.strip()] if normalized.strip() else []


def _extract_checkpoint_keywords(checkpoint: str) -> List[str]:
    """Extract evidence-vocabulary keywords from a checkpoint string."""
    t = checkpoint or ""
    hits = [w for w in CHECKPOINT_VOCAB if w and w in t]
    if hits:
        seen = set()
        out: List[str] = []
        for h in hits:
            if h in seen:
                continue
            seen.add(h)
            out.append(h)
        return out

    segments = re.split(r"[，；。;,.、()\[\]（）]+", t)
    picked: List[str] = []
    for seg in segments:
        s = seg.strip()
        if not s:
            continue
        if len(s) < 4:
            continue
        if len(s) > 26:
            s = s[:26]
        if any("一" <= ch <= "鿿" for ch in s):
            picked.append(s)
        if len(picked) >= 3:
            break
    return picked or [t[:16].strip()] if t.strip() else []


async def _llm_check_sheet_by_checkpoints(
    *,
    llm,
    ws_title: str,
    ws,
    checkpoints: Sequence[str],
    attachments: Optional[Dict[str, object]] = None,
    batch_size: int = 4,
    sleep_seconds: float = 0.2,
    attachments_preview: Optional[Dict[str, object]] = None,
    on_progress: Optional[Callable[[str, str], None]] = None,
) -> List[Finding]:
    if not checkpoints:
        return []
    attachments = attachments if attachments is not None else attachments_preview

    flat: List[str] = []
    for raw in checkpoints:
        for item in _split_checkpoints(raw):
            if item and item.strip():
                flat.append(item.strip())

    seen = set()
    deduped: List[str] = []
    for x in flat:
        if x in seen:
            continue
        seen.add(x)
        deduped.append(x)

    if not deduped:
        return []

    sheet_text = _build_sheet_text_for_llm(ws, max_cells=260, max_chars=24000)
    if not sheet_text.strip():
        return [
            Finding(
                issue_type="LLM判定：检查要点无法复核（Sheet无文本）",
                severity="P1",
                sheet=ws_title,
                cell=None,
                snippet="",
                basis="Sheet内未提取到可用于复核的文本单元格。",
                suggestion="确认该Sheet是否为图片/对象或空白；如为图片型底稿需先OCR或改用可读文本版本。",
                status="unknown",
                unknown_reason="Sheet内无文本可复核",
                risk_type="证据不足",
            )
        ]

    system_prompt = (
        "你是一名严格的IT审计/财务审计质量复核专家。\n"
        "你将收到：\n"
        "1) 某个底稿Sheet的文本化内容（每行含单元格坐标与文字）；\n"
        "2) 该Sheet对应的“检查要点”清单；\n"
        "3) （可选）该Sheet所引用的附件目录证据（路径/类型/解析状态/提取内容）。\n\n"
        "你的任务：逐条检查要点，判断该Sheet是否存在相关问题（未覆盖/证据不足/表述不清/范围不全/仅访谈等）。\n"
        "重要判断规则（避免误报）：\n"
        "1) 如果Sheet内容明确写明“未执行/未开展/未进行/不存在/未对...审阅/未清查”，则这本身意味着控制未执行或存在缺陷。此时不要将“缺少过程证据”作为独立问题点重复输出；应将问题表述为“控制未执行/未开展（无清查过程证据属结果）”。\n"
        "2) 仅当Sheet声称已执行（如“已审阅/已清查/已复核/已下发确认/已收集反馈”），但未提供相应过程证据时，才输出“缺少过程证据/证据不足”。\n"
        "输出要求：必须输出严格JSON对象：{\"results\": [...]}。\n"
        "results每个元素必须包含字段：\n"
        "- id: 整数\n"
        "- checkpoint: 字符串（原检查要点）\n"
        "- status: \"pass\"/\"fail\"/\"unknown\" (无问题/有问题/不确定)\n"
        "- conclusion: 一句话结论（当status=pass时也建议给出一句话）\n"
        "- reasons: 字符串数组，2-5条要点（说明判断依据）\n"
        "- evidence_refs: 数组，每个元素含 {sheet, cell_or_range, attachment(可选), excerpt(原文摘录)}。没有attachment时excerpt必须逐字来自sheet_text对应单元格；如果结论依赖附件目录/Agent调查证据，必须填写已提供的相对路径，并让excerpt逐字来自该附件提取文本。\n"
        "  * status=fail时必须至少1个evidence_ref；无法引用原文时status必须为unknown\n"
        "  * excerpt不可编造，必须是sheet_text中能找到的原句片段\n"
        "- severity: \"P0\"/\"P1\"/\"P2\"（当status!=pass时必填）\n"
        "- risk_type: \"覆盖性\"/\"一致性\"/\"证据不足\"/\"方法性\"/\"逻辑性\"/\"跨字段一致性\"之一\n"
        "- fix_suggestion: 对象，含 {supplement_explanation}\n"
        "- unknown_reason: status=unknown时必填，≥10字符\n"
    )

    findings: List[Finding] = []
    cell_ref_re = re.compile(r"^[A-Z]{1,3}\d{1,7}$")
    for start in range(0, len(deduped), max(1, int(batch_size))):
        chunk = deduped[start: start + max(1, int(batch_size))]
        end = start + len(chunk)
        attachments_text = _attachments_context_for_sheet(ws, attachments or {}) if attachments else ""
        payload = {
            "sheet": ws_title,
            "checkpoints": [{"id": start + i + 1, "checkpoint": cp} for i, cp in enumerate(chunk)],
            "sheet_text": sheet_text,
            "attachments": attachments_text,
        }
        user_prompt = "请按要求逐条复核以下检查要点：\n" + json.dumps(payload, ensure_ascii=False, indent=2)

        def _consume_results(objs: List[object]) -> None:
            for obj in objs:
                if not isinstance(obj, dict):
                    continue
                raw_status = str(obj.get("status", "")).strip()
                if raw_status == "无问题":
                    status = "pass"
                elif raw_status == "有问题":
                    status = "fail"
                elif raw_status == "不确定":
                    status = "unknown"
                else:
                    status = raw_status
                if status == "pass":
                    continue
                checkpoint = str(obj.get("checkpoint", "")).strip()
                raw_sev = str(obj.get("severity", "")).strip()
                severity = _SEVERITY_FROM_CHINESE.get(raw_sev, raw_sev)
                if severity not in ("P0", "P1", "P2"):
                    severity = "P1" if status != "unknown" else "P2"
                risk_type = str(obj.get("risk_type", "")).strip()
                base_issue = "检查要点存在问题" if status == "fail" else "检查要点信息不足/不确定"
                provided_issue = str(obj.get("issue_type", "")).strip()
                issue_type = provided_issue or (f"{base_issue}（{risk_type}）" if risk_type else base_issue)
                basis = str(obj.get("basis", "")).strip()
                suggestion = str(obj.get("suggestion", "")).strip()
                conclusion = str(obj.get("conclusion", "")).strip()
                reasons_raw = obj.get("reasons", [])
                reasons_list = [str(r).strip() for r in reasons_raw if r] if isinstance(reasons_raw, list) else []
                unknown_reason = str(obj.get("unknown_reason", "")).strip()
                fix_suggestion_obj = obj.get("fix_suggestion") or {}
                if not isinstance(fix_suggestion_obj, dict):
                    fix_suggestion_obj = {}
                sheet_indicates_not_done = False
                sheet_signal_text = (basis or "") + "\n" + (conclusion or "") + "\n" + (sheet_text or "")
                if any(k in sheet_signal_text for k in ("未对", "未进行", "未开展", "未执行", "不存在", "未审阅", "未清查", "未复核")):
                    if any(k in checkpoint for k in ("清查全过程", "过程证据", "留痕", "反馈", "下发", "收集")) or any(
                        k in issue_type for k in ("证据", "缺失", "不足", "留痕", "反馈", "全过程")
                    ):
                        sheet_indicates_not_done = True
                related_cells = obj.get("related_cells", [])
                picked_cells: List[str] = []
                seen_cells = set()
                if isinstance(related_cells, list) and related_cells:
                    for c in related_cells:
                        cc = str(c).strip().upper()
                        if not cell_ref_re.match(cc):
                            continue
                        if cc in seen_cells:
                            continue
                        cell_text = _get_cell_value(ws, cc)
                        if not cell_text:
                            continue
                        seen_cells.add(cc)
                        picked_cells.append(cc)
                        if len(picked_cells) >= 6:
                            break
                if basis:
                    hits = re.findall(r"\b[A-Z]{1,3}\d{1,7}\b", basis.upper())
                    for h in hits:
                        hh = str(h).strip().upper()
                        if not cell_ref_re.match(hh):
                            continue
                        if hh in seen_cells:
                            continue
                        cell_text = _get_cell_value(ws, hh)
                        if not cell_text:
                            continue
                        seen_cells.add(hh)
                        picked_cells.append(hh)
                        if len(picked_cells) >= 6:
                            break
                cell = ",".join(picked_cells) if picked_cells else None

                evidence_refs_list: List[dict] = []
                raw_refs = obj.get("evidence_refs")
                if isinstance(raw_refs, list) and raw_refs:
                    for ref in raw_refs:
                        if not isinstance(ref, dict):
                            continue
                        ev_cell = str(ref.get("cell_or_range", "")).strip()
                        ev_sheet = str(ref.get("sheet", "")).strip() or ws_title
                        ev_attachment = str(ref.get("attachment", "")).strip()
                        ev_excerpt = str(ref.get("excerpt", "")).strip()
                        if ev_cell:
                            evidence_refs_list.append({
                                "sheet": ev_sheet,
                                "cell_or_range": ev_cell,
                                "attachment": ev_attachment,
                                "excerpt": ev_excerpt,
                            })
                local_refs = [ref for ref in evidence_refs_list if not ref.get("attachment")]
                attachment_refs = [ref for ref in evidence_refs_list if ref.get("attachment")]
                evidence_refs_list = _verify_evidence_refs(local_refs, ws)
                evidence_refs_list.extend(
                    _verify_attachment_evidence_refs(attachment_refs, attachments or {}, ws=ws)
                )
                if not evidence_refs_list and picked_cells:
                    for cc in picked_cells:
                        ctext = _get_cell_value(ws, cc) or ""
                        if ctext:
                            evidence_refs_list.append({
                                "sheet": ws_title,
                                "cell_or_range": cc,
                                "excerpt": ctext[:2000],
                            })

                if status == "fail" and not evidence_refs_list:
                    status = "unknown"
                    unknown_reason = unknown_reason or "无法引用原始证据佐证该判定，降级为不确定"
                    severity = "P2"

                missing_evidence = obj.get("missing_evidence", [])
                missing_text = ""
                if isinstance(missing_evidence, list) and missing_evidence:
                    missing_text = "缺失证据: " + "、".join(str(x).strip() for x in missing_evidence if str(x).strip())
                if not fix_suggestion_obj.get("required_evidence_type") and missing_evidence:
                    if isinstance(missing_evidence, list) and missing_evidence:
                        fix_suggestion_obj["required_evidence_type"] = "、".join(str(x).strip() for x in missing_evidence if str(x).strip())[:300]
                if not fix_suggestion_obj.get("supplement_explanation") and suggestion:
                    fix_suggestion_obj["supplement_explanation"] = suggestion[:300]

                basis_parts: List[str] = []
                if checkpoint:
                    basis_parts.append("检查要点: " + checkpoint)
                if conclusion:
                    basis_parts.append("结论: " + conclusion)
                if reasons_list:
                    basis_parts.append("理由: " + " | ".join(reasons_list[:5]))
                if basis:
                    basis_parts.append("依据: " + basis)
                if missing_text:
                    basis_parts.append(missing_text)
                if evidence_refs_list:
                    refs_text = "; ".join(
                        f"{r.get('cell_or_range', '')}: {r.get('excerpt', '')[:200]}"
                        for r in evidence_refs_list[:3] if r.get('excerpt')
                    )
                    if refs_text:
                        basis_parts.append("引用: " + refs_text)
                if unknown_reason:
                    basis_parts.append("不确定原因: " + unknown_reason)
                basis2 = "\n".join(p for p in basis_parts if p).strip()

                snippet = ""
                if picked_cells:
                    parts: List[str] = []
                    for cc in picked_cells[:6]:
                        cell_text = _get_cell_value(ws, cc) or ""
                        if not cell_text:
                            continue
                        parts.append(f"{cc}: {_truncate(cell_text, 60)}")
                    if parts:
                        snippet = _truncate(" | ".join(parts), 220)
                if not snippet and evidence_refs_list:
                    parts = []
                    for ref in evidence_refs_list[:3]:
                        ex = ref.get("excerpt", "")
                        cc = ref.get("cell_or_range", "")
                        if ex:
                            parts.append(f"{cc}: {_truncate(ex, 60)}")
                    if parts:
                        snippet = _truncate(" | ".join(parts), 220)
                if not snippet and basis2:
                    snippet = _truncate(basis2.replace("\n", " "), 220)

                if sheet_indicates_not_done and "检查要点" in (issue_type or "") and any(
                    k in (issue_type or "") for k in ("证据不足", "缺失", "全过程", "留痕")
                ):
                    issue_type = "检查要点-控制未执行/未开展（因此无过程证据）"
                    if severity not in ("P0", "P1"):
                        severity = "P1"
                    if not suggestion:
                        suggestion = (
                            "明确该控制在审计期间未执行的事实与影响；作为缺陷记录并提出整改：建立权限清查机制（导出清单-下发确认-收集反馈-例外处置-复核留痕），并补充后续期间执行记录。"
                        )

                findings.append(
                    Finding(
                        issue_type="LLM判定：" + issue_type,
                        severity=severity,
                        sheet=ws_title,
                        cell=cell,
                        snippet=snippet,
                        basis=_truncate(basis2 or "LLM判定存在问题/不确定", 3000),
                        suggestion=_truncate(
                            suggestion or "对照检查要点补充执行步骤与证据，并在底稿中保留可复核来源。",
                            1200,
                        ),
                        status=status,
                        risk_type=risk_type or ("证据不足" if status == "fail" else ""),
                        evidence_refs=json.dumps(evidence_refs_list, ensure_ascii=False),
                        conclusion=conclusion,
                        reasons=json.dumps(reasons_list, ensure_ascii=False),
                        fix_suggestion_detail=json.dumps(fix_suggestion_obj, ensure_ascii=False),
                        unknown_reason=unknown_reason,
                    )
                )

        stage = f"checkpoints:{ws_title}"
        parsed, last_error = await _llm_request_json_list(
            llm=llm,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            stage=stage,
            max_attempts=2,
        )
        if parsed is not None:
            _consume_results(parsed)
        else:
            if len(chunk) > 1:
                _llm_stat(stage, "fallback_single", 1)
                for i, cp in enumerate(chunk):
                    payload1 = {
                        "sheet": ws_title,
                        "checkpoints": [{"id": start + i + 1, "checkpoint": cp}],
                        "sheet_text": sheet_text,
                        "attachments": attachments_text,
                    }
                    user_prompt1 = "请按要求逐条复核以下检查要点：\n" + json.dumps(payload1, ensure_ascii=False, indent=2)
                    parsed1, err1 = await _llm_request_json_list(
                        llm=llm, system_prompt=system_prompt, user_prompt=user_prompt1,
                        stage=stage, max_attempts=2,
                    )
                    if parsed1 is not None:
                        _consume_results(parsed1)
                    else:
                        findings.append(_checkpoint_failure_finding(ws_title, cp, err1))
            else:
                findings.append(_checkpoint_failure_finding(ws_title, "；".join(chunk), last_error))

        if sleep_seconds and sleep_seconds > 0:
            await asyncio.sleep(float(sleep_seconds))

        # Emit intra-chunk progress so long stages don't appear frozen.
        # Best-effort: callbacks must not break the review.
        if on_progress is not None:
            try:
                msg = f"已处理 {min(end, len(deduped))} / {len(deduped)} 个检查要点"
                on_progress(stage, msg)
            except Exception:
                pass

    return findings


def _checkpoint_failure_finding(ws_title: str, checkpoint: str, error) -> Finding:
    return Finding(
        issue_type="LLM判定：检查要点复核失败",
        severity="P1",
        sheet=ws_title,
        cell=None,
        snippet="",
        basis=_truncate(f"检查要点: {checkpoint}\nLLM调用失败: {error}", 1200),
        suggestion="检查LLM接口配置（.env）、网络连通性；必要时减少检查要点条数或缩短Sheet文本后重试。",
        status="unknown",
        unknown_reason="LLM调用失败，无法复核",
        risk_type="证据不足",
    )
