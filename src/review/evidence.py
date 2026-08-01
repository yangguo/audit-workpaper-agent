"""Immutable input and workbook-evidence builders for shadow artifacts."""
import hashlib
import json
import os
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from review.contracts import CellEvidence, EvidenceGraph, InputFile, SheetEvidence
from review.excel_utils import _detect_layout, _is_empty, _normalize_sheet_id


_DEFAULT_MAX_CELLS = 50_000
_HASH_CHUNK_SIZE = 1024 * 1024


def sha256_file(path: str | Path) -> str:
    """Return a SHA-256 checksum without loading the complete file into memory."""
    digest = hashlib.sha256()
    with Path(path).open("rb") as source:
        while chunk := source.read(_HASH_CHUNK_SIZE):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_path(path: str | Path) -> str:
    """Hash a file or directory with names and bytes in deterministic order."""
    source = Path(path)
    if source.is_file():
        return sha256_file(source)
    if not source.is_dir():
        raise FileNotFoundError(str(source))

    digest = hashlib.sha256()
    for child in sorted(source.rglob("*"), key=lambda p: p.relative_to(source).as_posix()):
        if not child.is_file() or child.is_symlink():
            continue
        relative = child.relative_to(source).as_posix().encode("utf-8")
        digest.update(relative)
        digest.update(b"\0")
        with child.open("rb") as handle:
            while chunk := handle.read(_HASH_CHUNK_SIZE):
                digest.update(chunk)
    return digest.hexdigest()


def _path_size(path: Path) -> int:
    if path.is_file():
        return path.stat().st_size
    return sum(
        child.stat().st_size
        for child in path.rglob("*")
        if child.is_file() and not child.is_symlink()
    )


def _input_file(role: str, path: str) -> InputFile:
    source = Path(path)
    if not source.exists() or source.is_symlink():
        raise FileNotFoundError(str(source))
    return InputFile(
        role=role,  # type: ignore[arg-type]
        path=str(source),
        filename=source.name,
        sha256=sha256_path(source),
        size=_path_size(source),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if source.is_file()
            else "inode/directory"
        ),
    )


def build_input_files(
    *,
    workpaper_path: str,
    checkpoints_path: str = "",
    attachments_dir: str = "",
    attachments_preview_path: str = "",
) -> list[InputFile]:
    """Build ordered, content-addressed records for review input files."""
    inputs = [_input_file("workpaper", workpaper_path)]
    if checkpoints_path:
        inputs.append(_input_file("checkpoints", checkpoints_path))
    if attachments_dir:
        inputs.append(_input_file("attachments_dir", attachments_dir))
    if attachments_preview_path:
        inputs.append(_input_file("attachments_preview", attachments_preview_path))
    return inputs


def _snapshot_limit(max_cells: int | None) -> int:
    if max_cells is not None:
        try:
            candidate = int(max_cells)
        except (TypeError, ValueError):
            candidate = 0
    else:
        try:
            candidate = int(
                os.getenv("REVIEW_EVIDENCE_SNAPSHOT_MAX_CELLS", _DEFAULT_MAX_CELLS)
            )
        except (TypeError, ValueError):
            candidate = 0
    return candidate if candidate > 0 else _DEFAULT_MAX_CELLS


def _json_value(value: Any) -> str:
    """Represent openpyxl values deterministically for JSON and hashing."""
    if isinstance(value, ArrayFormula):
        return json.dumps(
            {"t": value.t, "ref": value.ref, "text": value.text},
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
    if isinstance(value, DataTableFormula):
        return json.dumps(
            {
                "t": value.t,
                "ref": value.ref,
                "ca": _formula_flag(value.ca),
                "dt2D": _formula_flag(value.dt2D),
                "dtr": _formula_flag(value.dtr),
                "r1": value.r1,
                "r2": value.r2,
                "del1": _formula_flag(value.del1),
                "del2": _formula_flag(value.del2),
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.hex()
    return str(value)


def _formula_flag(value: Any) -> bool:
    return value is True or value == 1 or str(value).lower() in {"1", "true"}


def _digest_payload(payload: dict[str, Any]) -> str:
    serialized = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def _cell_evidence(cell, source_sha256: str, sheet_name: str) -> CellEvidence:
    value = _json_value(cell.value)
    formula = value if cell.data_type == "f" or value.startswith("=") else None
    content_hash = _digest_payload(
        {
            "value": value,
            "formula": formula,
            "data_type": cell.data_type or "",
        }
    )
    evidence_hash = _digest_payload(
        {
            "schema": "evidence-v1",
            "source_sha256": source_sha256,
            "sheet_name": sheet_name,
            "coordinate": cell.coordinate,
            "content_hash": content_hash,
        }
    )
    return CellEvidence(
        evidence_id=f"ev:{evidence_hash}",
        sheet_name=sheet_name,
        coordinate=cell.coordinate,
        value=value,
        formula=formula,
        data_type=cell.data_type or "",
        content_hash=content_hash,
    )


def _sheet_hash(
    ws,
    cells: list[CellEvidence],
    *,
    normalized_name: str,
    layout_header_row: int | None,
    standard_column: int | None,
    execution_columns: list[int],
    merged_ranges: list[str],
) -> str:
    return _digest_payload(
        {
            "sheet_name": ws.title,
            "normalized_name": normalized_name,
            "max_row": ws.max_row or 0,
            "max_column": ws.max_column or 0,
            "layout_header_row": layout_header_row,
            "standard_column": standard_column,
            "execution_columns": execution_columns,
            "merged_ranges": merged_ranges,
            "cells": [
                {"coordinate": cell.coordinate, "content_hash": cell.content_hash}
                for cell in cells
            ],
        }
    )


def build_evidence_graph(
    wb: openpyxl.Workbook,
    *,
    source_sha256: str,
    max_cells: int | None = None,
) -> EvidenceGraph:
    """Create a row-major, bounded evidence snapshot of a loaded workbook."""
    limit = _snapshot_limit(max_cells)
    captured = 0
    omitted = 0
    sheets: list[SheetEvidence] = []

    for ws in wb.worksheets:
        header_row, standard_column, execution_columns = _detect_layout(ws)
        normalized_name = _normalize_sheet_id(ws.title)
        merged_ranges = sorted(str(rng) for rng in ws.merged_cells.ranges)
        cells: list[CellEvidence] = []
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if _is_empty(cell.value):
                    continue
                if captured >= limit:
                    omitted += 1
                    continue
                cells.append(_cell_evidence(cell, source_sha256, ws.title))
                captured += 1

        sheets.append(
            SheetEvidence(
                name=ws.title,
                normalized_name=normalized_name,
                sheet_hash=_sheet_hash(
                    ws,
                    cells,
                    normalized_name=normalized_name,
                    layout_header_row=header_row,
                    standard_column=standard_column or None,
                    execution_columns=execution_columns,
                    merged_ranges=merged_ranges,
                ),
                max_row=ws.max_row or 0,
                max_column=ws.max_column or 0,
                layout_header_row=header_row,
                standard_column=standard_column or None,
                execution_columns=execution_columns,
                merged_ranges=merged_ranges,
                cells=cells,
            )
        )

    return EvidenceGraph(
        source_sha256=source_sha256,
        sheets=sheets,
        captured_cell_count=captured,
        omitted_cell_count=omitted,
        capture_status="truncated" if omitted else "complete",
    )
