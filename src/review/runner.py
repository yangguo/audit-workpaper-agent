"""Background review runner + in-process registry.

Decouples the long-running review pipeline from the agent tool call: the tool
starts the review as a background asyncio task and returns immediately with a
review_id; the frontend polls GET /review/{review_id}/status until completed,
then fetches GET /findings/{review_id}.

Single uvicorn worker => one event loop => a module-level registry is safe.
Starting a new review cancels any currently-running one to prevent stacking.
"""
import asyncio
import logging
import os
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from review.attachments import build_attachment_index, load_attachments_preview_xlsx
from review.checkpoints import load_checkpoints_xlsx
from review.contracts import ExecutionComponentRef, PolicyPackRef, ReviewManifest
from review.evidence import build_evidence_graph, build_input_files, sha256_file
from review.execution_context import (
    ReviewExecutionContext,
    build_review_execution_context,
    capture_runtime_config,
    component_ref_from_path,
)
from review.evidence_provenance import EvidenceProvenanceIndex, verify_finding_evidence
from review.evaluators import execute_policy_plan
from review.finding_comparison import compare_finding_sets
from review.findings import build_v2_findings, project_v2_findings_to_v1
from review.finding_taxonomy import (
    AssertionCatalog,
    assertion_catalog_directory,
    fallback_assertion_catalog,
    load_assertion_catalog,
)
from review.judgement import build_judgement_requests, execute_judgement_requests
from review.llm import LLM_CALL_STATS, get_review_llm
from review.planner import build_review_plan
from review.policy import load_policy_pack
from review.pipeline import run_review
from review.result_quality import build_quality_envelope
from review.remediation import enrich_finding_quality
from storage.findings_store import load_findings, save_findings
from storage.review_artifact_store import ReviewArtifactStore

_logger = logging.getLogger("review.runner")

# review_id -> {status, task, started_at, source, stats?, error?}
_REGISTRY: Dict[str, dict] = {}
_SAFE_COMPONENT = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]*$")


def _stage_b_policy_config() -> tuple[str, str, str, str | None]:
    mode = os.getenv("REVIEW_POLICY_MODE", "shadow").strip().lower()
    if mode not in {"shadow", "off"}:
        raise ValueError("REVIEW_POLICY_MODE must be shadow or off")
    pack_id = os.getenv("REVIEW_POLICY_PACK_ID", "itgc-core").strip()
    pack_version = os.getenv("REVIEW_POLICY_PACK_VERSION", "1.0.0").strip()
    root = os.getenv("REVIEW_POLICY_PACK_ROOT", "").strip() or None
    return mode, pack_id, pack_version, root


def _stage_c_judgement_config() -> tuple[str, str, str, str | None, int]:
    mode = os.getenv("REVIEW_JUDGEMENT_MODE", "off").strip().lower()
    if mode not in {"shadow", "off"}:
        raise ValueError("REVIEW_JUDGEMENT_MODE must be shadow or off")
    pack_id = os.getenv("REVIEW_JUDGEMENT_PACK_ID", "itgc-judgement").strip()
    pack_version = os.getenv("REVIEW_JUDGEMENT_PACK_VERSION", "1.0.0").strip()
    root = (
        os.getenv("REVIEW_JUDGEMENT_PACK_ROOT", "").strip()
        or os.getenv("REVIEW_POLICY_PACK_ROOT", "").strip()
        or None
    )
    try:
        max_requests = int(os.getenv("REVIEW_JUDGEMENT_MAX_REQUESTS", "200"))
    except (TypeError, ValueError) as exc:
        raise ValueError("REVIEW_JUDGEMENT_MAX_REQUESTS must be an integer") from exc
    if max_requests <= 0:
        raise ValueError("REVIEW_JUDGEMENT_MAX_REQUESTS must be greater than zero")
    return mode, pack_id, pack_version, root, max_requests


def _assertion_catalog_config() -> tuple[str, str, str | None]:
    """Read the declarative finding-assertion catalog configuration."""

    pack_id = os.getenv("REVIEW_ASSERTION_CATALOG_ID", "review-quality").strip()
    version = os.getenv("REVIEW_ASSERTION_CATALOG_VERSION", "1.0.0").strip()
    root = os.getenv("REVIEW_ASSERTION_CATALOG_ROOT", "").strip() or None
    return pack_id, version, root


def _assertion_catalog_component(
    *, pack_id: str, version: str, root: str | None
) -> ExecutionComponentRef:
    """Fingerprint the assertion declarations that govern V1 classification."""

    catalog_dir = assertion_catalog_directory(
        pack_id=pack_id,
        version=version,
        root=root,
    )
    return component_ref_from_path(
        component_id="review-quality-assertions",
        version=version,
        path=catalog_dir / "assertions.json",
    )


def _result_quality_config() -> str:
    mode = os.getenv("REVIEW_RESULT_QUALITY_MODE", "shadow").strip().lower()
    if mode not in {"off", "shadow", "on"}:
        raise ValueError("REVIEW_RESULT_QUALITY_MODE must be off, shadow or on")
    return mode


def _requested_sheet_names(sheets: Optional[str]) -> list[str]:
    return [name.strip() for name in (sheets or "").split(",") if name.strip()]


def _execution_engine_version(policy_mode: str) -> str:
    default = "stage-b-policy-shadow" if policy_mode == "shadow" else "stage-a-shadow"
    return os.getenv("REVIEW_ENGINE_VERSION", default).strip() or default


def _policy_component_ref(
    *,
    component_id: str,
    pack_id: str,
    version: str,
    root: str | None,
) -> ExecutionComponentRef:
    """Fingerprint a policy pack directory without putting its path in output."""

    if not _SAFE_COMPONENT.fullmatch(pack_id) or not _SAFE_COMPONENT.fullmatch(version):
        return component_ref_from_path(
            component_id=component_id,
            version=version or "unknown",
            path=Path(__file__).parent / ".missing-policy-pack",
        )
    pack_root = (
        Path(root).expanduser()
        if root
        else Path(__file__).resolve().parents[2] / "policy_packs"
    )
    return component_ref_from_path(
        component_id=component_id,
        version=version,
        path=pack_root / pack_id / version,
    )


def _workpaper_sha256(context: ReviewExecutionContext | None) -> str:
    if context is None:
        return ""
    for item in context.manifest.inputs:
        if item.role == "workpaper":
            return item.sha256
    return ""


def _attach_result_quality(
    *,
    findings: list[dict],
    workbook,
    file_path: str,
    attachments: Optional[dict[str, object]],
    execution_context: ReviewExecutionContext | None = None,
) -> tuple[list[dict], dict]:
    """Attach quality metadata while preserving legacy V1 by default."""
    mode = _result_quality_config()
    if mode == "off":
        return findings, {
            "mode": "off",
            "total_findings": len(findings),
            "citation_status": {},
            "rejected_refs": 0,
            "downgraded_findings": 0,
        }

    source_sha256 = _workpaper_sha256(execution_context) or sha256_file(file_path)
    graph = build_evidence_graph(workbook, source_sha256=source_sha256)
    index = EvidenceProvenanceIndex(
        graph,
        attachments=attachments or {},
        workbook=workbook,
    )
    engine_version = (
        execution_context.manifest.engine_version
        if execution_context is not None
        else (
            os.getenv("REVIEW_ENGINE_VERSION", "stage-a-quality-shadow").strip()
            or "stage-a-quality-shadow"
        )
    )
    status_counts: dict[str, int] = {}
    gate_counts: dict[str, dict[str, int]] = {}
    rejected_refs = 0
    downgraded_findings = 0
    enriched: list[dict] = []
    for finding in findings:
        original = dict(finding)
        gate_payload = original.pop("quality_gates", {})
        checked, verification = verify_finding_evidence(original, index)
        checked.pop("quality_gates", None)
        output = checked if mode == "on" else original
        if mode == "on" and checked.get("status") != original.get("status"):
            downgraded_findings += 1
        quality = build_quality_envelope(
            output,
            input_sha256=source_sha256,
            input_set_sha256=(
                execution_context.input_set_sha256 if execution_context else ""
            ),
            execution_sha256=(
                execution_context.execution_sha256 if execution_context else ""
            ),
            engine_version=engine_version,
            verified_refs=verification.accepted_refs,
            rejected_count=verification.rejected_count,
            rejection_codes=verification.rejection_codes,
            citation_status=verification.status,
            gates=gate_payload if isinstance(gate_payload, dict) else {},
        )
        output["quality"] = quality
        output.setdefault("finding_id", quality["finding_id"])
        enriched.append(output)
        status_counts[verification.status] = status_counts.get(verification.status, 0) + 1
        rejected_refs += verification.rejected_count
        for gate_name, gate_value in quality.get("gates", {}).items():
            if not isinstance(gate_value, dict):
                continue
            gate_status = str(gate_value.get("status", "") or "")
            if gate_status:
                gate_counts.setdefault(gate_name, {})[gate_status] = (
                    gate_counts.setdefault(gate_name, {}).get(gate_status, 0) + 1
                )

    return enriched, {
        "mode": mode,
        "total_findings": len(enriched),
        "citation_status": status_counts,
        "rejected_refs": rejected_refs,
        "downgraded_findings": downgraded_findings,
        "input_sha256": source_sha256,
        "input_set_sha256": (
            execution_context.input_set_sha256 if execution_context else ""
        ),
        "execution_sha256": (
            execution_context.execution_sha256 if execution_context else ""
        ),
        "engine_version": engine_version,
        "gate_status": gate_counts,
    }


def _now() -> str:
    return datetime.now().isoformat()


def _now_time_only() -> str:
    return datetime.now().strftime("%H:%M:%S")


def _make_progress_cb(review_id: str):
    """Return an on_progress callback that updates the registry entry's progress.

    Best-effort: swallows all exceptions so a callback bug can never break the review.
    Maintains a rolling `recent_events` list (last 15).
    """
    def _cb(payload: dict) -> None:
        try:
            entry = _REGISTRY.get(review_id)
            if entry is None:
                return
            prev = entry.get("progress") or {}
            events = list(prev.get("recent_events") or [])
            events.append({"t": _now_time_only(), "msg": str(payload.get("msg", ""))})
            events = events[-15:]
            entry["progress"] = {
                "stage": payload.get("stage", ""),
                "current_sheet": payload.get("current_sheet", ""),
                "llm_calls": payload.get("llm_calls", {}) or {},
                "findings_so_far": payload.get("findings_so_far") or {
                    "P0": 0, "P1": 0, "P2": 0, "total": 0,
                },
                "recent_events": events,
                "updated_at": _now(),
            }
        except Exception:
            pass
    return _cb


def get_status(review_id: str) -> Optional[dict]:
    entry = _REGISTRY.get(review_id)
    if entry is None:
        return None
    status = {
        "review_id": review_id,
        "status": entry["status"],
        "started_at": entry.get("started_at"),
        "source": entry.get("source"),
        "stats": entry.get("stats"),
        "progress": entry.get("progress"),
        "error": entry.get("error"),
    }
    if "artifact_status" in entry:
        status["artifact_status"] = entry["artifact_status"]
    if "artifact_error" in entry:
        status["artifact_error"] = entry["artifact_error"]
    return status


def list_running() -> List[str]:
    return [rid for rid, e in _REGISTRY.items() if e["status"] == "running"]


def cancel_all_running() -> int:
    """Cancel all currently-running reviews. Returns the number cancelled."""
    n = 0
    for rid, entry in list(_REGISTRY.items()):
        if entry["status"] != "running":
            continue
        task: Optional[asyncio.Task] = entry.get("task")
        if task is not None and not task.done():
            task.cancel()
        entry["status"] = "cancelled"
        n += 1
        _logger.info("cancelled running review %s", rid)
    return n


async def start_review(
    *,
    file_path: str,
    checkpoints_path: str = "",
    attachments_dir: str = "",
    attachments_preview_path: str = "",
    sheets: Optional[str] = None,
    source: str = "",
) -> str:
    """Start a background review. Returns the review_id immediately.

    Cancels any prior running review first (prevents stacking / quota burn).
    """
    cancel_all_running()
    review_id = uuid.uuid4().hex
    task = asyncio.create_task(_run_review(
        review_id=review_id,
        file_path=file_path,
        checkpoints_path=checkpoints_path,
        attachments_dir=attachments_dir,
        attachments_preview_path=attachments_preview_path,
        sheets=sheets,
        source=source,
    ))
    _REGISTRY[review_id] = {
        "status": "running",
        "task": task,
        "started_at": _now(),
        "source": source,
    }
    _logger.info("started background review %s source=%s", review_id, source)
    return review_id


async def _run_review(
    *,
    review_id: str,
    file_path: str,
    checkpoints_path: str,
    attachments_dir: str,
    attachments_preview_path: str,
    sheets: Optional[str],
    source: str,
) -> None:
    try:
        workspace_path = os.getenv("WORKSPACE_PATH", os.getcwd())
        store = ReviewArtifactStore(workspace_path=workspace_path)
        artifact_setup_error = None
        execution_context: ReviewExecutionContext | None = None
        policy_root: str | None = None
        # V1 must remain available even if a configured optional artifact
        # component is unavailable. The fallback catalog classifies every
        # finding explicitly as unclassified / human-review-required.
        assertion_catalog: AssertionCatalog = fallback_assertion_catalog()
        try:
            snapshot_paths = await asyncio.to_thread(
                store.snapshot_inputs,
                review_id,
                workpaper_path=file_path,
                checkpoints_path=checkpoints_path,
                attachments_dir=attachments_dir,
                attachments_preview_path=attachments_preview_path,
            )
            pinned_file_path = snapshot_paths["workpaper"]
            pinned_checkpoints_path = snapshot_paths.get("checkpoints", "")
            pinned_attachments_dir = snapshot_paths.get("attachments_dir", "")
            pinned_attachments_path = snapshot_paths.get("attachments_preview", "")
        except Exception as e:
            artifact_setup_error = f"{type(e).__name__}: {e}"
            _logger.exception("review input snapshot %s failed", review_id)
            entry = _REGISTRY.get(review_id)
            if entry is not None:
                entry["artifact_status"] = "error"
                entry["artifact_error"] = artifact_setup_error
            pinned_file_path = file_path
            pinned_checkpoints_path = checkpoints_path
            pinned_attachments_dir = attachments_dir
            pinned_attachments_path = attachments_preview_path

        if artifact_setup_error is None:
            try:
                policy_mode, policy_id, policy_version, policy_root = (
                    _stage_b_policy_config()
                )
                (
                    judgement_mode,
                    judgement_id,
                    judgement_version,
                    judgement_root,
                    _,
                ) = _stage_c_judgement_config()
                assertion_catalog_id, assertion_catalog_version, assertion_catalog_root = (
                    _assertion_catalog_config()
                )
                assertion_catalog = load_assertion_catalog(
                    pack_id=assertion_catalog_id,
                    version=assertion_catalog_version,
                    root=assertion_catalog_root,
                )
                policy_pack = (
                    PolicyPackRef(id=policy_id, version=policy_version)
                    if policy_mode == "shadow"
                    else None
                )
                judgement_policy_pack = (
                    PolicyPackRef(id=judgement_id, version=judgement_version)
                    if judgement_mode == "shadow"
                    else None
                )
                components: list[ExecutionComponentRef] = []
                components.append(
                    _assertion_catalog_component(
                        pack_id=assertion_catalog_id,
                        version=assertion_catalog.version,
                        root=assertion_catalog_root,
                    )
                )
                if policy_pack is not None:
                    components.append(
                        _policy_component_ref(
                            component_id="policy-pack",
                            pack_id=policy_id,
                            version=policy_version,
                            root=policy_root,
                        )
                    )
                if judgement_policy_pack is not None:
                    components.append(
                        _policy_component_ref(
                            component_id="judgement-policy-pack",
                            pack_id=judgement_id,
                            version=judgement_version,
                            root=judgement_root,
                        )
                    )
                inputs = build_input_files(
                    workpaper_path=pinned_file_path,
                    checkpoints_path=pinned_checkpoints_path,
                    attachments_dir=pinned_attachments_dir,
                    attachments_preview_path=pinned_attachments_path,
                )
                execution_context = build_review_execution_context(
                    review_id=review_id,
                    source=source,
                    requested_sheets=_requested_sheet_names(sheets),
                    inputs=inputs,
                    snapshot_paths={
                        key: value
                        for key, value in {
                            "workpaper": pinned_file_path,
                            "checkpoints": pinned_checkpoints_path,
                            "attachments_dir": pinned_attachments_dir,
                            "attachments_preview": pinned_attachments_path,
                        }.items()
                        if value
                    },
                    policy_pack=policy_pack,
                    judgement_policy_pack=judgement_policy_pack,
                    engine_version=_execution_engine_version(policy_mode),
                    components=components,
                    runtime_config=capture_runtime_config(),
                )
                # Write the manifest before V1 reads or sends the frozen input.
                store.begin(execution_context.manifest)
            except Exception as exc:
                artifact_setup_error = f"{type(exc).__name__}: {exc}"
                _logger.exception("review execution context %s failed", review_id)
                entry = _REGISTRY.get(review_id)
                if entry is not None:
                    entry["artifact_status"] = "error"
                    entry["artifact_error"] = artifact_setup_error

        wb = openpyxl.load_workbook(pinned_file_path, data_only=True)
        checkpoints = (
            load_checkpoints_xlsx(pinned_checkpoints_path)
            if pinned_checkpoints_path
            else {}
        )
        if pinned_attachments_dir:
            attachments = await asyncio.to_thread(
                build_attachment_index, pinned_attachments_dir
            )
        elif pinned_attachments_path:
            attachments = await asyncio.to_thread(
                load_attachments_preview_xlsx, pinned_attachments_path
            )
        else:
            attachments = {}
        # Stash review_id on the attachments dict so downstream helpers (OCR
        # artifact persistence, evidence snapshots) can locate the right
        # review artifact directory.
        if attachments:
            attachments["review_id"] = review_id
        LLM_CALL_STATS.clear()
        LLM_CALL_STATS.clear()
        llm = get_review_llm()
        findings, stats = await run_review(
            wb=wb,
            checkpoints=checkpoints,
            attachments=attachments,
            attachments_preview=attachments,
            sheets=sheets,
            llm=llm,
            on_progress=_make_progress_cb(review_id),
            assertion_catalog=assertion_catalog,
        )
        quality_input_findings = [dict(item) for item in findings]
        legacy_findings = [
            {key: value for key, value in item.items() if key != "quality_gates"}
            for item in quality_input_findings
        ]
        try:
            findings, quality_stats = _attach_result_quality(
                findings=quality_input_findings,
                workbook=wb,
                file_path=pinned_file_path,
                attachments=attachments,
                execution_context=execution_context,
            )
            if quality_stats.get("mode") != "off":
                findings, grouping_stats = enrich_finding_quality(
                    findings,
                    input_sha256=str(quality_stats.get("input_sha256", "") or ""),
                )
                quality_stats.update(grouping_stats)
        except Exception as exc:
            # Quality metadata is additive. A malformed/temporarily
            # unavailable provenance index must not turn a completed V1 review
            # into a failed review; the shadow artifact records its own error.
            _logger.exception("review quality capture %s failed", review_id)
            findings = legacy_findings
            quality_stats = {
                "mode": os.getenv("REVIEW_RESULT_QUALITY_MODE", "shadow"),
                "total_findings": len(findings),
                "citation_status": {},
                "rejected_refs": 0,
                "downgraded_findings": 0,
                "error": f"{type(exc).__name__}: {exc}",
            }
        stats["quality"] = quality_stats
        save_findings(review_id, findings, stats, source=source)
        entry = _REGISTRY.get(review_id)
        if entry is not None:
            entry["status"] = "completed"
            entry["stats"] = stats
            if artifact_setup_error is None and execution_context is not None:
                entry["artifact_status"] = "pending"
                entry["shadow_task"] = asyncio.create_task(
                    _capture_shadow_artifact(
                        review_id=review_id,
                        file_path=pinned_file_path,
                        checkpoints_path=pinned_checkpoints_path,
                        attachments_dir=pinned_attachments_dir,
                        attachments_preview_path=pinned_attachments_path,
                        sheets=sheets,
                        source=source,
                        findings=legacy_findings,
                        stats=stats,
                        llm=llm,
                        attachments=attachments,
                        execution_context=execution_context,
                        policy_root=policy_root,
                    )
                )
        _logger.info("background review %s completed: %d findings", review_id, len(findings))
    except asyncio.CancelledError:
        entry = _REGISTRY.get(review_id)
        if entry is not None:
            entry["status"] = "cancelled"
        _logger.info("background review %s cancelled", review_id)
        raise
    except Exception as e:
        entry = _REGISTRY.get(review_id)
        if entry is not None:
            entry["status"] = "error"
            entry["error"] = f"{type(e).__name__}: {e}"
        _logger.exception("background review %s failed", review_id)


async def _capture_shadow_artifact(
    *,
    review_id: str,
    file_path: str,
    checkpoints_path: str,
    attachments_dir: str = "",
    attachments_preview_path: str,
    sheets: Optional[str],
    source: str,
    findings: list[dict],
    stats: dict,
    llm=None,
    attachments: Optional[dict[str, object]] = None,
    execution_context: ReviewExecutionContext | None = None,
    policy_root: str | None = None,
) -> None:
    """Persist a V2 artifact without affecting the completed V1 review."""
    entry = _REGISTRY.get(review_id)
    if entry is not None:
        entry["artifact_status"] = "running"
    workspace_path = os.getenv("WORKSPACE_PATH", os.getcwd())

    try:
        error = await asyncio.to_thread(
            _write_shadow_artifact,
            review_id=review_id,
            file_path=file_path,
            checkpoints_path=checkpoints_path,
            attachments_dir=attachments_dir,
            attachments_preview_path=attachments_preview_path,
            sheets=sheets,
            source=source,
            findings=findings,
            stats=stats,
            workspace_path=workspace_path,
            execution_context=execution_context,
            policy_root=policy_root,
        )
    except Exception as e:
        error = f"{type(e).__name__}: {e}"
        _logger.exception("shadow artifact capture %s could not start", review_id)

    if error is None:
        try:
            error = await _capture_stage_c_shadow(
                review_id=review_id,
                file_path=file_path,
                sheets=sheets,
                llm=llm,
                attachments=attachments,
                workspace_path=workspace_path,
            )
        except Exception as e:
            error = f"{type(e).__name__}: {e}"
            _logger.exception("stage c shadow capture %s failed", review_id)

        store = ReviewArtifactStore(workspace_path=workspace_path)
        if error is None:
            try:
                await asyncio.to_thread(store.complete, review_id)
            except Exception as e:
                error = f"{type(e).__name__}: {e}"
                _logger.exception("shadow artifact completion %s failed", review_id)
        if error is not None:
            try:
                await asyncio.to_thread(store.fail, review_id, error)
            except Exception:
                _logger.exception("shadow artifact failure record %s failed", review_id)

    entry = _REGISTRY.get(review_id)
    if entry is None:
        return
    if error is None:
        entry["artifact_status"] = "completed"
    else:
        entry["artifact_status"] = "error"
        entry["artifact_error"] = error


def _load_stage_c_snapshot(file_path: str):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    graph = build_evidence_graph(wb, source_sha256=sha256_file(file_path))
    return wb, graph


async def _capture_stage_c_shadow(
    *,
    review_id: str,
    file_path: str,
    sheets: Optional[str],
    llm,
    attachments: Optional[dict[str, object]],
    workspace_path: str,
) -> Optional[str]:
    """Run opt-in Stage-C judgement after V1 and Stage-B capture complete."""
    mode, pack_id, pack_version, pack_root, max_requests = _stage_c_judgement_config()
    if mode == "off":
        return None
    if llm is None:
        return "RuntimeError: Stage C LLM is unavailable"

    store = ReviewArtifactStore(workspace_path=workspace_path)
    try:
        policy_pack = load_policy_pack(
            pack_id=pack_id,
            version=pack_version,
            root=pack_root,
        )
        wb, graph = await asyncio.to_thread(_load_stage_c_snapshot, file_path)
        requests = build_judgement_requests(
            workbook=wb,
            evidence_graph=graph,
            policy_pack=policy_pack,
            sheets=sheets,
            attachments=attachments,
            max_requests=max_requests,
        )
        executions = await execute_judgement_requests(requests, llm=llm)
        engine_version = os.getenv(
            "REVIEW_ENGINE_VERSION", "stage-b-policy-shadow"
        ).strip()
        v2_findings = build_v2_findings(
            requests=requests,
            executions=executions,
            policy_pack=policy_pack,
            engine_version=engine_version,
        )
        policy_metadata = {"id": policy_pack.id, "version": policy_pack.version}
        judgement_payload = {
            "schema_version": "stage-c-judgements/1",
            "engine_version": engine_version,
            "source_sha256": graph.source_sha256,
            "policy_pack": policy_metadata,
            "requests": [request.model_dump(mode="json") for request in requests],
            "results": [result.model_dump(mode="json") for result in executions],
            "stats": {
                "requests": len(requests),
                "results": len(executions),
                "by_verification_status": _count_values(
                    [result.verification_status for result in executions]
                ),
            },
        }
        v2_payload = {
            "schema_version": "stage-c-v2-findings/1",
            "engine_version": engine_version,
            "source_sha256": graph.source_sha256,
            "policy_pack": policy_metadata,
            "findings": v2_findings,
            "v1_projection": project_v2_findings_to_v1(v2_findings),
            "stats": {
                "total_findings": len(v2_findings),
                "by_status": _count_values(
                    [str(item.get("status", "")) for item in v2_findings]
                ),
            },
        }
        await asyncio.to_thread(
            store.write_judgements, review_id, judgement_payload
        )
        await asyncio.to_thread(store.write_v2_findings, review_id, v2_payload)
        legacy_payload = load_findings(review_id) or {}
        comparison = compare_finding_sets(
            legacy_payload.get("findings", [])
            if isinstance(legacy_payload, dict)
            else [],
            v2_findings,
        )
        await asyncio.to_thread(store.write_comparison, review_id, comparison)
        return None
    except Exception as e:
        _logger.exception("stage c shadow capture %s failed", review_id)
        return f"{type(e).__name__}: {e}"


def _count_values(values: list[str]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for value in values:
        counts[value] = counts.get(value, 0) + 1
    return counts


def _write_shadow_artifact(
    *,
    review_id: str,
    file_path: str,
    checkpoints_path: str,
    attachments_dir: str = "",
    attachments_preview_path: str,
    sheets: Optional[str],
    source: str,
    findings: list[dict],
    stats: dict,
    workspace_path: str,
    execution_context: ReviewExecutionContext | None = None,
    policy_root: str | None = None,
) -> Optional[str]:
    """Run synchronous artifact capture off the event-loop thread."""
    store = ReviewArtifactStore(workspace_path=workspace_path)
    try:
        if execution_context is None:
            policy_mode, policy_id, policy_version, legacy_policy_root = (
                _stage_b_policy_config()
            )
            judgement_mode, judgement_id, judgement_version, _, _ = (
                _stage_c_judgement_config()
            )
            engine_version = _execution_engine_version(policy_mode)
            inputs = build_input_files(
                workpaper_path=file_path,
                checkpoints_path=checkpoints_path,
                attachments_dir=attachments_dir,
                attachments_preview_path=attachments_preview_path,
            )
            manifest = ReviewManifest(
                review_id=review_id,
                source=source,
                requested_sheets=_requested_sheet_names(sheets),
                inputs=inputs,
                policy_pack=(
                    PolicyPackRef(id=policy_id, version=policy_version)
                    if policy_mode == "shadow"
                    else None
                ),
                judgement_policy_pack=(
                    PolicyPackRef(id=judgement_id, version=judgement_version)
                    if judgement_mode == "shadow"
                    else None
                ),
                engine_version=engine_version,
            )
            store.begin(manifest)
            active_policy_root = legacy_policy_root
        else:
            manifest = execution_context.manifest
            inputs = manifest.inputs
            engine_version = manifest.engine_version
            active_policy_root = policy_root

        wb = openpyxl.load_workbook(file_path, data_only=False)
        source_sha256 = next(
            (item.sha256 for item in inputs if item.role == "workpaper"), ""
        )
        graph = build_evidence_graph(wb, source_sha256=source_sha256)
        store.write_evidence(review_id, graph)
        store.write_v1_findings(review_id, findings, stats)
        if manifest.policy_pack is not None:
            policy_pack = load_policy_pack(
                pack_id=manifest.policy_pack.id,
                version=manifest.policy_pack.version,
                root=active_policy_root,
            )
            plan = build_review_plan(
                wb,
                graph,
                policy_pack,
                sheets=sheets,
                engine_version=engine_version,
            )
            store.write_review_plan(review_id, plan.to_dict())
            policy_findings = execute_policy_plan(plan, policy_pack)
            store.write_policy_findings(review_id, policy_findings)
        return None
    except Exception as e:
        error = f"{type(e).__name__}: {e}"
        _logger.exception("shadow artifact capture %s failed", review_id)
        try:
            store.fail(review_id, error)
        except Exception:
            _logger.exception("shadow artifact failure record %s failed", review_id)
        return error
