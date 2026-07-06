"""Evidence-vs-step consistency LLM check (ported from analyze_excel.py, async)."""
import asyncio
import json
import os
from typing import Dict, List

from openpyxl.utils import get_column_letter

from review.attachments import _extract_attachment_refs, _match_preview_items
from review.constants import EVIDENCE_KEYWORDS
from review.excel_utils import _detect_layout, _get_cell_value, _normalize_sheet_id, _truncate
from review.llm import _llm_request_json_list, _llm_stat
from review.models import AttachmentPreviewItem, Finding, _SEVERITY_FROM_CHINESE
from review.validation import _verify_evidence_refs

_EXCERPT_MAX_LEN = 2000


async def _llm_check_evidence_vs_steps(
    *,
    llm,
    ws_title: str,
    ws,
    attachments_preview: Dict[str, object],
    batch_size: int = 6,
    sleep_seconds: float = 0.2,
) -> List[Finding]:
    if not attachments_preview:
        return []
    findings: List[Finding] = []
    header_row, standard_col, execution_cols = _detect_layout(ws)
    if header_row is None or standard_col <= 0 or not execution_cols:
        return findings
    start_row = max(5, (header_row or 1) + 2)

    cases: List[Dict[str, object]] = []
    next_id = 1
    for row in range(start_row, (ws.max_row or 0) + 1):
        a_cell = f"{get_column_letter(standard_col)}{row}"
        a_text = _get_cell_value(ws, a_cell)
        if not a_text:
            continue
        for c in execution_cols:
            c_cell = f"{get_column_letter(c)}{row}"
            c_text = _get_cell_value(ws, c_cell)
            if not c_text:
                continue
            filenames, rel_paths, indices = _extract_attachment_refs(c_text)
            if not filenames and not rel_paths and not indices and not any(k in c_text for k in EVIDENCE_KEYWORDS):
                continue
            matched, missing = _match_preview_items(
                attachments_preview, filenames=filenames, rel_paths=rel_paths, indices=indices,
            )
            if (
                not matched
                and not filenames
                and not rel_paths
                and not indices
                and any(k in c_text for k in EVIDENCE_KEYWORDS)
                and isinstance(attachments_preview.get("by_sheet_norm"), dict)
            ):
                pool = attachments_preview.get("by_sheet_norm", {}).get(_normalize_sheet_id(ws_title))
                if isinstance(pool, list) and pool:
                    matched = [it for it in pool if isinstance(it, AttachmentPreviewItem)][:10]
            if missing and (filenames or rel_paths or indices):
                findings.append(
                    Finding(
                        issue_type="附件证据编号/文件未匹配（可能引用错误）",
                        severity="P1",
                        sheet=ws_title,
                        cell=c_cell,
                        snippet=_truncate(c_text, 220),
                        basis=_truncate(
                            "标准程序: " + _truncate(a_text, 160)
                            + "\n引用未匹配: " + "、".join(sorted({m for m in missing if m})),
                            1200,
                        ),
                        suggestion="核对附件编号/命名/路径与证据清单是否一致；必要时在底稿中给出可复核的相对路径或文件名，并补充证据来源说明。",
                    )
                )
            if matched:
                evidences: List[Dict[str, str]] = []
                for it in matched[:10]:
                    evidences.append({
                        "path": str(it.rel_path or it.filename or ""),
                        "status": str(it.status or ""),
                        "description": _truncate(str(it.description or "").replace("\r\n", "\n").replace("\r", "\n"), 1200),
                    })
                cases.append({
                    "id": next_id,
                    "sheet": ws_title,
                    "row": row,
                    "standard_cell": a_cell,
                    "execution_cell": c_cell,
                    "standard_text": a_text,
                    "execution_text": c_text,
                    "evidences": evidences,
                })
                next_id += 1

    if not cases:
        return findings

    max_items = 0
    try:
        max_items = int(os.getenv("LLM_EVIDENCE_STEPS_MAX_ITEMS", "0") or 0)
    except Exception:
        max_items = 0
    if max_items and max_items > 0 and len(cases) > max_items:
        def _sample_evenly(seq, k):
            if k <= 0 or len(seq) <= k:
                return list(seq)
            step = float(len(seq)) / float(k)
            picked = []
            used = set()
            for i in range(k):
                idx = int(i * step)
                if idx < 0:
                    idx = 0
                if idx >= len(seq):
                    idx = len(seq) - 1
                if idx in used:
                    continue
                used.add(idx)
                picked.append(seq[idx])
            while len(picked) < k and len(picked) < len(seq):
                idx = len(picked)
                if idx in used or idx >= len(seq):
                    break
                used.add(idx)
                picked.append(seq[idx])
            return picked

        _llm_stat(f"evidence_steps:{ws_title}", "sampled", 1)
        original = len(cases)
        cases = _sample_evenly(cases, max_items)
        findings.append(
            Finding(
                issue_type="证据-步骤一致性抽样复核（为控制LLM调用规模）",
                severity="P2",
                sheet=ws_title,
                cell=None,
                snippet="",
                basis=_truncate(f"原匹配记录数: {original}；本次LLM抽样复核: {len(cases)}（可通过环境变量LLM_EVIDENCE_STEPS_MAX_ITEMS调整）", 1200),
                suggestion="如需全量复核，请增大LLM_EVIDENCE_STEPS_MAX_ITEMS，或先缩小Sheet范围（-s）再运行。",
            )
        )

    system_prompt = (
        "你是一名严格的IT审计/财务审计质量复核专家。\n"
        "你将收到多条“标准审计程序/执行审计程序”记录，以及执行中引用的附件证据预览信息（附件路径/状态/内容描述）。\n"
        "你的任务：判断证据是否与审计步骤匹配、是否足以支持执行描述与测试结论（如“已验证/无异常/符合”等）。\n"
        "重点关注：\n"
        "1) 证据是否能证明该步骤要验证的控制点/属性（而非无关截图）。\n"
        "2) 执行描述是否仅列附件但未说明核查点/结论依据。\n"
        "3) 证据状态异常/描述模糊时，提出需要补充的证据类型与下一步动作。\n"
        "输出要求：必须输出严格JSON对象：{\"results\": [...]}。\n"
        "results中每个元素对应输入id，且必须包含字段：\n"
        "- id: 整数\n"
        "- status: \"pass\"/\"fail\"/\"unknown\" (无问题/有问题/不确定)\n"
        "- conclusion: 一句话结论\n"
        "- reasons: 字符串数组，2-5条要点\n"
        "- evidence_refs: 数组，每个元素含 {sheet, cell_or_range, attachment(可选), excerpt(原文摘录)}。excerpt必须逐字来自执行/标准/附件描述。\n"
        "  * status=fail时必须至少1个evidence_ref；无法引用时status必须为unknown\n"
        "- severity: \"P0\"/\"P1\"/\"P2\"（当status!=pass时必填）\n"
        "- risk_type: \"覆盖性\"/\"一致性\"/\"证据不足\"/\"方法性\"/\"逻辑性\"/\"跨字段一致性\"之一\n"
        "- fix_suggestion: 对象，含 {missing_field, supplement_explanation, required_evidence_type}\n"
        "- unknown_reason: 当status=unknown时必填，≥10字符\n"
        "向后兼容字段（可同时输出）：basis, suggestion, issue_type, missing_evidence\n"
        "不要输出Markdown代码块，不要输出多余文字。"
    )

    id_to_case = {int(c.get("id")): c for c in cases if isinstance(c.get("id"), int)}
    for start in range(0, len(cases), max(1, int(batch_size))):
        chunk = cases[start: start + max(1, int(batch_size))]
        payload = {"sheet": ws_title, "items": chunk}
        user_prompt = "请按要求逐条复核以下记录：\n" + json.dumps(payload, ensure_ascii=False, indent=2)

        def _consume_results(objs: List[object]) -> None:
            for obj in objs:
                if not isinstance(obj, dict):
                    continue
                rid = obj.get("id")
                if not isinstance(rid, int):
                    try:
                        rid = int(str(rid).strip())
                    except Exception:
                        continue
                raw_status = str(obj.get("status", "")).strip()
                if raw_status == "无问题":
                    status = "pass"
                elif raw_status == "有问题":
                    status = "fail"
                elif raw_status == "不确定":
                    status = "unknown"
                else:
                    status = raw_status
                if status == "pass":
                    continue
                case = id_to_case.get(int(rid))
                if not case:
                    continue
                c_cell = str(case.get("execution_cell") or "").strip() or None
                c_text = str(case.get("execution_text") or "")
                a_text = str(case.get("standard_text") or "")
                raw_sev = str(obj.get("severity", "")).strip()
                severity = _SEVERITY_FROM_CHINESE.get(raw_sev, raw_sev)
                if severity not in ("P0", "P1", "P2"):
                    severity = "P1" if status != "unknown" else "P2"
                issue_type = str(obj.get("issue_type", "")).strip() or (
                    "证据与审计步骤不匹配" if status == "fail" else "证据与审计步骤信息不足/不确定"
                )
                basis = str(obj.get("basis", "")).strip()
                suggestion = str(obj.get("suggestion", "")).strip()
                conclusion = str(obj.get("conclusion", "")).strip()
                reasons_raw = obj.get("reasons", [])
                reasons_list = [str(r).strip() for r in reasons_raw if r] if isinstance(reasons_raw, list) else []
                risk_type = str(obj.get("risk_type", "")).strip()
                unknown_reason = str(obj.get("unknown_reason", "")).strip()
                fix_suggestion_obj = obj.get("fix_suggestion") or {}
                if not isinstance(fix_suggestion_obj, dict):
                    fix_suggestion_obj = {}
                missing_evidence = obj.get("missing_evidence", [])
                missing_text = ""
                if isinstance(missing_evidence, list) and missing_evidence:
                    missing_text = "缺失证据: " + "、".join(str(x).strip() for x in missing_evidence if str(x).strip())

                evidence_refs_list: List[dict] = []
                raw_refs = obj.get("evidence_refs")
                if isinstance(raw_refs, list) and raw_refs:
                    for ref in raw_refs:
                        if not isinstance(ref, dict):
                            continue
                        ev_cell = str(ref.get("cell_or_range", "")).strip()
                        ev_sheet = str(ref.get("sheet", "")).strip() or ws_title
                        ev_attachment = str(ref.get("attachment", "")).strip()
                        ev_excerpt = str(ref.get("excerpt", "")).strip()
                        if ev_cell or ev_excerpt:
                            evidence_refs_list.append({
                                "sheet": ev_sheet,
                                "cell_or_range": ev_cell,
                                "attachment": ev_attachment,
                                "excerpt": ev_excerpt,
                            })
                if not evidence_refs_list and c_cell:
                    evidence_refs_list.append({
                        "sheet": ws_title,
                        "cell_or_range": c_cell,
                        "excerpt": c_text[:_EXCERPT_MAX_LEN] if c_text else "",
                    })
                evidence_refs_list = _verify_evidence_refs(evidence_refs_list, ws)
                if status == "fail" and not evidence_refs_list:
                    status = "unknown"
                    unknown_reason = unknown_reason or "无法引用原始证据佐证该判定，降级为不确定"
                    severity = "P2"

                if not fix_suggestion_obj.get("required_evidence_type") and missing_evidence:
                    if isinstance(missing_evidence, list) and missing_evidence:
                        fix_suggestion_obj["required_evidence_type"] = "、".join(str(x).strip() for x in missing_evidence if str(x).strip())[:300]
                if not fix_suggestion_obj.get("supplement_explanation") and suggestion:
                    fix_suggestion_obj["supplement_explanation"] = suggestion[:300]

                basis_parts: List[str] = []
                if a_text:
                    basis_parts.append("标准程序: " + _truncate(a_text, 260))
                if conclusion:
                    basis_parts.append("结论: " + conclusion)
                if reasons_list:
                    basis_parts.append("理由: " + " | ".join(reasons_list[:5]))
                if basis:
                    basis_parts.append("LLM依据: " + basis)
                if missing_text:
                    basis_parts.append(missing_text)
                if evidence_refs_list:
                    refs_text = "; ".join(
                        f"{r.get('cell_or_range', '')}: {r.get('excerpt', '')[:200]}"
                        for r in evidence_refs_list[:3] if r.get('excerpt')
                    )
                    if refs_text:
                        basis_parts.append("引用: " + refs_text)
                if unknown_reason:
                    basis_parts.append("不确定原因: " + unknown_reason)
                final_basis = "\n".join(p for p in basis_parts if p).strip()

                findings.append(
                    Finding(
                        issue_type="LLM判定：证据-步骤一致性-" + issue_type,
                        severity=severity,
                        sheet=ws_title,
                        cell=c_cell,
                        snippet=_truncate(c_text, 220),
                        basis=_truncate(final_basis or "LLM判定存在问题/不确定", 3000),
                        suggestion=_truncate(
                            suggestion
                            or "补充与该审计步骤直接对应的截图/导出清单/日志/审批等证据，并在底稿中写明「证据→核查点→结论」的对应关系。",
                            1200,
                        ),
                        status=status,
                        risk_type=risk_type or ("证据不足" if status == "fail" else ""),
                        evidence_refs=json.dumps(evidence_refs_list, ensure_ascii=False),
                        conclusion=conclusion,
                        reasons=json.dumps(reasons_list, ensure_ascii=False),
                        fix_suggestion_detail=json.dumps(fix_suggestion_obj, ensure_ascii=False),
                        unknown_reason=unknown_reason,
                    )
                )

        stage = f"evidence_steps:{ws_title}"
        parsed, last_error = await _llm_request_json_list(
            llm=llm, system_prompt=system_prompt, user_prompt=user_prompt,
            stage=stage, max_attempts=3,
        )
        if parsed is not None:
            _consume_results(parsed)
        else:
            if len(chunk) > 1:
                _llm_stat(stage, "fallback_single", 1)
                for item in chunk:
                    payload1 = {"sheet": ws_title, "items": [item]}
                    user_prompt1 = "请按要求逐条复核以下记录：\n" + json.dumps(payload1, ensure_ascii=False, indent=2)
                    parsed1, err1 = await _llm_request_json_list(
                        llm=llm, system_prompt=system_prompt, user_prompt=user_prompt1,
                        stage=stage, max_attempts=3,
                    )
                    if parsed1 is not None:
                        _consume_results(parsed1)
                    else:
                        findings.append(_evidence_steps_failure_finding(ws_title, err1))
            else:
                findings.append(_evidence_steps_failure_finding(ws_title, last_error))

        if sleep_seconds and sleep_seconds > 0:
            await asyncio.sleep(float(sleep_seconds))

    return findings


def _evidence_steps_failure_finding(ws_title: str, error) -> Finding:
    return Finding(
        issue_type="LLM判定：证据-步骤一致性复核失败",
        severity="P1",
        sheet=ws_title,
        cell=None,
        snippet="",
        basis=_truncate(f"LLM调用失败: {error}", 1200),
        suggestion="检查LLM接口配置（.env）、网络连通性；必要时减少检查行数或缩短证据描述后重试。",
        status="unknown",
        unknown_reason="LLM调用失败，无法复核",
        risk_type="证据不足",
    )
