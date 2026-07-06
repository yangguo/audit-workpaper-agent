"""openpyxl helpers for the review engine (ported from analyze_excel.py)."""
from typing import Iterable, List, Optional, Tuple

from openpyxl.utils import get_column_letter

from review.models import _EXCERPT_CONSTRUCTED_MARKER


def _is_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _get_cell_value(ws, cell_ref: str) -> Optional[str]:
    """Get a cell value, resolving merged-cell anchors to the top-left value."""
    cell = ws[cell_ref]
    value = cell.value
    if value is None and ws.merged_cells.ranges:
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                break
    if _is_empty(value):
        return None
    return str(value).strip()


def _get_cell_text(ws, cell_ref: str) -> str:
    """Safely get cell text, stripping the constructed-excerpt marker suffix."""
    if not cell_ref or not ws:
        return ""
    clean_ref = cell_ref.replace(_EXCERPT_CONSTRUCTED_MARKER, "").strip()
    if not clean_ref:
        return ""
    try:
        cell = ws[clean_ref]
        val = cell.value
        return str(val).strip() if val is not None else ""
    except Exception:
        return ""


def _truncate(text: str, n: int = 160) -> str:
    s = (text or "").strip()
    if len(s) <= n:
        return s
    return s[:n] + "..."


def _detect_layout(ws) -> Tuple[Optional[int], int, List[int]]:
    """Detect the header row, the standard-program column and exec columns."""
    max_scan_row = min(ws.max_row or 0, 40)
    max_scan_col = min(ws.max_column or 0, 30)
    for r in range(1, max_scan_row + 1):
        standard_col = None
        exec_cols: List[int] = []
        for c in range(1, max_scan_col + 1):
            v = _get_cell_value(ws, f"{get_column_letter(c)}{r}")
            if not v:
                continue
            if ("标准" in v and "审计程序" in v) or ("标准审计程序" in v):
                standard_col = c
            if "执行" in v and "审计程序" in v:
                exec_cols.append(c)
        if standard_col and exec_cols:
            exec_cols = sorted({c for c in exec_cols if c != standard_col})
            if exec_cols:
                return r, standard_col, exec_cols
    return None, 0, []


def _extract_sheet_text_cells(ws) -> Iterable[Tuple[str, str]]:
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if _is_empty(cell.value):
                continue
            value = cell.value
            if isinstance(value, str):
                text = value.strip()
            else:
                text = str(value).strip()
            if not text:
                continue
            yield cell.coordinate, text


def _build_sheet_text_for_llm(ws, max_cells: int = 260, max_chars: int = 24000) -> str:
    parts: List[str] = []
    total_chars = 0
    for coord, text in _extract_sheet_text_cells(ws):
        line = f"{coord}: {text}"
        if len(parts) >= max_cells:
            break
        if total_chars + len(line) + 1 > max_chars:
            break
        parts.append(line)
        total_chars += len(line) + 1
    return "\n".join(parts)


def _normalize_sheet_id(text: str) -> str:
    s = (text or "").strip().upper()
    s = s.replace(" ", "").replace("-", "").replace("_", "")
    return s
