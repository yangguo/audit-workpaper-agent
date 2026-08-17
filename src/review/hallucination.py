"""Hallucination-reduction helpers: cross-validation and adversarial challenge.

Ported from analyze_excel.py. _challenge_finding_with_llm is adapted to async
over the project's ChatOpenAI-based llm helper.
"""
import json
import os
import re
from typing import Any, List, Mapping, Optional

from openpyxl.utils import get_column_letter

from review.excel_utils import _get_cell_text
from review.evidence_agent import _excerpt_grounded
from review.llm import _llm_chat
from review.validation import _excerpt_matches

_EXCEPTION_FLAG_TOKENS = ("是", "有异常", "Y", "异常", "缺陷", "未通过")
_CHALLENGER_FULL_TEXT_DEFAULT = True
_CHALLENGER_PER_REF = 2000
_CHALLENGER_TOTAL = 6000
_CHALLENGER_RADIUS = 800


def _finding_value(finding: Any, name: str, default: Any = "") -> Any:
    """Read a legacy Finding object or the immutable mapping used by gates."""

    if isinstance(finding, Mapping):
        return finding.get(name, default)
    return getattr(finding, name, default)


def _challenger_full_text_enabled() -> bool:
    raw = os.getenv("REVIEW_CHALLENGER_FULL_TEXT", "")
    if not raw:
        return _CHALLENGER_FULL_TEXT_DEFAULT
    return raw.strip().lower() not in {"0", "false", "no", "off"}


def _build_attachment_evidence_snippets(refs: List[dict], per_ref: int = _CHALLENGER_PER_REF,
                                         total_limit: int = _CHALLENGER_TOTAL,
                                         radius: int = _CHALLENGER_RADIUS) -> List[str]:
    """Pull bounded evidence snippets for both embedded-media OCR and real-attachments.

    - For ``.embedded_media/`` paths: OCR ``full_text`` (HTML markup preserved)
      centred on the agent's cited excerpt.
    - For real attachment paths: directly-extracted ``attachment_text`` so
      xlsx/docx contents reach the challenger LLM too.
    """
    out: List[str] = []
    total = 0
    for r in refs:
        if not isinstance(r, dict):
            continue
        attachment = str(r.get("attachment", "") or "").strip()
        if not attachment:
            continue
        body = ""
        is_ocr = attachment.startswith(".embedded_media/")
        if is_ocr:
            body = str(r.get("full_text", "") or "").strip()
        else:
            body = str(r.get("attachment_text", "") or "").strip()
        if not body:
            continue
        excerpt = str(r.get("excerpt", "") or "").strip()
        if excerpt and len(body) > per_ref:
            anchor = excerpt[:24]
            position = body.find(anchor)
            if position < 0 and len(excerpt) >= 16:
                position = body.find(excerpt[:16])
            if position >= 0:
                start = max(0, position - radius)
                end = min(len(body), position + per_ref - radius)
                body = body[start:end]
            else:
                body = body[:per_ref]
        else:
            body = body[:per_ref]
        body = re.sub(r"\s+", " ", body).strip()
        if not body:
            continue
        remaining = max(0, total_limit - total)
        if remaining <= 0:
            break
        if len(body) > remaining:
            body = body[:remaining]
        kind = "OCR 抓到的实际内容" if is_ocr else "附件内可直接读取的文本"
        out.append(f"[附件 {attachment}] ({kind})\n{body}")
        total += len(body) + len(attachment) + 32
    return out


def _cross_validate_finding(finding, wb) -> List[str]:
    """Deterministic cross-checks against the workbook.

    Returns issue codes; empty means no issues found.
    """
    issues: List[str] = []
    sheet = _finding_value(finding, "sheet")
    cell_refs: List[str] = []
    cell = _finding_value(finding, "cell")
    if cell:
        for c in str(cell).split(","):
            c = c.strip()
            if c:
                cell_refs.append(c)
    try:
        evidence_refs = _finding_value(finding, "evidence_refs", [])
        refs = (
            json.loads(evidence_refs)
            if isinstance(evidence_refs, str) and evidence_refs
            else evidence_refs
        )
    except Exception:
        refs = []
    if isinstance(refs, list):
        for r in refs:
            if isinstance(r, dict) and r.get("cell_or_range"):
                cell_refs.append(r["cell_or_range"])

    if not wb or sheet not in wb.sheetnames:
        return issues
    ws = wb[sheet]

    if _finding_value(finding, "status") == "pass":
        for c in cell_refs:
            txt = _get_cell_text(ws, c)
            if txt and any(tok in txt for tok in _EXCEPTION_FLAG_TOKENS):
                issues.append("exception_flag_contradicts_pass")
                break

    for r in refs if isinstance(refs, list) else []:
        if not isinstance(r, dict):
            continue
        cell = r.get("cell_or_range", "")
        excerpt = r.get("excerpt", "")
        if cell and excerpt:
            actual = _get_cell_text(ws, cell)
            if actual and not _excerpt_matches(excerpt, actual):
                issues.append("evidence_excerpt_mismatch")
                break

    # Ground evidence excerpts against cached OCR full text so an agent that
    # hallucinates a parameter value (e.g. quoting "FAILED_LOGIN_ATTEMPTS=60"
    # when the screenshot actually shows 10) is caught even when the workbook
    # cells can't disprove it.
    for r in refs if isinstance(refs, list) else []:
        if not isinstance(r, dict):
            continue
        attachment = r.get("attachment", "") or ""
        full_text = r.get("full_text", "") or ""
        excerpt = r.get("excerpt", "") or ""
        if not attachment.startswith(".embedded_media/"):
            continue
        if not full_text or not excerpt:
            continue
        if not _excerpt_grounded(excerpt, full_text):
            issues.append("evidence_excerpt_mismatch_attachment")
            break

    if (
        _finding_value(finding, "status") == "fail"
        and _finding_value(finding, "severity") == "P0"
    ):
        if not refs:
            issues.append("high_severity_no_evidence")

    return issues


def _build_minimal_context(finding, ws, max_chars: int = 2000) -> str:
    """Build a minimal context (500-2000 chars) for an LLM re-review.

    When ``REVIEW_CHALLENGER_FULL_TEXT`` is enabled (default), also injects
    bounded OCR snippets from each embedded-media evidence_ref so the
    adversarial challenger can verify the finding against actual screenshot
    content instead of guessing from workbook cells alone.
    """
    if not ws:
        return ""
    parts: List[str] = []
    try:
        refs = json.loads(finding.evidence_refs) if finding.evidence_refs else []
    except Exception:
        refs = []
    if not isinstance(refs, list):
        refs = []
    if _challenger_full_text_enabled():
        parts.extend(_build_attachment_evidence_snippets(refs))
    target_cells: List[str] = []
    for r in refs[:6]:
        if isinstance(r, dict) and r.get("cell_or_range"):
            target_cells.append(r["cell_or_range"])
    if not target_cells and finding.cell:
        for c in str(finding.cell).split(","):
            c = c.strip()
            if c:
                target_cells.append(c)
    if not target_cells and finding.snippet:
        hits = re.findall(r"\b[A-Z]{1,3}\d{1,7}\b", finding.snippet)
        target_cells.extend(hits[:3])

    seen_rows: set = set()
    seen_cells: set = set()
    for cell in target_cells[:5]:
        actual = _get_cell_text(ws, cell)
        if actual and cell not in seen_cells:
            parts.append(f"{cell}: {actual[:160]}")
            seen_cells.add(cell)
        m = re.match(r"^([A-Z]+)(\d+)$", cell)
        if not m:
            continue
        col_letters, row_num = m.group(1), int(m.group(2))
        for hdr_row in range(1, 4):
            key = ("hdr", hdr_row)
            if key in seen_rows:
                continue
            seen_rows.add(key)
            for c in range(1, min(ws.max_column + 1, 12)):
                v = ws.cell(row=hdr_row, column=c).value
                if v:
                    parts.append(f"{get_column_letter(c)}{hdr_row}: {str(v)[:80]}")
        for r_off in (-1, 1):
            r = row_num + r_off
            if r < 1 or r > (ws.max_row or 0):
                continue
            key = ("row", r)
            if key in seen_rows:
                continue
            seen_rows.add(key)
            for c in range(1, min(ws.max_column + 1, 10)):
                v = ws.cell(row=r, column=c).value
                if v:
                    parts.append(f"{get_column_letter(c)}{r}: {str(v)[:120]}")

    text = "\n".join(parts)
    if len(text) > max_chars:
        text = text[:max_chars] + "..."
    return text


async def _challenge_finding_with_llm(*, llm, finding, minimal_context: str) -> Optional[str]:
    """Run a 'challenge' LLM call to verify a high-severity finding.

    Returns "agree" / "disagree" / None on error.

    NOTE: ``max_tokens`` is set generously (512) on purpose — reasoning
    models often eat the first 64 tokens on CoT. If we cap too tight the
    response is empty and the challenger errors out, which on a P0 path
    silently fails the gate. We also parse the verdict leniently so a
    trailing "agree" / "成立" after a reasoning paragraph is still caught.
    """
    if not llm or not minimal_context:
        return None
    challenge_prompt = (
        "你是一名严格的审计质量复核专家，正在以质疑者的角度审阅以下复核发现。\n"
        "你的任务：判断该发现是否真实成立，或仅是表面/缺证据/逻辑不严。\n\n"
        f"【finding JSON】\n{finding.basis[:1000]}\n\n"
        f"【相关最小上下文（底稿原文片段）】\n{minimal_context[:1500]}\n\n"
        "请回答：agree（成立）/disagree（不成立/无依据）。最后一行只写一个词。"
    )
    try:
        answer = await _llm_chat(
            llm=llm,
            messages=[
                {"role": "system", "content": "你是审计复核的质疑者。"},
                {"role": "user", "content": challenge_prompt},
            ],
            stage="challenge",
            max_attempts=2,
            max_tokens=512,
        )
        text = (answer or "").strip()
        if not text:
            return None
        text_lower = text.lower()
        # Prefer the verdict on the last non-empty line so we ignore any
        # reasoning prose. Look for negative signals first to avoid matching
        # "不成立" against the more common "成立" substring.
        for line in reversed([ln.strip() for ln in text.splitlines() if ln.strip()]):
            ln = line.lower()
            if "disagree" in ln or "不同意" in ln or "不成立" in ln or "无法成立" in ln:
                return "disagree"
            if (
                line == "agree"
                or ln == "agree"
                or "同意" in ln
                or "成立" in ln
            ):
                return "agree"
        # Fall back to scanning the whole response.
        if "disagree" in text_lower or "不同意" in text or "不成立" in text:
            return "disagree"
        if "agree" in text_lower or "同意" in text or "成立" in text:
            return "agree"
        return None
    except Exception:
        return None
