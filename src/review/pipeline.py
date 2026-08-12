"""Pipeline orchestration: runs all review stages and merges findings.

Ported from analyze_excel.py's generate_report review core (no xlsx/txt rendering).
"""
import dataclasses
import json
import logging
import os
from typing import Callable, Dict, List, Optional, Tuple

import openpyxl

from review.attachments import _check_attachment_references, _verify_attachment_evidence_refs
from review.checkpoints import _llm_check_sheet_by_checkpoints
from review.evidence_steps import _llm_check_evidence_vs_steps
from review.evidence_agent import investigate_sheet
from review.excel_utils import _detect_layout, _normalize_sheet_id
from review.finding_taxonomy import (
    AssertionCatalog,
    apply_legacy_taxonomy,
    classify_finding,
    default_assertion_catalog,
)
from review.findings_review import _llm_review_findings
from review.hallucination import (
    _build_minimal_context,
    _challenge_finding_with_llm,
)
from review.llm import LLM_CALL_STATS
from review.models import Finding, _SEVERITY_DISPLAY
from review.procedure_pairs import (
    _check_procedure_pairs,
    _check_sheet_scope,
    _llm_check_procedure_pairs,
)
from review.quality_gates import (
    QualityGateContext,
    build_quality_gate_context,
    run_assertion_gates,
)

_logger = logging.getLogger("review.pipeline")

_SEVERITY_ORDER = {"P0": 0, "P1": 1, "P2": 2}

def _apply_static_finding_taxonomy(finding: dict) -> dict:
    """Compatibility wrapper for the former V1 taxonomy helper.

    The adapter only recognizes exact legacy producer labels.  Semantic
    classification is performed later by ``classify_finding`` using the
    versioned assertion catalog.
    """

    updated = apply_legacy_taxonomy(finding)
    finding.clear()
    finding.update(updated)
    return finding


def _deterministic_crosscheck_mode() -> str:
    mode = os.getenv("REVIEW_DETERMINISTIC_CROSSCHECK_MODE", "all_findings").strip().lower()
    if mode not in {"all_findings", "p0_only", "off"}:
        raise ValueError(
            "REVIEW_DETERMINISTIC_CROSSCHECK_MODE must be all_findings, p0_only or off"
        )
    return mode


def _gate(
    status: str,
    *,
    reason: str = "",
    issues: Optional[List[str]] = None,
    duration_ms: int | None = None,
) -> dict[str, object]:
    result: dict[str, object] = {"status": status}
    if reason:
        result["reason"] = reason
    if issues is not None:
        result["issues"] = list(issues)
    if duration_ms is not None:
        result["duration_ms"] = max(0, duration_ms)
    return result


def _assertion_gate_ids(finding: dict, catalog: AssertionCatalog) -> list[str]:
    assertion = catalog.maybe_assertion(
        str(finding.get("assertion_id", "") or "")
    ) or catalog.assertion("finding.unclassified")
    return list(assertion.deterministic_gate_ids)


def _model_re_review_gate(
    finding: dict, review_result: dict | None
) -> dict[str, object]:
    """Record the independent model re-review without title-based inference."""

    if str(finding.get("origin", "") or "").strip() == "llm":
        return _gate(
            "not_run",
            reason="LLM-origin finding has no separate model verifier",
        )
    if not isinstance(review_result, dict):
        return _gate(
            "not_run", reason="model re-review produced no result for this finding"
        )
    reviewed_status = str(review_result.get("llm_status", "") or "")
    finding_status = str(finding.get("status", "") or "")
    return _gate(
        "passed" if reviewed_status == finding_status else "flagged",
        reason=(
            "re-review agrees with the rule result"
            if reviewed_status == finding_status
            else "re-review status differs from the rule result"
        ),
    )


def _emit_progress(on_progress, stage: str, current_sheet: str, findings, msg: str) -> None:
    """Best-effort progress report. Never raises — pipeline must not break on a bad callback."""
    if on_progress is None:
        return
    try:
        sev = {"P0": 0, "P1": 0, "P2": 0}
        for f in findings:
            s = getattr(f, "severity", None) or "P2"
            if s not in sev:
                s = "P2"
            sev[s] += 1
        on_progress({
            "stage": stage,
            "current_sheet": current_sheet or "",
            "llm_calls": {k: int(v.get("calls", 0)) for k, v in LLM_CALL_STATS.items()},
            "findings_so_far": {
                "P0": sev["P0"], "P1": sev["P1"], "P2": sev["P2"],
                "total": len(findings),
            },
            "msg": msg,
        })
    except Exception:
        pass


def _parse_sheet_filter(raw: Optional[str]) -> Optional[List[str]]:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or text.lower() in {"all", "*", "全部"}:
        return None
    parts = [p.strip() for p in text.replace("，", ",").split(",") if p.strip()]
    seen, out = set(), []
    for p in parts:
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out or None


def _finding_to_dict(
    f: Finding, assertion_catalog: AssertionCatalog | None = None
) -> dict:
    d = dataclasses.asdict(f)
    for k, default in (("evidence_refs", []), ("reasons", []), ("fix_suggestion_detail", {})):
        v = d.get(k)
        if isinstance(v, str) and v.strip():
            try:
                d[k] = json.loads(v)
            except Exception:
                d[k] = v
        else:
            d[k] = default
    d["severity_display"] = _SEVERITY_DISPLAY.get(f.severity, f.severity)
    return classify_finding(d, assertion_catalog or default_assertion_catalog())


def _path_key(value: object) -> str:
    return str(value or "").strip().replace("\\", "/").casefold()


def _accepted_agent_evidence_for_sheet(attachments, sheet: object) -> List[dict]:
    """Return only Agent evidence that already passed path/excerpt validation."""
    if not attachments or not isinstance(attachments, dict):
        return []
    by_sheet = attachments.get("agent_evidence_by_sheet") or {}
    if not isinstance(by_sheet, dict):
        return []
    records = by_sheet.get(_normalize_sheet_id(str(sheet or "")))
    if not isinstance(records, list):
        return []
    result: List[dict] = []
    seen = set()
    for raw in records:
        if not isinstance(raw, dict):
            continue
        path = str(raw.get("path", "") or "").strip()
        excerpt = str(raw.get("excerpt", "") or "").strip()
        key = _path_key(path)
        if not key or not excerpt or key in seen:
            continue
        seen.add(key)
        result.append({"path": path, "excerpt": excerpt})
    return result

def _backfill_embedded_evidence_refs(findings_dicts, attachments=None):
    """Backfill only a path plus verbatim Agent excerpt that can be verified.

    A document title, fuzzy topic match, cached OCR text, or an LLM-supplied
    citation cannot identify a particular source. The finding's basis must
    explicitly name the path and the evidence Agent must have supplied a
    same-sheet excerpt that validates against the pinned attachment index.
    """
    if not findings_dicts or not attachments:
        return findings_dicts
    ocr_cache = (attachments or {}).get("ocr_by_path") or {}

    def _ocr_full_text(rel_path: str) -> str:
        """Return the full OCR text for an embedded-media path, or empty string."""
        candidates: List[str] = []
        key = rel_path.lower()
        cached = ocr_cache.get(key)
        if isinstance(cached, dict) and str(cached.get("status", "")).lower() == "ok":
            candidates.append(str(cached.get("content", "") or ""))
        if "::" in key:
            candidates.append(ocr_cache.get(key.replace("::", "__")) or "")
        for raw in candidates:
            if raw:
                return raw
        return ""

    def _attachment_extracted_text(rel_path: str) -> str:
        """Return the directly-extracted text for a real attachment from the index."""
        if not attachments:
            return ""
        items = attachments.get("items") or []
        target = str(rel_path or "").strip().lower().replace("\\", "/")
        for item in items:
            if not item:
                continue
            rel = str(getattr(item, "rel_path", "") or "").lower().replace("\\", "/")
            if rel == target:
                text = str(getattr(item, "extracted_text", "") or "").strip()
                if text:
                    return text
        return ""

    for fnd in findings_dicts:
        if not isinstance(fnd, dict):
            continue
        basis = str(fnd.get("basis") or "")
        if not basis:
            continue
        accepted = _accepted_agent_evidence_for_sheet(attachments, fnd.get("sheet"))
        if not accepted:
            continue
        refs_raw = fnd.get("evidence_refs") or []
        refs = [dict(ref) for ref in refs_raw if isinstance(ref, dict)]
        refs_by_path = {
            _path_key(ref.get("attachment")): ref
            for ref in refs
            if _path_key(ref.get("attachment"))
        }
        changed = False
        folded_basis = basis.casefold()
        for evidence in accepted:
            path = evidence["path"]
            key = _path_key(path)
            if not key or path.casefold() not in folded_basis:
                continue
            candidate = {
                "sheet": fnd.get("sheet"),
                "cell_or_range": "",
                "attachment": path,
                "excerpt": evidence["excerpt"],
            }
            verified = _verify_attachment_evidence_refs([candidate], attachments)
            if not verified or not verified[0].get("attachment"):
                continue
            normalized = dict(verified[0])
            if path.startswith(".embedded_media/"):
                full_text = _ocr_full_text(path)
                if full_text:
                    normalized["full_text"] = full_text
            else:
                attachment_text = _attachment_extracted_text(path)
                if attachment_text:
                    normalized["attachment_text"] = attachment_text
            if key in refs_by_path:
                existing = refs_by_path[key]
                if not str(existing.get("excerpt", "") or "").strip():
                    existing.update(normalized)
                    changed = True
                continue
            refs.append(normalized)
            refs_by_path[key] = normalized
            changed = True
        if changed:
            fnd["evidence_refs"] = refs
    return findings_dicts


def _sort_findings(findings: List[Finding]) -> List[Finding]:
    return sorted(findings, key=lambda f: (
        _SEVERITY_ORDER.get(f.severity, 9),
        f.sheet or "",
        str(f.cell or ""),
    ))


async def run_review(
    *,
    wb: openpyxl.Workbook,
    checkpoints: Optional[Dict[str, List[str]]] = None,
    attachments: Optional[Dict[str, object]] = None,
    sheets: Optional[str] = None,
    llm,
    attachments_preview: Optional[Dict[str, object]] = None,
    on_progress: Optional[Callable[[dict], None]] = None,
    assertion_catalog: AssertionCatalog | None = None,
    quality_gate_context: QualityGateContext | None = None,
) -> Tuple[List[dict], dict]:
    """Run the full review pipeline. Returns (findings_dicts, stats)."""
    catalog = assertion_catalog or default_assertion_catalog()
    gate_context = quality_gate_context or build_quality_gate_context(
        workbook=wb,
        evidence_registry=None,
        assertion_catalog=catalog,
    )
    checkpoints = checkpoints or {}
    attachments = attachments if attachments is not None else attachments_preview
    attachments = attachments or {}
    filtered = _parse_sheet_filter(sheets)
    warning = ""
    if filtered is None:
        target = list(wb.sheetnames)
    else:
        # Resolve requested names to actual workbook tabs, tolerating the
        # case/dash/space variants an LLM naturally produces (e.g. "pe6" -> "PE-6").
        # _normalize_sheet_id is the same helper attachments/evidence_steps use
        # for sheet-id matching.
        norm_to_actual = {_normalize_sheet_id(s): s for s in wb.sheetnames}
        resolved: List[str] = []
        unmatched: List[str] = []
        seen = set()
        for req in filtered:
            actual = norm_to_actual.get(_normalize_sheet_id(req))
            if actual is None:
                unmatched.append(req)
                continue
            if actual in seen:
                continue
            seen.add(actual)
            resolved.append(actual)
        reviewable = [
            s for s in resolved
            if _detect_layout(wb[s])[0] is not None or checkpoints.get(s)
        ]
        if not reviewable:
            _logger.warning(
                "sheets=%r yielded no reviewable sheets (resolved=%r unmatched=%r); "
                "falling back to all sheets", sheets, resolved, unmatched,
            )
            target = list(wb.sheetnames)
            detail = "无可审阅内容（无审计程序布局/检查要点）" if resolved else "未在底稿中找到"
            warning = f"指定的 Sheet（{sheets}）{detail}，已回退到全部 Sheet。"
        else:
            target = resolved
            if unmatched:
                warning = f"部分指定 Sheet 未匹配：{', '.join(unmatched)}；已审阅：{', '.join(resolved)}。"
    _logger.info(
        "run_review start: sheets_arg=%r target=%r wb_sheets=%r "
        "checkpoints_keys=%r attachment_items=%r warning=%r",
        sheets, target, list(wb.sheetnames),
        list(checkpoints.keys()), len(attachments.get("items", []) if attachments else []),
        warning,
    )

    findings: List[Finding] = []
    agent_stats = {
        "mode": str(os.environ.get("REVIEW_EVIDENCE_AGENT_MODE", "fallback")),
        "runs": 0,
        "tool_calls": 0,
        "accepted_evidence": 0,
        "unresolved": 0,
        "errors": 0,
        "ocr": {"calls": 0, "success": 0, "errors": 0, "timeouts": 0},
        "details": [],
    }
    _emit_progress(on_progress, "starting", "", findings, f"开始审阅，共 {len(target)} 个 sheet")
    for sheet in target:
        if sheet not in wb.sheetnames:
            _logger.info("  sheet=%r skipped (not in workbook)", sheet)
            continue
        ws = wb[sheet]
        _emit_progress(on_progress, "checkpoints", sheet, findings, f"开始处理 {sheet}")
        _logger.info(
            "  sheet=%r cp=%r attachments=%r",
            sheet, bool(checkpoints.get(sheet)), bool(attachments),
        )
        agent_result = await investigate_sheet(ws=ws, attachments=attachments, llm=llm)
        if agent_result.get("status") != "skipped":
            agent_stats["runs"] += 1
            agent_stats["tool_calls"] += int(agent_result.get("tool_calls", 0) or 0)
            agent_stats["accepted_evidence"] += len(agent_result.get("evidence", []) or [])
            agent_stats["unresolved"] += len(agent_result.get("unresolved", []) or [])
            if agent_result.get("status") == "error":
                agent_stats["errors"] += 1
            ocr_result = agent_result.get("ocr") or {}
            if isinstance(ocr_result, dict):
                ocr_stats = agent_stats["ocr"]
                for key in ("calls", "success", "errors", "timeouts"):
                    ocr_stats[key] += int(ocr_result.get(key, 0) or 0)
            evidence = agent_result.get("evidence", []) or []
            agent_stats["details"].append({
                "sheet": sheet,
                "status": agent_result.get("status"),
                "tool_calls": int(agent_result.get("tool_calls", 0) or 0),
                "evidence": list(evidence),
                "unresolved": list(agent_result.get("unresolved", []) or []),
                "tool_trace": list(agent_result.get("tool_trace", []) or []),
                "ocr": dict(ocr_result) if isinstance(ocr_result, dict) else {},
            })
            if evidence:
                by_sheet = attachments.setdefault("agent_evidence_by_sheet", {})
                if isinstance(by_sheet, dict):
                    by_sheet[_normalize_sheet_id(sheet)] = list(evidence)
        # 1) checkpoint-based review
        if checkpoints.get(sheet):
            findings += await _llm_check_sheet_by_checkpoints(
                llm=llm, ws_title=sheet, ws=ws,
                checkpoints=checkpoints[sheet],
                attachments=attachments or None,
                on_progress=lambda stage, msg: _emit_progress(
                    on_progress, "checkpoints", sheet, findings, msg,
                ),
            )
        _emit_progress(on_progress, "checkpoints", sheet, findings, f"完成 {sheet} checkpoint 评审")
        # 2) attachment-reference matching
        if attachments:
            findings += _check_attachment_references(sheet, ws, attachments)
        # 3) evidence <-> step consistency
        if attachments:
            findings += await _llm_check_evidence_vs_steps(
                llm=llm, ws_title=sheet, ws=ws,
                attachments=attachments,
            )
        _emit_progress(on_progress, "evidence_steps", sheet, findings, f"完成 {sheet} 证据-步骤一致性检查")
        # 4) rule-based procedure-pair checks
        findings += _check_procedure_pairs(sheet, ws)
        # 5) sheet-scope checks
        findings += _check_sheet_scope(sheet, ws)
        # 6) A-C correspondence LLM judgement
        _, ac_findings = await _llm_check_procedure_pairs(
            llm=llm, wb=wb, target_sheets=[sheet],
            attachments=attachments,
        )
        findings += ac_findings
        _emit_progress(on_progress, "procedure_pairs", sheet, findings, f"完成 {sheet} 程序配对检查")

    findings_sorted = _sort_findings(findings)

    # Lift attachment paths into evidence_refs[].attachment / full_text /
    # attachment_text BEFORE the reviewer runs so it can see the actual OCR
    # content for each finding rather than guessing from cell text alone.
    pre_review_dicts: List[dict] = []
    for fnd in findings_sorted:
        d = _finding_to_dict(fnd, catalog)
        pre_review_dicts.append(d)
    _backfill_embedded_evidence_refs(pre_review_dicts, attachments)
    # Mirror the backfilled refs onto the Finding objects so the reviewer (which
    # consumes Finding instances, not dicts) sees the same content.
    for fnd, d in zip(findings_sorted, pre_review_dicts):
        refs = d.get("evidence_refs") or []
        if refs:
            object.__setattr__(
                fnd, "evidence_refs", json.dumps(refs, ensure_ascii=False)
            )

    # LLM re-review of rule-based (non-LLM-tagged) findings
    _emit_progress(on_progress, "findings_review", "", findings, "进入发现复核")
    review = await _llm_review_findings(wb, findings_sorted, llm, attachments=attachments)

    # Assertion-selected deterministic gates + adversarial challenge. The
    # catalog determines gate applicability; direct pipeline callers without
    # a frozen quality context receive explicit not_run records instead of an
    # implicit attachment re-scan or false positive pass.
    deterministic_mode = _deterministic_crosscheck_mode()
    cross_issues: Dict[int, List[str]] = {}
    challenge: Dict[int, Optional[str]] = {}
    deterministic_gates: Dict[int, dict[str, dict[str, object]]] = {}
    challenge_gates: Dict[int, dict[str, object]] = {}
    _emit_progress(on_progress, "hallucination", "", findings_sorted, "进入交叉验证/对抗挑战")
    for idx, f in enumerate(findings_sorted, start=1):
        gate_finding = _finding_to_dict(f, catalog)
        gate_ids = _assertion_gate_ids(gate_finding, catalog)
        if deterministic_mode == "off":
            deterministic_gates[idx] = {
                gate_id: _gate(
                    "not_run",
                    reason="deterministic gates disabled by configuration",
                    duration_ms=0,
                )
                for gate_id in gate_ids
            }
        elif deterministic_mode == "p0_only" and f.severity != "P0" and not f.needs_review:
            deterministic_gates[idx] = {
                gate_id: _gate(
                    "not_run",
                    reason="p0_only policy excludes non-P0 finding",
                    duration_ms=0,
                )
                for gate_id in gate_ids
            }
        else:
            deterministic_gates[idx] = run_assertion_gates(
                gate_finding, gate_context
            )
        cross_issues[idx] = [
            issue
            for outcome in deterministic_gates[idx].values()
            if isinstance(outcome, dict)
            for issue in outcome.get("issues", [])
            if isinstance(issue, str)
        ]

        if f.severity == "P0":
            ws = wb[f.sheet] if f.sheet in wb.sheetnames else None
            ctx = _build_minimal_context(f, ws)
            if not ctx:
                challenge_gates[idx] = _gate(
                    "not_run", reason="minimal context unavailable for challenger"
                )
            else:
                try:
                    challenge[idx] = await _challenge_finding_with_llm(
                        llm=llm, finding=f, minimal_context=ctx,
                    )
                    verdict = challenge[idx]
                    if verdict == "agree":
                        challenge_gates[idx] = _gate("passed", reason="challenger agreed")
                    elif verdict == "disagree":
                        challenge_gates[idx] = _gate("flagged", reason="challenger disagreed")
                    else:
                        challenge_gates[idx] = _gate(
                            "error", reason="challenger returned no verdict"
                        )
                except Exception as exc:
                    challenge_gates[idx] = _gate(
                        "error", reason=f"{type(exc).__name__}: {exc}"
                    )
        else:
            challenge_gates[idx] = _gate(
                "not_run", reason="adversarial challenger is limited to P0 findings"
            )

    out: List[dict] = []
    for idx, f in enumerate(findings_sorted, start=1):
        d = _finding_to_dict(f, catalog)
        if idx in review:
            d.update(review[idx])
        d = classify_finding(d, catalog)
        d["cross_validate_issues"] = cross_issues.get(idx, [])
        d["challenge_verdict"] = challenge.get(idx)
        model_gate = _model_re_review_gate(d, review.get(idx))
        d["quality_gates"] = {
            **deterministic_gates.get(idx, {}),
            "model_re_review": model_gate,
            "adversarial_challenge": challenge_gates.get(
                idx, _gate("not_run", reason="no adversarial gate record")
            ),
        }
        out.append(d)
    # Lift only already-validated Agent attachment excerpts whose exact path
    # appears in the V1 basis. This makes the structured UI citation point to
    # the evidence actually inspected without guessing from a document title.
    _backfill_embedded_evidence_refs(out, attachments)
    # Backfilled validated attachment paths are part of the claim subject for
    # assertions that require attachment evidence, so classify one final time.
    out = [classify_finding(finding, catalog) for finding in out]

    _emit_progress(on_progress, "done", "", findings_sorted, "审阅完成")
    stats = {
        "total_findings": len(out),
        "by_severity": _counts_by(out, "severity"),
        "by_status": _counts_by(out, "status"),
        "by_risk_type": _counts_by(out, "risk_type"),
        "llm_call_stats": {k: dict(v) for k, v in LLM_CALL_STATS.items()},
        "evidence_agent": agent_stats,
        "warning": warning,
        "quality_gates": _gate_counts(out),
    }
    return out, stats


def _counts_by(items: List[dict], key: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for it in items:
        k = str(it.get(key, "") or "")
        counts[k] = counts.get(k, 0) + 1
    return counts


def _gate_counts(items: List[dict]) -> Dict[str, Dict[str, int]]:
    counts: Dict[str, Dict[str, int]] = {}
    for finding in items:
        gates = finding.get("quality_gates") or {}
        if not isinstance(gates, dict):
            continue
        for name, outcome in gates.items():
            if not isinstance(outcome, dict):
                continue
            status = str(outcome.get("status", "") or "")
            if not status:
                continue
            counts.setdefault(str(name), {})[status] = (
                counts.setdefault(str(name), {}).get(status, 0) + 1
            )
    return counts
